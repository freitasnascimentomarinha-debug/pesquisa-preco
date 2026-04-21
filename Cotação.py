import base64
import os
import streamlit as st
import requests
import pandas as pd
import streamlit.components.v1 as components
from datetime import datetime
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import re

JUSTIFICATIVA_COTACAO = (
    "Justificativa quanto aos critérios de Pesquisa de Preços (IN 65/2021)\n\n"
    "A presente pesquisa de preços foi realizada em estrita observância aos critérios e parâmetros "
    "estabelecidos pela Instrução Normativa SEGES/ME nº 65, de 7 de julho de 2021. Foram priorizados "
    "dados provenientes de aquisições e contratações similares no âmbito da Administração Pública, "
    "assegurando a contemporaneidade dos registros e a compatibilidade técnica com o objeto pretendido.\n\n"
    "Os valores obtidos refletem os preços efetivamente praticados no mercado, garantindo a "
    "economicidade e a ampla fundamentação na formação do preço estimado para a contratação. "
    "Ressalta-se que a utilização de painéis de preços e bancos de dados oficiais confere "
    "transparência, rastreabilidade e segurança documental ao procedimento de aferição do valor de mercado.\n\n"
    "Dessa forma, entende-se que a metodologia adotada atende aos princípios da razoabilidade, "
    "economicidade e motivação do ato administrativo, conferindo robustez à formação do preço estimado."
)

def _pdf_safe(text):
    """Garante texto seguro para renderização PDF"""
    if not text or str(text) == 'nan' or str(text) == 'None':
        return ''
    try:
        return str(text).encode('latin-1', 'replace').decode('latin-1')
    except Exception:
        return str(text)

# Configuração da página
st.set_page_config(
    page_title="AtaCotada - Cotação",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para tema azul escuro com dourado
st.markdown("""
    <style>
        :root {
            --azul-escuro: #001a4d;
            --azul-claro: #0033cc;
            --dourado: #d4af37;
            --branco: #ffffff;
            --cinza-claro: #f0f0f0;
        }
        
        * {
            margin: 0;
            padding: 0;
        }
        
        body, .main {
            background-color: #001a4d;
            color: #ffffff;
        }
        
        .stApp {
            background-color: #001a4d;
        }
        
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
        }
        
        /* Header customizado */
        .header-container {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            padding: 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
            text-align: center;
        }
        
        .logo-text {
            color: #ffffff;
            font-size: 14px;
            font-weight: 600;
            letter-spacing: 2px;
            margin-bottom: 0.5rem;
        }
        
        .sistema-nome {
            color: #d4af37;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 3px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);
            margin: 0.5rem 0;
            font-family: 'Arial Black', sans-serif;
        }
        
        .subtitulo {
            color: #ffffff;
            font-size: 14px;
            margin-top: 0.5rem;
            letter-spacing: 1px;
        }
        
        /* Filtros container */
        .filtros-container {
            background: linear-gradient(135deg, #0a2540 0%, #164863 100%);
            padding: 1.5rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            border-left: 5px solid #d4af37;
        }
        
        /* Cards de estatísticas */
        .stats-card {
            background: linear-gradient(135deg, #0a2540 0%, #0f4c75 100%);
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 5px solid #d4af37;
            margin-bottom: 1rem;
        }
        
        /* Botões */
        .stButton > button {
            background-color: #d4af37;
            color: #001a4d;
            border: none;
            border-radius: 6px;
            padding: 0.75rem 1.5rem;
            font-weight: bold;
            font-size: 16px;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .stButton > button:hover {
            background-color: #ffd700;
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(212, 175, 55, 0.3);
        }
        
        /* Inputs e Selects */
        .stTextInput > div > div > input,
        .stMultiSelect > div > div > div,
        .stSelectbox > div > div > select,
        .stSlider > div > div > div {
            background-color: #0a2540 !important;
            color: #ffffff !important;
            border: 2px solid #0033cc !important;
            border-radius: 6px !important;
        }
        
        /* Labels e textos de input - BRANCO */
        label, .stText, .stLabel, .stMultiSelect label, .stSelectbox label, .stTextInput label {
            color: #ffffff !important;
        }
        
        /* Checkbox labels e texto - MAIS ESPECÍFICO */
        .stCheckbox {
            background-color: transparent;
        }
        
        .stCheckbox > label {
            color: #ffffff !important;
            background-color: transparent;
        }
        
        .stCheckbox > div > label {
            color: #ffffff !important;
        }
        
        .stCheckbox label {
            color: #ffffff !important;
        }
        
        .stCheckbox label span {
            color: #ffffff !important;
        }
        
        .stCheckbox label div {
            color: #ffffff !important;
        }
        
        .stCheckbox input[type="checkbox"] + label {
            color: #ffffff !important;
        }
        
        /* Pegar todo texto dentro de checkbox */
        .stCheckbox * {
            color: #ffffff !important;
        }
        
        /* Labels dos widgets */
        div[role="radiogroup"] label,
        .stMultiSelect > div > div[class*="stMultiSelect"] label {
            color: #ffffff !important;
        }
        
        /* Todos os span e div dentro dos filtros */
        .stMultiSelect > div > div > span,
        .stSelectbox > div > div > span {
            color: #ffffff !important;
        }
        
        /* Títulos */
        h1, h2, h3 {
            color: #d4af37;
            text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.5);
        }
        
        /* Expanders */
        .streamlit-expanderHeader {
            background-color: #0a2540;
            color: #d4af37;
            border: 2px solid #0033cc;
            border-radius: 6px;
        }
        
        /* Info boxes */
        .stAlert {
            background-color: #0a2540;
            border: 2px solid #d4af37;
            border-radius: 6px;
        }
        
        /* Tabelas */
        .streamlit-dataframe {
            background-color: #0a2540;
            color: #ffffff;
        }
        
        table {
            color: #ffffff;
            background-color: #0a2540;
        }
        
        th {
            background-color: #0033cc !important;
            color: #d4af37 !important;
            font-weight: bold;
        }
        
        td {
            background-color: #0f4c75 !important;
            color: #ffffff;
        }
        
        tr:hover {
            background-color: #164863 !important;
        }
        
        /* Markdown */
        .markdown-text-container {
            color: #ffffff;
        }
        
        /* SIDEBAR - NAVEGAÇÃO DE PÁGINAS MODERNO */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0a0a0a 0%, #1a1a1a 100%);
            border-right: 2px solid #d4af37;
        }
        
        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            background: transparent;
        }
        
        /* Botões de navegação das páginas */
        [data-testid="stSidebar"] button {
            background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%);
            color: #ffffff !important;
            border: 2px solid #333333 !important;
            border-radius: 8px;
            margin: 0.5rem 0;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.3s ease;
            padding: 0.75rem 1rem !important;
        }
        
        [data-testid="stSidebar"] button:hover {
            background: linear-gradient(135deg, #2d2d2d 0%, #3d3d3d 100%);
            border: 2px solid #d4af37 !important;
            color: #d4af37 !important;
            transform: translateX(5px);
            box-shadow: 0 4px 12px rgba(212, 175, 55, 0.2);
        }
        
        /* Botão ativo */
        [data-testid="stSidebar"] button[kind="primary"] {
            background: linear-gradient(135deg, #d4af37 0%, #ffd700 100%);
            color: #001a4d !important;
            border: 2px solid #d4af37 !important;
            font-weight: bold;
            box-shadow: 0 4px 12px rgba(212, 175, 55, 0.4);
        }
        
        /* Links na sidebar */
        [data-testid="stSidebar"] a {
            color: #ffffff !important;
            text-decoration: none;
            transition: all 0.3s ease;
        }
        
        [data-testid="stSidebarNav"] span {
            color: #ffffff !important;
        }
        
        [data-testid="stSidebar"] a:hover,
        [data-testid="stSidebarNav"] a:hover span {
            color: #d4af37 !important;
            text-shadow: 0 0 10px rgba(212, 175, 55, 0.5);
        }
        
        /* Texto e labels na sidebar */
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stText,
        [data-testid="stSidebar"] div {
            color: #ffffff !important;
        }
        
        /* Separadores na sidebar */
        [data-testid="stSidebar"] hr {
            border-color: #333333;
            margin: 1rem 0;
        }
        
        /* Cabeçalho da sidebar */
        [data-testid="stSidebar"] .stMarkdown h1,
        [data-testid="stSidebar"] .stMarkdown h2,
        [data-testid="stSidebar"] .stMarkdown h3 {
            color: #d4af37;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.8);
            border-bottom: 2px solid #d4af37;
            padding-bottom: 0.5rem;
            margin-bottom: 1rem;
        }
    </style>
""", unsafe_allow_html=True)

# ===== SIDEBAR - Navegação customizada =====
_acanto_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Projeto Adesões", "acanto.png")
if os.path.exists(_acanto_path):
    with open(_acanto_path, "rb") as _f:
        _acanto_b64 = base64.b64encode(_f.read()).decode()
else:
    _acanto_b64 = None

with st.sidebar:
    if _acanto_b64:
        st.markdown(f'<div style="text-align:center;padding:1rem 0 0.5rem 0;"><img src="data:image/png;base64,{_acanto_b64}" style="max-width:70%;height:auto;"></div>', unsafe_allow_html=True)
    st.markdown("## MENU")
    st.markdown("---")
    st.page_link("Cotação.py", label="Cotação", icon="⚓")
    st.page_link("pages/Detalhes_Compra.py", label="Detalhes Compra", icon="🔍")
    st.page_link("pages/Adesões.py", label="Adesões", icon="🤝")
    st.page_link("pages/Notas_Fiscais.py", label="Notas Fiscais", icon="📄")
    st.page_link("pages/Banco_de_Fornecedores.py", label="Fornecedores", icon="🏢")
    st.page_link("pages/Consulta.py", label="Consulta CNPJ", icon="💻")
    st.page_link("pages/Web_Scraping.py", label="Web Scraping", icon="🕷️")
    st.page_link("pages/O_Babilaca_(IA).py", label="O Babilaca (IA)", icon="🧠")
    st.markdown("---")
    st.markdown("## LINKS ÚTEIS")
    st.markdown("""
    <div style="margin-bottom: 1rem;">
        <a href="https://freitasnascimentomarinha-debug.github.io/ShootMail/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            📧 Disparador de Emails
        </a>
    </div>
    <div style="margin-bottom: 1rem;">
        <a href="https://detetive-obtencao.vercel.app/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            🚨 Detetive Obtenção
        </a>
    </div>
    """, unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;color:#d4af37;font-size:10px;font-weight:600;padding:0.3rem 0;white-space:nowrap;">Centro de Operações do Abastecimento</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-footer">Marinha do Brasil<br>AtaCotada v1.0</div>', unsafe_allow_html=True)

# Função para formatar preço em reais
def formatar_preco_reais(valor):
    if valor is None:
        return 'Preço não disponível'
    else:
        return f'R$ {valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

# Função para formatar número em formato brasileiro
def formatar_numero_br(valor):
    """Formata número para padrão brasileiro (vírgula como separador decimal)"""
    return f'{valor:.2f}'.replace('.', ',')

def formatar_moeda_br(valor):
    """Formata valor monetário para padrão brasileiro com separador de milhares (1.234.567,89)"""
    if valor is None or (isinstance(valor, float) and valor != valor):  # NaN check
        return ""
    return f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

# Função para encontrar coluna com variações de nome
def encontrar_coluna(df, nomes_possiveis):
    """Procura por uma coluna entre várias variações de nome (case-insensitive)"""
    colunas_lower = {col.lower(): col for col in df.columns}
    # Preferência por correspondência exata (case-insensitive)
    for nome in nomes_possiveis:
        chave = nome.lower()
        if chave in colunas_lower:
            return colunas_lower[chave]

    # Se não houver correspondência exata, procurar por substring nos nomes das colunas
    for col_lower, col_original in colunas_lower.items():
        for nome in nomes_possiveis:
            if nome.lower() in col_lower:
                return col_original

    return None

def organizar_colunas_resultado(df):
    """Reordena as colunas para priorizar os campos mais úteis na análise."""
    colunas_prioritarias = [
        (["idCompra", "id_compra", "compraId"], "ID Compra"),
        (["dataCompra", "data_compra"], "Data Compra"),
        (["descricaoItem", "descricao", "description", "item_descricao"], "Descrição do Item"),
        (["siglaUnidadeFornecimento", "unidadeFornecimento", "unidade_fornecimento"], "Sigla Unidade Fornecimento"),
        (["precoUnitario", "valorUnitario", "preco_unitario"], "Preço Unitário"),
        (["quantidade", "quantidadeItem", "qtd"], "Quantidade"),
        (["nomeUasg", "orgaoEntidade.razaoSocial", "orgao_nome"], "Nome UASG"),
        (["nomeFornecedor", "razaoSocialFornecedor", "fornecedor_nome"], "Nome Fornecedor"),
        (["niFornecedor", "cnpjFornecedor", "fornecedor_cnpj"], "NI Fornecedor"),
        (["marca", "nomeMarca"], "Marca"),
        (["objetoCompra", "objeto", "objetoCompraDescricao"], "Objeto Compra"),
    ]

    colunas_ordenadas = []
    renomear = {}
    for aliases, label in colunas_prioritarias:
        coluna = encontrar_coluna(df, aliases)
        if coluna and coluna not in colunas_ordenadas:
            colunas_ordenadas.append(coluna)
            renomear[coluna] = label

    colunas_restantes = [col for col in df.columns if col not in colunas_ordenadas]
    df_ordenado = df[colunas_ordenadas + colunas_restantes].copy()
    return df_ordenado.rename(columns=renomear)

def preparar_dataframe_exibicao(df):
    """Prepara a grade de resultados com colunas ordenadas e formatações de leitura."""
    df_exibicao = organizar_colunas_resultado(df)

    if "Preço Unitário" in df_exibicao.columns:
        df_exibicao["Preço Unitário"] = df_exibicao["Preço Unitário"].apply(formatar_preco_reais)

    if "Data Compra" in df_exibicao.columns:
        datas = pd.to_datetime(df_exibicao["Data Compra"], errors='coerce')
        df_exibicao["Data Compra"] = datas.dt.strftime('%d/%m/%Y').where(
            datas.notna(),
            df_exibicao["Data Compra"].fillna('').astype(str)
        )

    return df_exibicao

def exibir_tabela_com_destaque(df_editor):
    """Exibe uma visualização com destaque para linhas selecionadas."""
    if 'Selecionar' not in df_editor.columns:
        st.dataframe(df_editor, hide_index=True, use_container_width=True)
        return

    df_visual = df_editor.drop(columns=['Selecionar']).copy()
    mask = df_editor['Selecionar'].fillna(False)

    def estilo_linha(row):
        if bool(mask.loc[row.name]):
            return ['background-color: #fff5cc; color: #1a1a1a; font-weight: 600'] * len(row)
        return [''] * len(row)

    st.dataframe(
        df_visual.style.apply(estilo_linha, axis=1),
        hide_index=True,
        use_container_width=True
    )

# Função para remover outliers usando o método IQR (Interquartile Range)
def remover_outliers_iqr(dataframe, coluna):
    """
    Remove outliers usando o método IQR.
    Outliers são valores fora da faixa [Q1 - 1.5*IQR, Q3 + 1.5*IQR]
    Retorna: (df_sem_outliers, quantidade_removida, limite_inferior, limite_superior)
    """
    # Garantir que a coluna exista
    if coluna not in dataframe.columns:
        return dataframe.copy(), 0, None, None

    # Calcular quartis e IQR ignorando valores nulos
    serie = dataframe[coluna].dropna()
    if serie.empty:
        return dataframe.copy(), 0, None, None

    Q1 = serie.quantile(0.25)
    Q3 = serie.quantile(0.75)
    IQR = Q3 - Q1
    limite_inferior = Q1 - 1.5 * IQR
    limite_superior = Q3 + 1.5 * IQR

    df_sem_outliers = dataframe[(dataframe[coluna] >= limite_inferior) & (dataframe[coluna] <= limite_superior)]
    outliers_removidos = len(dataframe) - len(df_sem_outliers)

    return df_sem_outliers, outliers_removidos, limite_inferior, limite_superior

# Função para gerar relatório em Excel formatado
def gerar_relatorio_excel(dataframe, estatisticas, outliers_info, col_precounitario):
    """
    Gera um relatório em Excel com os dados formatados (Layout Notas Fiscais).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Cotação"
    
    max_col = 12
    last_col = 'L'
    
    # Estilos (Padrão Notas Fiscais)
    navy = PatternFill(start_color='001A4D', end_color='001A4D', fill_type='solid')
    dark_navy = PatternFill(start_color='0A2540', end_color='0A2540', fill_type='solid')
    light_bg = PatternFill(start_color='F5F8FF', end_color='F5F8FF', fill_type='solid')
    alt_row = PatternFill(start_color='EDF2FA', end_color='EDF2FA', fill_type='solid')
    white_bg = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    gold_title = Font(name='Calibri', bold=True, color='D4AF37', size=20)
    white_subtitle = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    gray_info = Font(name='Calibri', color='B0B0B0', size=9, italic=True)
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    gold_stat = Font(name='Calibri', bold=True, color='D4AF37', size=10)
    data_font = Font(name='Calibri', size=10, color='333333')
    
    border_thin = Border(
        left=Side(style='thin', color='E8E8E8'),
        right=Side(style='thin', color='E8E8E8'),
        top=Side(style='thin', color='E8E8E8'),
        bottom=Side(style='thin', color='E8E8E8')
    )
    gold_bottom = Border(bottom=Side(style='medium', color='D4AF37'))
    
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # CABEÇALHO SUPERIOR
    item_str = "Pesquisa de Preços - Cotação"
    if not dataframe.empty:
        col_codigo = encontrar_coluna(dataframe, ['codigoItemCatalogo', 'codigo_item', 'codigo'])
        col_desc = encontrar_coluna(dataframe, ['descricaoItem', 'descricao', 'description'])
        
        codigo = str(dataframe.iloc[0][col_codigo]) if col_codigo else ""
        desc = str(dataframe.iloc[0][col_desc]) if col_desc else ""
        
        if codigo and desc:
            item_str = f"{codigo} - {desc}"
        elif desc:
            item_str = desc

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = "AtaCotada"
    ws['A1'].font = gold_title
    ws['A1'].fill = navy
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = item_str
    ws['A2'].font = white_subtitle
    ws['A2'].fill = navy
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 26
    
    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = "Marinha do Brasil - Centro de Operações do Abastecimento"
    ws['A3'].font = gray_info
    ws['A3'].fill = navy
    ws['A3'].alignment = center
    ws.row_dimensions[3].height = 18
    
    for row_h in range(1, 4):
        for col_h in range(1, max_col + 1):
            ws.cell(row=row_h, column=col_h).fill = navy

    # ESTATÍSTICAS
    preco_min = dataframe[col_precounitario].min()
    preco_med = dataframe[col_precounitario].mean()
    preco_mediano = dataframe[col_precounitario].median()
    preco_max = dataframe[col_precounitario].max()
    desvio = dataframe[col_precounitario].std()
    coef_var = (desvio / preco_med * 100) if preco_med != 0 else 0
    
    stats_list = [
        ("TOTAL", str(len(dataframe))),
        ("Mínimo", formatar_moeda_br(preco_min)),
        ("Médio", formatar_moeda_br(preco_med)),
        ("Mediana", formatar_moeda_br(preco_mediano)),
        ("Máximo", formatar_moeda_br(preco_max)),
        ("CV", f"{coef_var:.2f}%")
    ]
    
    for i, (label, val) in enumerate(stats_list):
        col_idx = i + 1
        ws.cell(row=5, column=col_idx).value = f"{label}: {val}"
        ws.cell(row=5, column=col_idx).font = gold_stat
        ws.cell(row=5, column=col_idx).fill = dark_navy
        ws.cell(row=5, column=col_idx).alignment = center
    ws.row_dimensions[5].height = 24

    # DADOS
    colunas_mapa = {
        'idCompra': 'ID Compra',
        'codigoItemCatalogo': 'Código Item',
        'descricaoItem': 'Descrição',
        'dataCompra': 'Data Compra',
        'unidadeFornecimento': 'Unidade',
        'quantidade': 'Quantidade',
        'precoUnitario': 'Preço Unitário',
        'niFornecedor': 'CNPJ Fornecedor',
        'nomeFornecedor': 'Fornecedor',
        'uf': 'UF',
        'codigoUasg': 'Cod. UASG',
        'nomeUasg': 'Nome UASG'
    }
    
    col_reais = []
    labels = []
    for col_esp, label in colunas_mapa.items():
        cr = encontrar_coluna(dataframe, [col_esp])
        if cr:
            col_reais.append(cr)
            labels.append(label)
            
    # Cabeçalho da tabela
    start_row = 7
    for c_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=start_row, column=c_idx)
        cell.value = label
        cell.font = header_font
        cell.fill = navy
        cell.alignment = center
        cell.border = gold_bottom
    ws.row_dimensions[start_row].height = 28
    
    # Linhas de dados
    for r_idx, row_data in enumerate(dataframe[col_reais].values, start_row + 1):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            cell.font = data_font
            cell.border = border_thin
            if (r_idx - start_row) % 2 == 0:
                cell.fill = alt_row
            else:
                cell.fill = white_bg
            
            # Formatação moeda na coluna Preço Unitário
            if labels[c_idx-1] == 'Preço Unitário' and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00'

    # JUSTIFICATIVA IN 65
    current_row = start_row + len(dataframe) + 2
    ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
    cell_just_title = ws.cell(row=current_row, column=1)
    paragraphs = JUSTIFICATIVA_COTACAO.split('\n\n')
    cell_just_title.value = paragraphs[0]
    cell_just_title.font = Font(name='Calibri', bold=True, color='001A4D', size=11)
    cell_just_title.alignment = center
    ws.row_dimensions[current_row].height = 30
    
    current_row += 1
    for par in paragraphs[1:]:
        ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
        cell_par = ws.cell(row=current_row, column=1)
        cell_par.value = par
        cell_par.font = Font(name='Calibri', size=10, color='333333')
        cell_par.alignment = Alignment(horizontal='justify', vertical='top', wrap_text=True)
        ws.row_dimensions[current_row].height = 60
        current_row += 1

    # Ajustar larguras
    column_widths = [12, 12, 40, 12, 12, 10, 15, 18, 30, 6, 12, 30]
    for i, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{start_row+1}'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Função para gerar PDF simples a partir dos dados
def gerar_relatorio_pdf_simples(dataframe, estatisticas, outliers_info, col_precounitario):
    """
    Gera um relatório PDF (Layout Notas Fiscais) com justificativa IN 65.
    """
    effective_w = 280
    x_start = 8
    
    item_str = 'Pesquisa de Preços - Cotação Pública'
    if not dataframe.empty:
        col_codigo = encontrar_coluna(dataframe, ['codigoItemCatalogo', 'codigo_item', 'codigo'])
        col_desc = encontrar_coluna(dataframe, ['descricaoItem', 'descricao', 'description'])
        
        codigo = str(dataframe.iloc[0][col_codigo]) if col_codigo else ""
        desc = str(dataframe.iloc[0][col_desc]) if col_desc else ""
        
        if codigo and desc:
            item_str = f"{codigo} - {desc}"
        elif desc:
            item_str = desc

    class PDFCotacao(FPDF):
        def header(self):
            self.set_fill_color(0, 26, 77)
            self.rect(0, 0, self.w, 30, 'F')
            self.set_fill_color(212, 175, 55)
            self.rect(0, 30, self.w, 1.2, 'F')
            self.set_font('Arial', 'B', 20)
            self.set_text_color(212, 175, 55)
            self.set_xy(10, 4)
            self.cell(60, 10, 'AtaCotada')
            self.set_font('Arial', 'B', 11)
            self.set_text_color(255, 255, 255)
            self.set_xy(10, 16)
            self.cell(0, 8, _pdf_safe(item_str))
            self.set_font('Arial', '', 7)
            self.set_text_color(180, 180, 180)
            self.set_xy(self.w - 70, 6)
            self.cell(60, 5, 'Marinha do Brasil', align='R')
            self.set_y(35)

        def footer(self):
            self.set_y(-10)
            self.set_fill_color(0, 26, 77)
            self.rect(0, self.h - 10, self.w, 10, 'F')
            self.set_font('Arial', 'I', 6.5)
            self.set_text_color(160, 160, 160)
            dh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.cell(0, 8, _pdf_safe(f'AtaCotada - Gerado em {dh}'), align='L')
            self.set_x(-30)
            self.cell(20, 8, f'{self.page_no()}/{{nb}}', align='R')

    pdf = PDFCotacao(orientation='L', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # ESTATÍSTICAS
    preco_min = dataframe[col_precounitario].min()
    preco_med = dataframe[col_precounitario].mean()
    preco_mediano = dataframe[col_precounitario].median()
    preco_max = dataframe[col_precounitario].max()
    desvio = dataframe[col_precounitario].std()
    coef_var = (desvio / preco_med * 100) if preco_med != 0 else 0
    
    stats_data = [
        ("Mínimo", f"R$ {formatar_moeda_br(preco_min)}"),
        ("Médio", f"R$ {formatar_moeda_br(preco_med)}"),
        ("Mediana", f"R$ {formatar_moeda_br(preco_mediano)}"),
        ("Máximo", f"R$ {formatar_moeda_br(preco_max)}"),
        ("Desvio Padrão", f"R$ {formatar_moeda_br(desvio)}"),
        ("CV", f"{coef_var:.2f}%".replace('.', ','))
    ]
    
    pdf.set_y(38)
    sw = effective_w / 6
    for label, val in stats_data:
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.set_fill_color(10, 37, 64)
        pdf.rect(x, y, sw - 1, 12, 'F')
        pdf.set_fill_color(212, 175, 55)
        pdf.rect(x, y, 1.2, 12, 'F')
        pdf.set_xy(x + 2, y + 1.5)
        pdf.set_font('Arial', '', 6)
        pdf.set_text_color(212, 175, 55)
        pdf.cell(sw - 4, 3, _pdf_safe(label))
        pdf.set_xy(x + 2, y + 5)
        pdf.set_font('Arial', 'B', 8)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(sw - 4, 5, _pdf_safe(val))
        pdf.set_xy(x + sw, y)
    pdf.ln(15)

    # DADOS
    colunas_mapa = {
        'idCompra': 'ID Compra',
        'dataCompra': 'Data',
        'codigoUasg': 'UASG',
        'unidadeFornecimento': 'Unid.',
        'quantidade': 'Qtd',
        'precoUnitario': 'V. Unitário',
        'niFornecedor': 'CNPJ',
        'nomeFornecedor': 'Fornecedor',
        'uf': 'UF',
        'nomeUasg': 'Órgão'
    }
    
    col_widths = [18, 16, 15, 12, 12, 18, 26, 60, 8, 95]
    
    col_reais = []
    labels = []
    for col_esp, label in colunas_mapa.items():
        cr = encontrar_coluna(dataframe, [col_esp])
        if cr:
            col_reais.append(cr)
            labels.append(label)

    # Tabela header
    pdf.set_fill_color(0, 26, 77)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Arial', 'B', 7)
    
    # Ajuste dinâmico caso alguma coluna não seja encontrada
    if len(labels) == len(col_widths):
        widths_to_use = col_widths
    else:
        cw = effective_w / len(labels) if len(labels) > 0 else effective_w
        widths_to_use = [cw] * len(labels)
        
    for i, l in enumerate(labels):
        pdf.cell(widths_to_use[i], 7, _pdf_safe(l), 1, 0, 'C', True)
    pdf.ln()

    # Linhas
    pdf.set_font('Arial', '', 6)
    pdf.set_text_color(51, 51, 51)
    for idx, row in dataframe.iterrows():
        if pdf.get_y() > 175:
            pdf.add_page()
            pdf.set_fill_color(0, 26, 77)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Arial', 'B', 7)
            for i, l in enumerate(labels):
                pdf.cell(widths_to_use[i], 7, _pdf_safe(l), 1, 0, 'C', True)
            pdf.ln()
            pdf.set_font('Arial', '', 6)
            pdf.set_text_color(51, 51, 51)
            
        fill = idx % 2 == 1
        pdf.set_fill_color(245, 248, 255)
        for i, col_r in enumerate(col_reais):
            val = row[col_r]
            w = widths_to_use[i]
            
            if labels[i] == 'V. Unitário':
                txt = f"R$ {formatar_moeda_br(val)}"
            elif labels[i] == 'Data':
                txt = str(val)[:10] if pd.notna(val) else ""
            else:
                txt = str(val) if pd.notna(val) else ""
                txt = txt.replace('\n', ' ').replace('\r', '')
                # Truncate string to avoid overlapping cells in PDF
                while len(txt) > 0 and pdf.get_string_width(_pdf_safe(txt)) > w - 2:
                    txt = txt[:-1]

            pdf.cell(w, 6, _pdf_safe(txt), 1, 0, 'C', fill)
        pdf.ln()

    # JUSTIFICATIVA
    pdf.add_page()
    pdf.ln(10)
    paragraphs = JUSTIFICATIVA_COTACAO.split('\n\n')
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(0, 26, 77)
    pdf.multi_cell(0, 8, _pdf_safe(paragraphs[0]), align='C')
    pdf.ln(5)
    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(51, 51, 51)
    for p in paragraphs[1:]:
        pdf.multi_cell(0, 6, _pdf_safe(p), align='J')
        pdf.ln(4)

    pdf_output = pdf.output(dest='S')
    if isinstance(pdf_output, bytearray):
        return bytes(pdf_output)
    return pdf_output

# Funções para consulta de fornecedores via API OpenCNPJ
def buscar_dados_fornecedor(cnpj):
    """Busca dados do fornecedor na API OpenCNPJ"""
    if not cnpj or pd.isna(cnpj):
        return None
    
    try:
        # Remover caracteres especiais do CNPJ
        cnpj_limpo = str(cnpj).replace('.', '').replace('/', '').replace('-', '').strip()
        if len(cnpj_limpo) != 14:
            return None
        
        api_url = f'https://api.opencnpj.org/{cnpj_limpo}'
        response = requests.get(api_url, timeout=5)
        
        if response.status_code == 200:
            dados = response.json()
            # Log para debug (comentado para produção)
            # print(f"DEBUG - Resposta API para {cnpj_limpo}: {dados}")
            return dados
        return None
    except Exception as e:
        # Silenciar exceções para não interromper o fluxo
        return None

def formatar_telefone(telefone):
    """Formata telefone para um padrão legível (21) 98130-7593"""
    if isinstance(telefone, dict):
        # Se é um dicionário com ddd e numero
        ddd = telefone.get('ddd', '').strip() if telefone.get('ddd') else ''
        numero = telefone.get('numero', '').strip() if telefone.get('numero') else ''
        
        if ddd and numero:
            # Formatar como (21) 98130-7593 ou (21) 3130-7593
            if len(numero) >= 8:
                numero_formatado = f"{numero[:4]}-{numero[4:]}"
            else:
                numero_formatado = numero
            return f"({ddd}) {numero_formatado}"
        elif numero:
            return numero
        elif ddd:
            return ddd
        else:
            return str(telefone)
    else:
        # Se é uma string, tentar extrair padrão (21) 98130-7593 se precisar
        telefone_str = str(telefone).strip()
        return telefone_str

def extrair_contatos_fornecedor(dados_fornecedor):
    """Extrai email, telefones, estado e cidade do retorno da API"""
    if not dados_fornecedor:
        return None
    
    try:
        # Variações possíveis para campos de email
        email = ''
        for campo_email in ['email', 'correio_eletronico', 'mail', 'e_mail', 'emailComercial']:
            email = dados_fornecedor.get(campo_email, '')
            if email:
                break
        
        # Variações possíveis para campos de estado/UF
        estado = ''
        for campo_estado in ['state', 'estado', 'uf', 'UF', 'state_code', 'sigla_estado']:
            valor = dados_fornecedor.get(campo_estado, '')
            if valor:
                estado = str(valor).upper().strip()
                break
        
        # Variações possíveis para campos de cidade
        cidade = ''
        for campo_cidade in ['city', 'cidade', 'municipio', 'municipality', 'city_name']:
            valor = dados_fornecedor.get(campo_cidade, '')
            if valor:
                cidade = str(valor).strip()
                break
        
        # Variações possíveis para campos de telefone
        telefones = []
        
        # Buscar em campos individuais
        for campo_tel in ['phone', 'telefone', 'phonePrimary', 'phoneSecondary', 
                         'telefone_comercial', 'telefone_principal', 'ddd_telefone',
                         'ddd_fax', 'fone', 'telephone']:
            valor = dados_fornecedor.get(campo_tel, '')
            if valor:
                # Formatar o telefone se for dicionário
                telefone_formatado = formatar_telefone(valor)
                if telefone_formatado and telefone_formatado not in telefones:
                    telefones.append(telefone_formatado)
        
        # Buscar em arrays
        for campo_array in ['phones', 'telefones', 'phone_numbers']:
            valores = dados_fornecedor.get(campo_array, [])
            if isinstance(valores, list):
                for v in valores:
                    if v:
                        # Formatar o telefone
                        telefone_formatado = formatar_telefone(v)
                        if telefone_formatado and telefone_formatado not in telefones:
                            telefones.append(telefone_formatado)
        
        # Remover duplicatas mantendo ordem
        telefones_unicos = []
        for t in telefones:
            if t not in telefones_unicos:
                telefones_unicos.append(t)
        
        # Formatar localização
        localizacao = ''
        if cidade and estado:
            localizacao = f"{cidade}, {estado}"
        elif cidade:
            localizacao = cidade
        elif estado:
            localizacao = estado
        
        return {
            'email': email if email else 'Não informado',
            'telefones': ', '.join(telefones_unicos) if telefones_unicos else 'Não informado',
            'cidade': cidade if cidade else 'Não informado',
            'estado': estado if estado else 'Não informado',
            'localizacao': localizacao if localizacao else 'Não informado',
            'debug_campos': list(dados_fornecedor.keys())  # Para debug
        }
    except Exception as e:
        return None



def gerar_html_fornecedores(dataframe, descricao_item=None):
    """Gera HTML com tabela de fornecedores e seus contatos"""
    
    # Encontrar as colunas corretas
    col_niFornecedor = encontrar_coluna(dataframe, ['niFornecedor', 'nifornecedor', 'ni_fornecedor'])
    col_cnpj = encontrar_coluna(dataframe, ['cnpj', 'CNPJ'])
    col_nomeFornecedor = encontrar_coluna(dataframe, ['nomeFornecedor', 'nomefornecedor', 'nome_fornecedor', 'fornecedor'])
    col_descricao = encontrar_coluna(dataframe, ['descricaoItem', 'descricao', 'description', 'item_descricao'])
    
    if not col_niFornecedor:
        return "<h3 style='color: red;'>Erro: Coluna 'niFornecedor' não encontrada no dataframe</h3>"
    
    # Se não foi passada descrição, tenta extrair do dataframe
    if not descricao_item and col_descricao and len(dataframe) > 0:
        descricao_item = str(dataframe.iloc[0][col_descricao])
    
    # Obter fornecedores únicos
    fornecedores_unicos = dataframe.drop_duplicates(subset=[col_niFornecedor])
    
    # CSS inline compacto
    css_style = """
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); font-family: 'Arial', sans-serif; padding: 2rem; min-height: 100vh; }
    .container { max-width: 1200px; margin: 0 auto; background: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3); overflow: hidden; }
    .header { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); color: #ffffff; padding: 2rem; text-align: center; }
    .header h1 { color: #d4af37; font-size: 36px; margin-bottom: 0.5rem; letter-spacing: 2px; }
    .header p { font-size: 14px; color: #ffffff; }
    .header .descricao-item { font-size: 10px; color: #d4af37; letter-spacing: 0.5px; margin-top: 0.5rem; font-style: italic; opacity: 0.8; }
    .content { padding: 2rem; }
    .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 1rem; margin-bottom: 2rem; }
    .stat-box { background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center; }
    .stat-box label { color: #d4af37; font-weight: bold; font-size: 12px; letter-spacing: 1px; }
    .stat-box .value { color: #ffffff; font-size: 28px; font-weight: bold; margin-top: 0.5rem; }
    table { width: 100%; border-collapse: collapse; margin-top: 1rem; }
    thead { background: #001a4d; color: #ffffff; }
    th { padding: 1rem; text-align: left; font-weight: bold; border-bottom: 2px solid #d4af37; }
    td { padding: 0.75rem 1rem; border-bottom: 1px solid #e0e0e0; }
    tbody tr:hover { background: #f5f5f5; }
    tbody tr:nth-child(even) { background: #f9f9f9; }
    .contato-info { font-size: 13px; color: #333; }
    .email { color: #0033cc; text-decoration: none; }
    .email:hover { text-decoration: underline; }
    .phone { color: #006600; }
    .footer { background: #f0f0f0; padding: 1rem 2rem; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }
    """
    
    # Coletar dados de cada fornecedor
    linhas = []
    com_email_count = 0
    com_telefone_count = 0
    
    for idx, row in fornecedores_unicos.iterrows():
        ni_fornecedor = row.get(col_niFornecedor, '')
        cnpj = row.get(col_cnpj, '') if col_cnpj else ''
        nome_fornecedor = row.get(col_nomeFornecedor, 'Não informado') if col_nomeFornecedor else 'Não informado'
        
        # Buscar dados na API usando niFornecedor
        dados_fornecedor = buscar_dados_fornecedor(ni_fornecedor)
        contatos = extrair_contatos_fornecedor(dados_fornecedor) if dados_fornecedor else None
        
        if contatos:
            email = contatos['email']
            telefones = contatos['telefones']
            localizacao = contatos['localizacao']
            
            if email and email != 'Não informado':
                com_email_count += 1
            if telefones and telefones != 'Não informado':
                com_telefone_count += 1
        else:
            email = 'Não informado'
            telefones = 'Não informado'
            localizacao = 'Não informado'
        
        # Formatar email como link
        email_html = f'<a href="mailto:{email}" class="email">{email}</a>' if email and email != 'Não informado' else 'Não informado'
        
        linha = f"<tr><td>{ni_fornecedor}</td><td>{cnpj}</td><td>{nome_fornecedor}</td><td>{localizacao}</td><td><span class='contato-info'>{email_html}</span></td><td><span class='contato-info phone'>{telefones}</span></td></tr>"
        linhas.append(linha)
    
    linhas_tabela = ''.join(linhas)
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    html_completo = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Contatos de Fornecedores - AtaCotada</title>
    <style>{css_style}</style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>AtaCotada</h1>
            <p>Contatos de Fornecedores - Marinha do Brasil</p>
            {f'<div class="descricao-item">Fornecedores de: {descricao_item[:100]}</div>' if descricao_item else ''}
        </div>
        <div class="content">
            <div class="stats">
                <div class="stat-box">
                    <label>TOTAL DE FORNECEDORES</label>
                    <div class="value">{len(fornecedores_unicos)}</div>
                </div>
                <div class="stat-box">
                    <label>COM EMAIL</label>
                    <div class="value">{com_email_count}</div>
                </div>
                <div class="stat-box">
                    <label>COM TELEFONE</label>
                    <div class="value">{com_telefone_count}</div>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>NI Fornecedor</th>
                        <th>CNPJ</th>
                        <th>Fornecedor</th>
                        <th>Localização</th>
                        <th>Email</th>
                        <th>Telefones</th>
                    </tr>
                </thead>
                <tbody>
                    {linhas_tabela}
                </tbody>
            </table>
        </div>
        <div class="footer">
            <p>Relatório gerado por AtaCotada em {data_hora}</p>
        </div>
    </div>
</body>
</html>"""
    
    return html_completo

# URLs atualizadas
consultarItemMaterial_base_url = 'https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/1_consultarMaterial'
consultarItemServico_base_url = 'https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/3_consultarServico'

def obter_itens(tipo_item, codigo_item_catalogo, pagina, tamanho_pagina):
    url = consultarItemMaterial_base_url if tipo_item == 'Material' else consultarItemServico_base_url
    params = {
        'pagina': pagina,   
        'tamanhoPagina':tamanho_pagina,  # Ajuste para 500 itens por página
        'codigoItemCatalogo': codigo_item_catalogo
    }
    try:
        response = requests.get(url, params=params)
        if response.status_code == 200:
            json_response = response.json()
            itens = json_response.get('resultado', [])
            paginas_restantes = json_response.get('paginasRestantes', 0)
            total_paginas = json_response.get('totalPaginas', 0)
            return itens, paginas_restantes, total_paginas
        elif response.status_code == 400:
            st.error("❌ Erro 400 - Requisição Inválida: O código do item pode estar inválido ou em formato incorreto. Verifique o código no catálogo de compras do governo federal.")
            return [], 0, 0
        else:
            st.error(f"Erro na consulta: {response.status_code}")
            return [], 0, 0
    except Exception as e:
        st.error(f"Erro ao realizar a requisição: {str(e)}")
        return [], 0, 0

# Streamlit UI
# Header customizado
col_header1, col_header2, col_header3 = st.columns([1, 2, 1])
with col_header2:
    st.markdown("""
        <div class="header-container">
            <div class="logo-text">⚓ MARINHA DO BRASIL ⚓</div>
            <div class="sistema-nome">AtaCotada</div>
            <div class="subtitulo">PESQUISA DE PREÇOS DE MATERIAIS E SERVIÇOS</div>
        </div>
    """, unsafe_allow_html=True)

# Disclaimer com estilo
st.markdown("""
    <div style="background-color: #0a2540; padding: 1rem; border-radius: 8px; border-left: 5px solid #d4af37; margin-bottom: 1.5rem;">
        <p style="color: #d4af37; font-weight: bold; margin-bottom: 0.5rem;">📚 Orientação</p>
        <p style="color: #ffffff;">Utilize o catálogo de materiais e serviços do Governo Federal para encontrar o código desejado.</p>
    </div>
""", unsafe_allow_html=True)

components.iframe("https://catalogo.compras.gov.br/cnbs-web/busca", height=450, width=1080, scrolling=True)

# Seção de consulta
st.markdown("""
    <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
        🔍 Consultar Item
    </div>
""", unsafe_allow_html=True)

col_tipo, col_codigo, col_botao = st.columns([1, 2, 1])

with col_tipo:
    tipo_item = st.selectbox("Tipo de Item", ['Material', 'Serviço'], key='tipo_item')

with col_codigo:
    codigo_item_catalogo = st.text_input("Código do Item de Catálogo", value="", key='codigo_item_catalogo', placeholder="Ex: 123456")

with col_botao:
    st.write("")
    st.write("")
    consultar = st.button('🔎 Consultar', use_container_width=True)

pagina = 1
tamanho_pagina = 500

if consultar:
    if codigo_item_catalogo:  # Verifica se o código do item de catálogo não está vazio
        # Validação do código do item
        codigo_item_catalogo = codigo_item_catalogo.strip()
        
        # Verificar comprimento máximo (limite de segurança/API)
        if len(codigo_item_catalogo) > 20:
            st.error("❌ Erro de Validação: O código do item não deve exceder 20 caracteres. Verifique o código e tente novamente.")
        # Verificar se contém apenas números e caracteres válidos
        elif not codigo_item_catalogo.isdigit():
            st.error("❌ Erro de Validação: O código do item deve conter apenas números. Verifique o formato e tente novamente.")
        else:
            itens, paginas_restantes, total_paginas = obter_itens(tipo_item, codigo_item_catalogo, pagina, tamanho_pagina)
            if itens:  # Ensure 'itens' is not empty before proceeding
                st.session_state['itens'] = itens      
            else:
                st.error("Nenhum item encontrado. Por favor, tente com um código diferente ou verifique a conexão com a API.")
    else:
        st.warning("Por favor, informe o código do item de catálogo para realizar a consulta.")

# Mostrar filtros e resultados se houver dados em session_state
if st.session_state.get('itens'):
    try:
        # Ensure 'itens' is in the expected format before normalization
        if isinstance(st.session_state['itens'], list) and all(isinstance(item, dict) for item in st.session_state['itens']):
            df_completo = pd.json_normalize(st.session_state['itens'])
            
            # Encontrar colunas com nomes variáveis
            col_unidade = encontrar_coluna(df_completo, ['unidadeFornecimento', 'unidade_fornecimento', 'unidade', 'unitFornecimento'])
            col_nome_unidade = encontrar_coluna(df_completo, ['nomeUnidadeFornecimento', 'nome_unidade_fornecimento', 'nomeUnidade', 'nome_unidade'])
            col_estado = encontrar_coluna(df_completo, ['uf', 'estado', 'estadoUf', 'estado_uf'])
            col_precounitario = encontrar_coluna(df_completo, ['precoUnitario', 'preco_unitario', 'preço_unitário'])
            col_uasg = encontrar_coluna(df_completo, ['nomeUasg', 'nome_uasg', 'uasg', 'nomeUAsg'])
            col_data_compra = encontrar_coluna(df_completo, ['dataCompra', 'data_compra', 'data'])

            # Filtro de janela de 1 ano (sempre a partir do dia presente)
            if col_data_compra:
                try:
                    df_completo[col_data_compra] = pd.to_datetime(df_completo[col_data_compra], errors='coerce')
                    hoje = pd.Timestamp.now().normalize()
                    um_ano_atras = hoje - pd.DateOffset(years=1)
                    df_completo = df_completo[
                        (df_completo[col_data_compra] >= um_ano_atras) & 
                        (df_completo[col_data_compra] <= hoje)
                    ].copy()
                except Exception as e:
                    st.warning(f"Erro ao filtrar por data: {e}")
            
            # Seção de Filtros
            st.markdown("""
                <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
                    ⚙️ Filtros de Consulta
                </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            # Inicializar variáveis de filtro
            unidade_selecionada = []
            nome_unidade_input = ""
            estado_selecionado = []
            uasg_input = ""
            remover_outliers = False
            
            with col1:
                # Filtro de unidade de fornecimento (se a coluna existir)
                if col_unidade:
                    unidades_disponiveis = df_completo[col_unidade].dropna().unique().tolist()
                    if unidades_disponiveis:
                        unidade_selecionada = st.multiselect(
                            "Unidade de Fornecimento",
                            options=sorted(unidades_disponiveis),
                            default=unidades_disponiveis,
                            key='filtro_unidade'
                        )
                
                # Filtro de nome unidade fornecimento (se a coluna existir)
                if col_nome_unidade:
                    nome_unidade_input = st.text_input(
                        "Nome Unidade Fornecimento",
                        value="",
                        key='filtro_nome_unidade',
                        placeholder="Filtrar por nome..."
                    )
                
                # Filtro de nome UASG (se a coluna existir) - ABAIXO DO FILTRO DA UNIDADE
                if col_uasg:
                    uasg_input = st.text_input(
                        "Nome UASG",
                        value="",
                        key='filtro_uasg',
                        placeholder="Filtrar por nome..."
                    )
            
            with col2:
                # Filtro de estado (se a coluna existir)
                if col_estado:
                    estados_disponiveis = df_completo[col_estado].dropna().unique().tolist()
                    if estados_disponiveis:
                        estado_selecionado = st.multiselect(
                            "Estado (UF)",
                            options=sorted(estados_disponiveis),
                            default=estados_disponiveis,
                            key='filtro_estado'
                        )
                
                # Checkbox para remover outliers
                if col_precounitario:
                    remover_outliers = st.checkbox(
                        "🔍 Remover outliers (IQR)",
                        value=False,
                        key='filtro_outliers'
                    )
                    st.write("")
            
            # Aplicar filtros iniciais (sem preço ainda)
            dataframe = df_completo.copy()
            
            # Filtro de unidade de fornecimento
            if col_unidade and unidade_selecionada:
                dataframe = dataframe[dataframe[col_unidade].isin(unidade_selecionada)]
            
            # Filtro de nome unidade fornecimento
            if col_nome_unidade and nome_unidade_input:
                dataframe = dataframe[dataframe[col_nome_unidade].str.contains(nome_unidade_input, case=False, na=False)]
            
            # Filtro de estado
            if col_estado and estado_selecionado:
                dataframe = dataframe[dataframe[col_estado].isin(estado_selecionado)]
            
            # Filtro de nome UASG
            if col_uasg and uasg_input:
                dataframe = dataframe[dataframe[col_uasg].str.contains(uasg_input, case=False, na=False)]
            
            # Remover outliers se selecionado (ANTES do slider de preço)
            outliers_info = None
            if remover_outliers and col_precounitario and col_precounitario in dataframe.columns:
                dataframe_temp, outliers_removidos, limite_inf, limite_sup = remover_outliers_iqr(dataframe, col_precounitario)
                outliers_info = {
                    'removidos': outliers_removidos,
                    'limite_inf': limite_inf,
                    'limite_sup': limite_sup
                }
                dataframe = dataframe_temp
            
            # Agora mostrar o slider de preço com os limites corretos (após remover outliers se necessário)
            faixa_preco = (0, 0)
            if col_precounitario and col_precounitario in dataframe.columns and len(dataframe) > 0:
                preco_min_filtrado = dataframe[col_precounitario].min()
                preco_max_filtrado = dataframe[col_precounitario].max()

                if float(preco_min_filtrado) < float(preco_max_filtrado):
                    faixa_preco = st.slider(
                        "Faixa de Preço Unitário (R$)",
                        min_value=float(preco_min_filtrado),
                        max_value=float(preco_max_filtrado),
                        value=(float(preco_min_filtrado), float(preco_max_filtrado)),
                        format="R$ %.2f",
                        step=0.01,
                        key='filtro_preco'
                    )
                else:
                    faixa_preco = (float(preco_min_filtrado), float(preco_max_filtrado))
                    st.info(
                        f"Todos os itens filtrados possuem o mesmo preço unitário: {formatar_preco_reais(preco_min_filtrado)}."
                    )
                
                # Mostrar valores selecionados em formato brasileiro
                st.markdown(f"""
                    <div style="background-color: #0a2540; padding: 0.8rem; border-radius: 6px; border-left: 4px solid #d4af37; margin-top: 0.5rem;">
                        <span style="color: #d4af37; font-weight: bold;">Intervalo selecionado:</span>
                        <span style="color: #ffffff;"> {formatar_preco_reais(faixa_preco[0])} até {formatar_preco_reais(faixa_preco[1])}</span>
                    </div>
                """, unsafe_allow_html=True)
            
            # Aplicar filtro de preço
            if col_precounitario and col_precounitario in dataframe.columns:
                dataframe = dataframe[
                    (dataframe[col_precounitario] >= faixa_preco[0]) &
                    (dataframe[col_precounitario] <= faixa_preco[1])
                ]
            
            # Mostrar estatísticas apenas dos dados filtrados
            if len(dataframe) > 0:
                # Seção de resultados
                st.markdown("""
                    <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
                        📊 Resultados
                    </div>
                """, unsafe_allow_html=True)
                
                # Mostrar contagem com destaque
                col_info_a, col_info_b, col_btn_excel, col_btn_pdf = st.columns([1, 1, 0.6, 0.6])
                with col_info_a:
                    st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                            <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">REGISTROS ENCONTRADOS</div>
                            <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">{len(dataframe)}</div>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Mostrar informações sobre outliers removidos
                if outliers_info:
                    with col_info_b:
                        st.markdown(f"""
                            <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                                <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">OUTLIERS REMOVIDOS</div>
                                <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">{outliers_info['removidos']}</div>
                            </div>
                        """, unsafe_allow_html=True)
                
                # Mostrar informações sobre outliers removidos quando não houver
                if not outliers_info:
                    with col_info_b:
                        st.markdown(f"""
                            <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                                <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">OUTLIERS REMOVIDOS</div>
                                <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">0</div>
                            </div>
                        """, unsafe_allow_html=True)
                
                # Mostrar faixa de outliers
                if outliers_info:
                    st.markdown(f"""
                        <div style="background-color: #0a2540; padding: 1rem; border-radius: 8px; border-left: 5px solid #d4af37; margin-bottom: 1.5rem;">
                            <span style="color: #d4af37; font-weight: bold;">📊 Faixa válida (sem outliers):</span><br>
                            <span style="color: #ffffff;">{formatar_preco_reais(outliers_info['limite_inf'])} até {formatar_preco_reais(outliers_info['limite_sup'])}</span>
                        </div>
                    """, unsafe_allow_html=True)

                st.markdown("### Itens encontrados")
                df_exibicao = preparar_dataframe_exibicao(dataframe)
                df_editor = df_exibicao.copy()
                df_editor.insert(0, 'Selecionar', False)

                df_editor = st.data_editor(
                    df_editor,
                    hide_index=True,
                    use_container_width=True,
                    disabled=[col for col in df_editor.columns if col != 'Selecionar'],
                    column_config={
                        'Selecionar': st.column_config.CheckboxColumn(
                            'Selecionar',
                            help='Marque os itens que devem entrar no relatório.'
                        ),
                        'Descrição do Item': st.column_config.TextColumn('Descrição do Item', width='large'),
                        'Objeto Compra': st.column_config.TextColumn('Objeto Compra', width='large'),
                        'Nome UASG': st.column_config.TextColumn('Nome UASG', width='medium'),
                        'Nome Fornecedor': st.column_config.TextColumn('Nome Fornecedor', width='medium')
                    },
                    key='resultado_cotacao_editor'
                )

                indices_selecionados = df_editor.index[df_editor['Selecionar'].fillna(False)].tolist()
                dataframe_relatorio = dataframe.iloc[indices_selecionados].copy() if indices_selecionados else dataframe.copy()
                usa_selecao = len(indices_selecionados) > 0

                st.caption('Visualização com destaque das linhas selecionadas:')
                exibir_tabela_com_destaque(df_editor)

                if usa_selecao:
                    st.info(
                        f"Relatórios e cálculos considerarão apenas {len(dataframe_relatorio)} item(ns) selecionado(s)."
                    )
                else:
                    st.caption('Marque a coluna Selecionar para gerar os relatórios somente com os itens desejados.')
                
                # Mostrar estatísticas se a coluna precoUnitario existir
                if col_precounitario and col_precounitario in dataframe.columns:
                    mean = dataframe_relatorio[col_precounitario].mean()
                    median = dataframe_relatorio[col_precounitario].median()
                    preco_min = dataframe_relatorio[col_precounitario].min()
                    preco_max = dataframe_relatorio[col_precounitario].max()
                    std = dataframe_relatorio[col_precounitario].std()
                    cv = ((std / mean) * 100) if mean != 0 else 0

                    estatisticas = {
                        'min': preco_min,
                        'mean': mean,
                        'median': median,
                        'max': preco_max,
                        'std': std,
                        'cv': cv
                    }
                    
                    st.markdown(
                        f"""
                            |Preço Mínimo|Preço Unitário Médio|Preço Unitário Mediano|Preço Máximo|Desvio Padrão|Coeficiente de Variação|
                            |:----------:|:------------------:|:--------------------:|:----------:|:-----------:|:---------------------:|
                            |**{formatar_preco_reais(preco_min)}**|**{formatar_preco_reais(mean)}**|**{formatar_preco_reais(median)}**|**{formatar_preco_reais(preco_max)}**|**{formatar_preco_reais(std)}**|**{formatar_numero_br(cv)}%**|
                        """
                    )

                    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename_excel = f"relatorio_AtaCotada_{data_hora}.xlsx"
                    filename_pdf = f"relatorio_AtaCotada_{data_hora}.pdf"
                    excel_bytes = gerar_relatorio_excel(dataframe_relatorio, estatisticas, outliers_info, col_precounitario)
                    pdf_bytes = gerar_relatorio_pdf_simples(dataframe_relatorio, estatisticas, outliers_info, col_precounitario)

                    with col_btn_excel:
                        st.download_button(
                            label="📊 Excel (seleção)" if usa_selecao else "📊 Excel",
                            data=excel_bytes,
                            file_name=filename_excel,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                    with col_btn_pdf:
                        st.download_button(
                            label="📄 PDF (seleção)" if usa_selecao else "📄 PDF",
                            data=pdf_bytes,
                            file_name=filename_pdf,
                            mime="application/pdf",
                            use_container_width=True
                        )

                    st.markdown("<br>", unsafe_allow_html=True)
                    col_fornecedores = st.columns([0.6, 1, 1])
                    with col_fornecedores[0]:
                        if st.button("📞 Consultar Fornecedores", use_container_width=True, key="btn_fornecedores"):
                            with st.spinner("Consultando dados de fornecedores..."):
                                col_descricao = encontrar_coluna(dataframe_relatorio, ['descricaoItem', 'descricao', 'description', 'item_descricao'])
                                descricao_item = None
                                if col_descricao and len(dataframe_relatorio) > 0:
                                    descricao_item = str(dataframe_relatorio.iloc[0][col_descricao])

                                html_fornecedores = gerar_html_fornecedores(dataframe_relatorio, descricao_item)
                                html_bytes = html_fornecedores.encode('utf-8')
                                html_filename = f"fornecedores_AtaCotada_{data_hora}.html"

                                st.download_button(
                                    label="📋 Abrir Relatório de Fornecedores",
                                    data=html_bytes,
                                    file_name=html_filename,
                                    mime="text/html",
                                    use_container_width=True,
                                    key="btn_download_fornecedores"
                                )
                else:
                    st.caption('A seleção é mantida na tabela, mas os cálculos do relatório dependem da coluna de preço unitário.')
            else:
                st.warning("Nenhum resultado encontrado com os filtros aplicados. Tente ajustar os critérios.")
                     
        else:
            st.error("Formato dos itens inválido para normalização.")
    except Exception as e:
        st.error(f"Erro ao processar os itens: {str(e)}")
        # Mostrar colunas disponíveis quando há erro
        if st.session_state.get('itens'):
            try:
                df_debug = pd.json_normalize(st.session_state['itens'])
                with st.expander("ℹ️ Colunas disponíveis (Debug)"):
                    st.write("Colunas encontradas no dataframe:")
                    st.write(list(df_debug.columns))
                    st.write("\nPrimeiras linhas dos dados:")
                    st.write(df_debug.head())
            except:
                pass

# Botão Nova Pesquisa
if st.session_state.get('itens'):
    st.markdown("""
        <div style="margin-top: 2rem; padding: 1.5rem; background: linear-gradient(135deg, #0a2540 0%, #164863 100%); border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
            <p style="color: #d4af37; font-weight: bold; margin-bottom: 1rem; font-size: 14px;">Deseja realizar uma nova pesquisa?</p>
        </div>
    """, unsafe_allow_html=True)
    
    col_nova_pesquisa = st.columns([1, 2, 1])[1]
    with col_nova_pesquisa:
        if st.button("🔄 Nova Pesquisa", use_container_width=True, key='nova_pesquisa'):
            st.session_state['itens'] = None
            st.rerun()

# Rodapé
st.markdown("""
    <div style="margin-top: 3rem; padding-top: 2rem; border-top: 1px solid #0033cc; text-align: center;">
        <p style="color: #d4af37; font-size: 12px; margin: 0; letter-spacing: 0.5px;">
            Idealizado e Desenvolvido por COpAb - Sobressalentes - V1.2026
        </p>
    </div>
""", unsafe_allow_html=True)
