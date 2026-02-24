import streamlit as st
import requests
import pandas as pd
import streamlit.components.v1 as components
from datetime import datetime
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuração da página
st.set_page_config(
    page_title="AtaCotada - Marinha do Brasil",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para tema azul escuro com dourado
st.markdown("""
    <style>
        :root {
            --azul-escuro: #001a4d;
            --azul-claro: #0033cc;
            --dourado: #d4af37;
            --branco: #ffffff;
            --cinza-claro: #f0f0f0;
        }
        
        * {
            margin: 0;
            padding: 0;
        }
        
        body, .main {
            background-color: #001a4d;
            color: #ffffff;
        }
        
        .stApp {
            background-color: #001a4d;
        }
        
        .main .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
        }
        
        /* Header customizado */
        .header-container {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            padding: 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
            text-align: center;
        }
        
        .logo-text {
            color: #ffffff;
            font-size: 14px;
            font-weight: 600;
            letter-spacing: 2px;
            margin-bottom: 0.5rem;
        }
        
        .sistema-nome {
            color: #d4af37;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 3px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);
            margin: 0.5rem 0;
            font-family: 'Arial Black', sans-serif;
        }
        
        .subtitulo {
            color: #ffffff;
            font-size: 14px;
            margin-top: 0.5rem;
            letter-spacing: 1px;
        }
        
        /* Filtros container */
        .filtros-container {
            background: linear-gradient(135deg, #0a2540 0%, #164863 100%);
            padding: 1.5rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            border-left: 5px solid #d4af37;
        }
        
        /* Cards de estatísticas */
        .stats-card {
            background: linear-gradient(135deg, #0a2540 0%, #0f4c75 100%);
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 5px solid #d4af37;
            margin-bottom: 1rem;
        }
        
        /* Botões */
        .stButton > button {
            background-color: #d4af37;
            color: #001a4d;
            border: none;
            border-radius: 6px;
            padding: 0.75rem 1.5rem;
            font-weight: bold;
            font-size: 16px;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .stButton > button:hover {
            background-color: #ffd700;
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(212, 175, 55, 0.3);
        }
        
        /* Inputs e Selects */
        .stTextInput > div > div > input,
        .stMultiSelect > div > div > div,
        .stSelectbox > div > div > select,
        .stSlider > div > div > div {
            background-color: #0a2540 !important;
            color: #ffffff !important;
            border: 2px solid #0033cc !important;
            border-radius: 6px !important;
        }
        
        /* Labels e textos de input - BRANCO */
        label, .stText, .stLabel, .stMultiSelect label, .stSelectbox label, .stTextInput label {
            color: #ffffff !important;
        }
        
        /* Checkbox labels e texto - MAIS ESPECÍFICO */
        .stCheckbox {
            background-color: transparent;
        }
        
        .stCheckbox > label {
            color: #ffffff !important;
            background-color: transparent;
        }
        
        .stCheckbox > div > label {
            color: #ffffff !important;
        }
        
        .stCheckbox label {
            color: #ffffff !important;
        }
        
        .stCheckbox label span {
            color: #ffffff !important;
        }
        
        .stCheckbox label div {
            color: #ffffff !important;
        }
        
        .stCheckbox input[type="checkbox"] + label {
            color: #ffffff !important;
        }
        
        /* Pegar todo texto dentro de checkbox */
        .stCheckbox * {
            color: #ffffff !important;
        }
        
        /* Labels dos widgets */
        div[role="radiogroup"] label,
        .stMultiSelect > div > div[class*="stMultiSelect"] label {
            color: #ffffff !important;
        }
        
        /* Todos os span e div dentro dos filtros */
        .stMultiSelect > div > div > span,
        .stSelectbox > div > div > span {
            color: #ffffff !important;
        }
        
        /* Títulos */
        h1, h2, h3 {
            color: #d4af37;
            text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.5);
        }
        
        /* Expanders */
        .streamlit-expanderHeader {
            background-color: #0a2540;
            color: #d4af37;
            border: 2px solid #0033cc;
            border-radius: 6px;
        }
        
        /* Info boxes */
        .stAlert {
            background-color: #0a2540;
            border: 2px solid #d4af37;
            border-radius: 6px;
        }
        
        /* Tabelas */
        .streamlit-dataframe {
            background-color: #0a2540;
            color: #ffffff;
        }
        
        table {
            color: #ffffff;
            background-color: #0a2540;
        }
        
        th {
            background-color: #0033cc !important;
            color: #d4af37 !important;
            font-weight: bold;
        }
        
        td {
            background-color: #0f4c75 !important;
            color: #ffffff;
        }
        
        tr:hover {
            background-color: #164863 !important;
        }
        
        /* Markdown */
        .markdown-text-container {
            color: #ffffff;
        }
    </style>
""", unsafe_allow_html=True)

# Função para formatar preço em reais
def formatar_preco_reais(valor):
    if valor is None:
        return 'Preço não disponível'
    else:
        return f'R$ {valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

# Função para formatar número em formato brasileiro
def formatar_numero_br(valor):
    """Formata número para padrão brasileiro (vírgula como separador decimal)"""
    return f'{valor:.2f}'.replace('.', ',')

def formatar_moeda_br(valor):
    """Formata valor monetário para padrão brasileiro com separador de milhares (1.234.567,89)"""
    if valor is None or (isinstance(valor, float) and valor != valor):  # NaN check
        return ""
    return f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

# Função para encontrar coluna com variações de nome
def encontrar_coluna(df, nomes_possiveis):
    """Procura por uma coluna entre várias variações de nome (case-insensitive)"""
    colunas_lower = {col.lower(): col for col in df.columns}
    # Preferência por correspondência exata (case-insensitive)
    for nome in nomes_possiveis:
        chave = nome.lower()
        if chave in colunas_lower:
            return colunas_lower[chave]

    # Se não houver correspondência exata, procurar por substring nos nomes das colunas
    for col_lower, col_original in colunas_lower.items():
        for nome in nomes_possiveis:
            if nome.lower() in col_lower:
                return col_original

    return None

# Função para remover outliers usando o método IQR (Interquartile Range)
def remover_outliers_iqr(dataframe, coluna):
    """
    Remove outliers usando o método IQR.
    Outliers são valores fora da faixa [Q1 - 1.5*IQR, Q3 + 1.5*IQR]
    Retorna: (df_sem_outliers, quantidade_removida, limite_inferior, limite_superior)
    """
    # Garantir que a coluna exista
    if coluna not in dataframe.columns:
        return dataframe.copy(), 0, None, None

    # Calcular quartis e IQR ignorando valores nulos
    serie = dataframe[coluna].dropna()
    if serie.empty:
        return dataframe.copy(), 0, None, None

    Q1 = serie.quantile(0.25)
    Q3 = serie.quantile(0.75)
    IQR = Q3 - Q1
    limite_inferior = Q1 - 1.5 * IQR
    limite_superior = Q3 + 1.5 * IQR

    df_sem_outliers = dataframe[(dataframe[coluna] >= limite_inferior) & (dataframe[coluna] <= limite_superior)]
    outliers_removidos = len(dataframe) - len(df_sem_outliers)

    return df_sem_outliers, outliers_removidos, limite_inferior, limite_superior

# Função para gerar relatório em Excel formatado
def gerar_relatorio_excel(dataframe, estatisticas, outliers_info, col_precounitario):
    """
    Gera um relatório em Excel com os dados formatados e editável.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Estilos
    header_fill = PatternFill(start_color="001A4D", end_color="001A4D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="D4AF37", size=14)
    stats_fill = PatternFill(start_color="0033CC", end_color="0033CC", fill_type="solid")
    stats_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws['A1'] = "AtaCotada"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:L1')
    ws.row_dimensions[1].height = 25
    
    # Subtítulo
    ws['A2'] = "MARINHA DO BRASIL - Pesquisa de Preços"
    ws['A2'].font = Font(size=10, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:L2')
    
    # Data
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws['A3'] = f"Relatório gerado em: {data_hora}"
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:L3')
    ws.row_dimensions[3].height = 18
    
    # ESTATÍSTICAS
    ws['A5'] = "ESTATÍSTICAS"
    ws['A5'].font = Font(bold=True, size=11, color="D4AF37")
    ws.row_dimensions[5].height = 18
    
    # Cabeçalho estatísticas
    stats_labels = ['Preço Mín', 'Preço Médio', 'Preço Mediano', 'Preço Máx', 'Desvio Padrão', 'Coef. Variação']
    for col, label in enumerate(stats_labels, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = label
        cell.font = stats_font
        cell.fill = stats_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Valores estatísticas
    preco_min = dataframe[col_precounitario].min()
    preco_med = dataframe[col_precounitario].mean()
    preco_mediano = dataframe[col_precounitario].median()
    preco_max = dataframe[col_precounitario].max()
    desvio = dataframe[col_precounitario].std()
    coef_var = (desvio / preco_med * 100) if preco_med != 0 else 0
    
    stats_values = [preco_min, preco_med, preco_mediano, preco_max, desvio, coef_var]
    for col, valor in enumerate(stats_values, 1):
        cell = ws.cell(row=7, column=col)
        if col == 6:  # Coeficiente de variação
            cell.value = valor
            cell.number_format = '0.00"%"'
        else:
            cell.value = valor
            cell.number_format = 'R$ #,##0.00'
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    ws.row_dimensions[7].height = 20
    
    # INFORMAÇÕES ADICIONAIS
    ws['A9'] = "INFORMAÇÕES"
    ws['A9'].font = Font(bold=True, size=11, color="D4AF37")
    
    total_registros = len(dataframe)
    outliers_removidos = outliers_info['removidos'] if outliers_info else 0
    
    ws['A10'] = f"Total de Registros: {total_registros}"
    ws['A11'] = f"Outliers Removidos: {outliers_removidos}"
    ws['A10'].font = Font(size=10, bold=True)
    ws['A11'].font = Font(size=10, bold=True)
    
    # DADOS DA TABELA
    ws['A13'] = "DADOS"
    ws['A13'].font = Font(bold=True, size=11, color="D4AF37")
    
    # Preparar colunas a serem exibidas
    colunas_mapa = {
        'idCompra': 'ID Compra',
        'niFornecedor': 'NI Fornecedor',
        'codigoItemCatalogo': 'Código Item',
        'descricaoItem': 'Descrição',
        'dataCompra': 'Data Compra',
        'unidadeFornecimento': 'Unidade',
        'quantidade': 'Quantidade',
        'precoUnitario': 'Preço Unitário',
        'cnpj': 'CNPJ',
        'nomeFornecedor': 'Fornecedor',
        'codigoUasg': 'Cod. UASG',
        'nomeUasg': 'Nome UASG'
    }
    
    colunas_encontradas = {}
    for col_esperada, label in colunas_mapa.items():
        col_real = encontrar_coluna(dataframe, [col_esperada, col_esperada.lower()])
        if col_real:
            colunas_encontradas[col_real] = label
    
    # Selecionar apenas as colunas encontradas
    df_export = dataframe[[col for col in colunas_encontradas.keys()]].copy()
    
    # Renomear colunas
    df_export.columns = [colunas_encontradas[col] for col in df_export.columns]
    
    # Formatar preços
    if 'Preço Unitário' in df_export.columns:
        df_export['Preço Unitário'] = df_export['Preço Unitário'].apply(lambda x: x if pd.isna(x) else float(x))
    
    # Escrever dados na planilha
    start_row = 14
    for r_idx, row in enumerate(dataframe_to_rows(df_export, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            
            # Formatar cabeçalho
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # Formatar preços
                if c_idx == 6 and isinstance(value, (int, float)):  # Coluna Preço Unitário
                    cell.value = value
                    cell.number_format = 'R$ #,##0.00'
            
            cell.border = border
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 12,  # ID Compra
        'B': 12,  # NI Fornecedor
        'C': 12,  # Código Item
        'D': 30,  # Descrição
        'E': 12,  # Data Compra
        'F': 12,  # Unidade
        'G': 10,  # Quantidade
        'H': 15,  # Preço Unitário
        'I': 16,  # CNPJ
        'J': 30,  # Fornecedor
        'K': 12,  # Cod. UASG
        'L': 30   # Nome UASG
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar painel (linhas/colunas do cabeçalho)
    ws.freeze_panes = 'A15'
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Função para gerar PDF simples a partir dos dados
def gerar_relatorio_pdf_simples(dataframe, estatisticas, outliers_info, col_precounitario):
    """
    Gera um relatório PDF simples e limpo com tabela completa de dados.
    """
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_margins(8, 8, 8)
    
    # CABEÇALHO
    pdf.set_font("Arial", "B", 14)
    pdf.set_text_color(212, 175, 55)
    pdf.cell(0, 8, "AtaCotada", ln=True, align="C")
    
    pdf.set_font("Arial", "", 9)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 4, "MARINHA DO BRASIL - Pesquisa de Preços", ln=True, align="C")
    
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pdf.cell(0, 4, f"Relatório gerado em: {data_hora}", ln=True, align="C")
    pdf.ln(2)
    
    # TABELA DE ESTATÍSTICAS
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(0, 26, 77)
    pdf.set_text_color(255, 255, 255)
    
    col_width_stats = 28
    pdf.cell(col_width_stats, 6, "Preço Mín", 1, 0, "C", True)
    pdf.cell(col_width_stats, 6, "Preço Médio", 1, 0, "C", True)
    pdf.cell(col_width_stats, 6, "Preço Mediano", 1, 0, "C", True)
    pdf.cell(col_width_stats, 6, "Preço Máx", 1, 0, "C", True)
    pdf.cell(col_width_stats, 6, "Desvio Padrão", 1, 0, "C", True)
    pdf.cell(col_width_stats, 6, "Coef. Variação", 1, 1, "C", True)
    
    pdf.set_font("Arial", "", 8)
    pdf.set_text_color(0, 0, 0)
    
    preco_min = dataframe[col_precounitario].min()
    preco_med = dataframe[col_precounitario].mean()
    preco_mediano = dataframe[col_precounitario].median()
    preco_max = dataframe[col_precounitario].max()
    desvio = dataframe[col_precounitario].std()
    coef_var = (desvio / preco_med * 100) if preco_med != 0 else 0
    
    pdf.cell(col_width_stats, 6, f"R$ {formatar_moeda_br(preco_min)}", 1, 0, "C")
    pdf.cell(col_width_stats, 6, f"R$ {formatar_moeda_br(preco_med)}", 1, 0, "C")
    pdf.cell(col_width_stats, 6, f"R$ {formatar_moeda_br(preco_mediano)}", 1, 0, "C")
    pdf.cell(col_width_stats, 6, f"R$ {formatar_moeda_br(preco_max)}", 1, 0, "C")
    pdf.cell(col_width_stats, 6, f"{desvio:.2f}".replace('.', ','), 1, 0, "C")
    pdf.cell(col_width_stats, 6, f"{coef_var:.2f}%".replace('.', ','), 1, 1, "C")
    
    pdf.ln(2)
    
    # INFORMAÇÕES
    pdf.set_font("Arial", "B", 8)
    total_registros = len(dataframe)
    outliers_removidos = outliers_info['removidos'] if outliers_info else 0
    
    pdf.cell(0, 4, f"Total de Preços Encontrados: {total_registros} | Outliers Removidos: {outliers_removidos}", ln=True)
    pdf.ln(2)
    
    # TABELA DE DADOS
    pdf.set_font("Arial", "B", 8)
    pdf.cell(0, 4, "TABELA DE DADOS", ln=True)
    pdf.ln(1)
    
    # Preparar colunas a serem exibidas (mesmas do Excel)
    colunas_mapa = {
        'idCompra': 'ID Compra',
        'niFornecedor': 'NI Fornecedor',
        'codigoItemCatalogo': 'Código Item',
        'descricaoItem': 'Descrição',
        'dataCompra': 'Data Compra',
        'unidadeFornecimento': 'Unidade',
        'quantidade': 'Quantidade',
        'precoUnitario': 'Preço Unit.',
        'cnpj': 'CNPJ',
        'nomeFornecedor': 'Fornecedor',
        'codigoUasg': 'Cod. UASG',
        'nomeUasg': 'Nome UASG'
    }
    
    colunas_encontradas = {}
    for col_esperada, label in colunas_mapa.items():
        col_real = encontrar_coluna(dataframe, [col_esperada, col_esperada.lower()])
        if col_real:
            colunas_encontradas[col_real] = label
    
    if colunas_encontradas:
        # Calcular largura das colunas dinamicamente
        # Largura disponível: A4 landscape = 297mm, menos margens (8+8=16) = 281mm
        largura_disponivel = 297 - 16
        num_colunas = len(colunas_encontradas)
        col_width_dados = largura_disponivel / num_colunas
        
        # Cabeçalho da tabela de dados
        pdf.set_font("Arial", "B", 7)
        pdf.set_fill_color(0, 26, 77)
        pdf.set_text_color(255, 255, 255)
        
        y_before_header = pdf.get_y()
        for label in colunas_encontradas.values():
            pdf.cell(col_width_dados, 5, label[:15], 1, 0, "C", True)
        pdf.ln()
        
        # Dados da tabela - TODOS os registros
        pdf.set_font("Arial", "", 6)
        pdf.set_text_color(0, 0, 0)
        
        for idx in range(len(dataframe)):
            # Verificar se precisa adicionar nova página
            if pdf.get_y() > 180:  # Deixar espaço para rodapé
                pdf.add_page()
                # Repetir cabeçalho em nova página
                pdf.set_font("Arial", "B", 7)
                pdf.set_fill_color(0, 26, 77)
                pdf.set_text_color(255, 255, 255)
                
                for label in colunas_encontradas.values():
                    pdf.cell(col_width_dados, 5, label[:15], 1, 0, "C", True)
                pdf.ln()
                
                pdf.set_font("Arial", "", 6)
                pdf.set_text_color(0, 0, 0)
            
            for col_real, label in colunas_encontradas.items():
                valor = dataframe.iloc[idx][col_real]
                if isinstance(valor, float) and col_real == encontrar_coluna(dataframe, ['precoUnitario', 'precunitario']):
                    cell_text = f"R$ {formatar_moeda_br(valor)}"
                else:
                    cell_text = str(valor)[:18] if valor else ""
                pdf.cell(col_width_dados, 5, cell_text, 1, 0, "C")
            pdf.ln()
    
    # RODAPÉ
    pdf.ln(1)
    pdf.set_font("Arial", "", 7)
    pdf.set_text_color(212, 175, 55)
    pdf.cell(0, 4, "Idealizado e Desenvolvido por COpAb - Sobressalentes - V1.2026", ln=True, align="C")
    
    pdf.set_font("Arial", "", 6)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 4, f"Página {pdf.page_no()}", align="C")
    
    pdf_output = pdf.output(dest='S')
    if isinstance(pdf_output, bytearray):
        return bytes(pdf_output)
    return pdf_output

# Funções para consulta de fornecedores via API OpenCNPJ
def buscar_dados_fornecedor(cnpj):
    """Busca dados do fornecedor na API OpenCNPJ"""
    if not cnpj or pd.isna(cnpj):
        return None
    
    try:
        # Remover caracteres especiais do CNPJ
        cnpj_limpo = str(cnpj).replace('.', '').replace('/', '').replace('-', '').strip()
        if len(cnpj_limpo) != 14:
            return None
        
        api_url = f'https://api.opencnpj.org/{cnpj_limpo}'
        response = requests.get(api_url, timeout=5)
        
        if response.status_code == 200:
            dados = response.json()
            # Log para debug (comentado para produção)
            # print(f"DEBUG - Resposta API para {cnpj_limpo}: {dados}")
            return dados
        return None
    except Exception as e:
        # Silenciar exceções para não interromper o fluxo
        return None

def formatar_telefone(telefone):
    """Formata telefone para um padrão legível (21) 98130-7593"""
    if isinstance(telefone, dict):
        # Se é um dicionário com ddd e numero
        ddd = telefone.get('ddd', '').strip() if telefone.get('ddd') else ''
        numero = telefone.get('numero', '').strip() if telefone.get('numero') else ''
        
        if ddd and numero:
            # Formatar como (21) 98130-7593 ou (21) 3130-7593
            if len(numero) >= 8:
                numero_formatado = f"{numero[:4]}-{numero[4:]}"
            else:
                numero_formatado = numero
            return f"({ddd}) {numero_formatado}"
        elif numero:
            return numero
        elif ddd:
            return ddd
        else:
            return str(telefone)
    else:
        # Se é uma string, tentar extrair padrão (21) 98130-7593 se precisar
        telefone_str = str(telefone).strip()
        return telefone_str

def extrair_contatos_fornecedor(dados_fornecedor):
    """Extrai email, telefones, estado e cidade do retorno da API"""
    if not dados_fornecedor:
        return None
    
    try:
        # Variações possíveis para campos de email
        email = ''
        for campo_email in ['email', 'correio_eletronico', 'mail', 'e_mail', 'emailComercial']:
            email = dados_fornecedor.get(campo_email, '')
            if email:
                break
        
        # Variações possíveis para campos de estado/UF
        estado = ''
        for campo_estado in ['state', 'estado', 'uf', 'UF', 'state_code', 'sigla_estado']:
            valor = dados_fornecedor.get(campo_estado, '')
            if valor:
                estado = str(valor).upper().strip()
                break
        
        # Variações possíveis para campos de cidade
        cidade = ''
        for campo_cidade in ['city', 'cidade', 'municipio', 'municipality', 'city_name']:
            valor = dados_fornecedor.get(campo_cidade, '')
            if valor:
                cidade = str(valor).strip()
                break
        
        # Variações possíveis para campos de telefone
        telefones = []
        
        # Buscar em campos individuais
        for campo_tel in ['phone', 'telefone', 'phonePrimary', 'phoneSecondary', 
                         'telefone_comercial', 'telefone_principal', 'ddd_telefone',
                         'ddd_fax', 'fone', 'telephone']:
            valor = dados_fornecedor.get(campo_tel, '')
            if valor:
                # Formatar o telefone se for dicionário
                telefone_formatado = formatar_telefone(valor)
                if telefone_formatado and telefone_formatado not in telefones:
                    telefones.append(telefone_formatado)
        
        # Buscar em arrays
        for campo_array in ['phones', 'telefones', 'phone_numbers']:
            valores = dados_fornecedor.get(campo_array, [])
            if isinstance(valores, list):
                for v in valores:
                    if v:
                        # Formatar o telefone
                        telefone_formatado = formatar_telefone(v)
                        if telefone_formatado and telefone_formatado not in telefones:
                            telefones.append(telefone_formatado)
        
        # Remover duplicatas mantendo ordem
        telefones_unicos = []
        for t in telefones:
            if t not in telefones_unicos:
                telefones_unicos.append(t)
        
        # Formatar localização
        localizacao = ''
        if cidade and estado:
            localizacao = f"{cidade}, {estado}"
        elif cidade:
            localizacao = cidade
        elif estado:
            localizacao = estado
        
        return {
            'email': email if email else 'Não informado',
            'telefones': ', '.join(telefones_unicos) if telefones_unicos else 'Não informado',
            'cidade': cidade if cidade else 'Não informado',
            'estado': estado if estado else 'Não informado',
            'localizacao': localizacao if localizacao else 'Não informado',
            'debug_campos': list(dados_fornecedor.keys())  # Para debug
        }
    except Exception as e:
        return None



def gerar_html_fornecedores(dataframe, descricao_item=None):
    """Gera HTML com tabela de fornecedores e seus contatos"""
    
    # Encontrar as colunas corretas
    col_niFornecedor = encontrar_coluna(dataframe, ['niFornecedor', 'nifornecedor', 'ni_fornecedor'])
    col_cnpj = encontrar_coluna(dataframe, ['cnpj', 'CNPJ'])
    col_nomeFornecedor = encontrar_coluna(dataframe, ['nomeFornecedor', 'nomefornecedor', 'nome_fornecedor', 'fornecedor'])
    col_descricao = encontrar_coluna(dataframe, ['descricaoItem', 'descricao', 'description', 'item_descricao'])
    
    if not col_niFornecedor:
        return "<h3 style='color: red;'>Erro: Coluna 'niFornecedor' não encontrada no dataframe</h3>"
    
    # Se não foi passada descrição, tenta extrair do dataframe
    if not descricao_item and col_descricao and len(dataframe) > 0:
        descricao_item = str(dataframe.iloc[0][col_descricao])
    
    # Obter fornecedores únicos
    fornecedores_unicos = dataframe.drop_duplicates(subset=[col_niFornecedor])
    
    # CSS inline compacto
    css_style = """
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); font-family: 'Arial', sans-serif; padding: 2rem; min-height: 100vh; }
    .container { max-width: 1200px; margin: 0 auto; background: #ffffff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3); overflow: hidden; }
    .header { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); color: #ffffff; padding: 2rem; text-align: center; }
    .header h1 { color: #d4af37; font-size: 36px; margin-bottom: 0.5rem; letter-spacing: 2px; }
    .header p { font-size: 14px; color: #ffffff; }
    .header .descricao-item { font-size: 10px; color: #d4af37; letter-spacing: 0.5px; margin-top: 0.5rem; font-style: italic; opacity: 0.8; }
    .content { padding: 2rem; }
    .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 1rem; margin-bottom: 2rem; }
    .stat-box { background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center; }
    .stat-box label { color: #d4af37; font-weight: bold; font-size: 12px; letter-spacing: 1px; }
    .stat-box .value { color: #ffffff; font-size: 28px; font-weight: bold; margin-top: 0.5rem; }
    table { width: 100%; border-collapse: collapse; margin-top: 1rem; }
    thead { background: #001a4d; color: #ffffff; }
    th { padding: 1rem; text-align: left; font-weight: bold; border-bottom: 2px solid #d4af37; }
    td { padding: 0.75rem 1rem; border-bottom: 1px solid #e0e0e0; }
    tbody tr:hover { background: #f5f5f5; }
    tbody tr:nth-child(even) { background: #f9f9f9; }
    .contato-info { font-size: 13px; color: #333; }
    .email { color: #0033cc; text-decoration: none; }
    .email:hover { text-decoration: underline; }
    .phone { color: #006600; }
    .footer { background: #f0f0f0; padding: 1rem 2rem; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }
    """
    
    # Coletar dados de cada fornecedor
    linhas = []
    com_email_count = 0
    com_telefone_count = 0
    
    for idx, row in fornecedores_unicos.iterrows():
        ni_fornecedor = row.get(col_niFornecedor, '')
        cnpj = row.get(col_cnpj, '') if col_cnpj else ''
        nome_fornecedor = row.get(col_nomeFornecedor, 'Não informado') if col_nomeFornecedor else 'Não informado'
        
        # Buscar dados na API usando niFornecedor
        dados_fornecedor = buscar_dados_fornecedor(ni_fornecedor)
        contatos = extrair_contatos_fornecedor(dados_fornecedor) if dados_fornecedor else None
        
        if contatos:
            email = contatos['email']
            telefones = contatos['telefones']
            localizacao = contatos['localizacao']
            
            if email and email != 'Não informado':
                com_email_count += 1
            if telefones and telefones != 'Não informado':
                com_telefone_count += 1
        else:
            email = 'Não informado'
            telefones = 'Não informado'
            localizacao = 'Não informado'
        
        # Formatar email como link
        email_html = f'<a href="mailto:{email}" class="email">{email}</a>' if email and email != 'Não informado' else 'Não informado'
        
        linha = f"<tr><td>{ni_fornecedor}</td><td>{cnpj}</td><td>{nome_fornecedor}</td><td>{localizacao}</td><td><span class='contato-info'>{email_html}</span></td><td><span class='contato-info phone'>{telefones}</span></td></tr>"
        linhas.append(linha)
    
    linhas_tabela = ''.join(linhas)
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    html_completo = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Contatos de Fornecedores - AtaCotada</title>
    <style>{css_style}</style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>AtaCotada</h1>
            <p>Contatos de Fornecedores - Marinha do Brasil</p>
            {f'<div class="descricao-item">Fornecedores de: {descricao_item[:100]}</div>' if descricao_item else ''}
        </div>
        <div class="content">
            <div class="stats">
                <div class="stat-box">
                    <label>TOTAL DE FORNECEDORES</label>
                    <div class="value">{len(fornecedores_unicos)}</div>
                </div>
                <div class="stat-box">
                    <label>COM EMAIL</label>
                    <div class="value">{com_email_count}</div>
                </div>
                <div class="stat-box">
                    <label>COM TELEFONE</label>
                    <div class="value">{com_telefone_count}</div>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>NI Fornecedor</th>
                        <th>CNPJ</th>
                        <th>Fornecedor</th>
                        <th>Localização</th>
                        <th>Email</th>
                        <th>Telefones</th>
                    </tr>
                </thead>
                <tbody>
                    {linhas_tabela}
                </tbody>
            </table>
        </div>
        <div class="footer">
            <p>Relatório gerado por AtaCotada em {data_hora}</p>
        </div>
    </div>
</body>
</html>"""
    
    return html_completo

# URLs atualizadas
consultarItemMaterial_base_url = 'https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/1_consultarMaterial'
consultarItemServico_base_url = 'https://dadosabertos.compras.gov.br/modulo-pesquisa-preco/3_consultarServico'

def obter_itens(tipo_item, codigo_item_catalogo, pagina, tamanho_pagina):
    url = consultarItemMaterial_base_url if tipo_item == 'Material' else consultarItemServico_base_url
    params = {
        'pagina': pagina,   
        'tamanhoPagina':tamanho_pagina,  # Ajuste para 500 itens por página
        'codigoItemCatalogo': codigo_item_catalogo
    }
    try:
        response = requests.get(url, params=params)
        if response.status_code == 200:
            json_response = response.json()
            itens = json_response.get('resultado', [])
            paginas_restantes = json_response.get('paginasRestantes', 0)
            total_paginas = json_response.get('totalPaginas', 0)
            return itens, paginas_restantes, total_paginas
        else:
            st.error(f"Erro na consulta: {response.status_code}")
            return [], 0
    except Exception as e:
        st.error(f"Erro ao realizar a requisição: {str(e)}")
        return [], 0

# Streamlit UI
# Header customizado
col_header1, col_header2, col_header3 = st.columns([1, 2, 1])
with col_header2:
    st.markdown("""
        <div class="header-container">
            <div class="logo-text">⚓ MARINHA DO BRASIL ⚓</div>
            <div class="sistema-nome">AtaCotada</div>
            <div class="subtitulo">PESQUISA DE PREÇOS DE MATERIAIS E SERVIÇOS</div>
        </div>
    """, unsafe_allow_html=True)

# Disclaimer com estilo
st.markdown("""
    <div style="background-color: #0a2540; padding: 1rem; border-radius: 8px; border-left: 5px solid #d4af37; margin-bottom: 1.5rem;">
        <p style="color: #d4af37; font-weight: bold; margin-bottom: 0.5rem;">📚 Orientação</p>
        <p style="color: #ffffff;">Utilize o catálogo de materiais e serviços do Governo Federal para encontrar o código desejado.</p>
    </div>
""", unsafe_allow_html=True)

components.iframe("https://catalogo.compras.gov.br/cnbs-web/busca", height=450, width=1080, scrolling=True)

# Seção de consulta
st.markdown("""
    <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
        🔍 Consultar Item
    </div>
""", unsafe_allow_html=True)

col_tipo, col_codigo, col_botao = st.columns([1, 2, 1])

with col_tipo:
    tipo_item = st.selectbox("Tipo de Item", ['Material', 'Serviço'], key='tipo_item')

with col_codigo:
    codigo_item_catalogo = st.text_input("Código do Item de Catálogo", value="", key='codigo_item_catalogo', placeholder="Ex: 123456")

with col_botao:
    st.write("")
    st.write("")
    consultar = st.button('🔎 Consultar', use_container_width=True)

pagina = 1
tamanho_pagina = 500

if consultar:
    if codigo_item_catalogo:  # Verifica se o código do item de catálogo não está vazio
        itens, paginas_restantes, total_paginas = obter_itens(tipo_item, codigo_item_catalogo, pagina, tamanho_pagina)
        if itens:  # Ensure 'itens' is not empty before proceeding
            st.session_state['itens'] = itens      
        else:
            st.error("Nenhum item encontrado. Por favor, tente com um código diferente ou verifique a conexão com a API.")
    else:
        st.warning("Por favor, informe o código do item de catálogo para realizar a consulta.")

# Mostrar filtros e resultados se houver dados em session_state
if st.session_state.get('itens'):
    try:
        # Ensure 'itens' is in the expected format before normalization
        if isinstance(st.session_state['itens'], list) and all(isinstance(item, dict) for item in st.session_state['itens']):
            df_completo = pd.json_normalize(st.session_state['itens'])
            
            # Encontrar colunas com nomes variáveis
            col_unidade = encontrar_coluna(df_completo, ['unidadeFornecimento', 'unidade_fornecimento', 'unidade', 'unitFornecimento'])
            col_nome_unidade = encontrar_coluna(df_completo, ['nomeUnidadeFornecimento', 'nome_unidade_fornecimento', 'nomeUnidade', 'nome_unidade'])
            col_estado = encontrar_coluna(df_completo, ['uf', 'estado', 'estadoUf', 'estado_uf'])
            col_precounitario = encontrar_coluna(df_completo, ['precoUnitario', 'preco_unitario', 'preço_unitário'])
            col_uasg = encontrar_coluna(df_completo, ['nomeUasg', 'nome_uasg', 'uasg', 'nomeUAsg'])
            
            # Seção de Filtros
            st.markdown("""
                <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
                    ⚙️ Filtros de Consulta
                </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            # Inicializar variáveis de filtro
            unidade_selecionada = []
            nome_unidade_input = ""
            estado_selecionado = []
            uasg_input = ""
            remover_outliers = False
            
            with col1:
                # Filtro de unidade de fornecimento (se a coluna existir)
                if col_unidade:
                    unidades_disponiveis = df_completo[col_unidade].dropna().unique().tolist()
                    if unidades_disponiveis:
                        unidade_selecionada = st.multiselect(
                            "Unidade de Fornecimento",
                            options=sorted(unidades_disponiveis),
                            default=unidades_disponiveis,
                            key='filtro_unidade'
                        )
                
                # Filtro de nome unidade fornecimento (se a coluna existir)
                if col_nome_unidade:
                    nome_unidade_input = st.text_input(
                        "Nome Unidade Fornecimento",
                        value="",
                        key='filtro_nome_unidade',
                        placeholder="Filtrar por nome..."
                    )
                
                # Filtro de nome UASG (se a coluna existir) - ABAIXO DO FILTRO DA UNIDADE
                if col_uasg:
                    uasg_input = st.text_input(
                        "Nome UASG",
                        value="",
                        key='filtro_uasg',
                        placeholder="Filtrar por nome..."
                    )
            
            with col2:
                # Filtro de estado (se a coluna existir)
                if col_estado:
                    estados_disponiveis = df_completo[col_estado].dropna().unique().tolist()
                    if estados_disponiveis:
                        estado_selecionado = st.multiselect(
                            "Estado (UF)",
                            options=sorted(estados_disponiveis),
                            default=estados_disponiveis,
                            key='filtro_estado'
                        )
                
                # Checkbox para remover outliers
                if col_precounitario:
                    remover_outliers = st.checkbox(
                        "🔍 Remover outliers (IQR)",
                        value=False,
                        key='filtro_outliers'
                    )
                    st.write("")
            
            # Aplicar filtros iniciais (sem preço ainda)
            dataframe = df_completo.copy()
            
            # Filtro de unidade de fornecimento
            if col_unidade and unidade_selecionada:
                dataframe = dataframe[dataframe[col_unidade].isin(unidade_selecionada)]
            
            # Filtro de nome unidade fornecimento
            if col_nome_unidade and nome_unidade_input:
                dataframe = dataframe[dataframe[col_nome_unidade].str.contains(nome_unidade_input, case=False, na=False)]
            
            # Filtro de estado
            if col_estado and estado_selecionado:
                dataframe = dataframe[dataframe[col_estado].isin(estado_selecionado)]
            
            # Filtro de nome UASG
            if col_uasg and uasg_input:
                dataframe = dataframe[dataframe[col_uasg].str.contains(uasg_input, case=False, na=False)]
            
            # Remover outliers se selecionado (ANTES do slider de preço)
            outliers_info = None
            if remover_outliers and col_precounitario and col_precounitario in dataframe.columns:
                dataframe_temp, outliers_removidos, limite_inf, limite_sup = remover_outliers_iqr(dataframe, col_precounitario)
                outliers_info = {
                    'removidos': outliers_removidos,
                    'limite_inf': limite_inf,
                    'limite_sup': limite_sup
                }
                dataframe = dataframe_temp
            
            # Agora mostrar o slider de preço com os limites corretos (após remover outliers se necessário)
            faixa_preco = (0, 0)
            if col_precounitario and col_precounitario in dataframe.columns and len(dataframe) > 0:
                preco_min_filtrado = dataframe[col_precounitario].min()
                preco_max_filtrado = dataframe[col_precounitario].max()
                
                faixa_preco = st.slider(
                    "Faixa de Preço Unitário (R$)",
                    min_value=float(preco_min_filtrado),
                    max_value=float(preco_max_filtrado),
                    value=(float(preco_min_filtrado), float(preco_max_filtrado)),
                    step=0.01,
                    key='filtro_preco'
                )
                
                # Mostrar valores selecionados em formato brasileiro
                st.markdown(f"""
                    <div style="background-color: #0a2540; padding: 0.8rem; border-radius: 6px; border-left: 4px solid #d4af37; margin-top: 0.5rem;">
                        <span style="color: #d4af37; font-weight: bold;">Intervalo selecionado:</span>
                        <span style="color: #ffffff;"> {formatar_preco_reais(faixa_preco[0])} até {formatar_preco_reais(faixa_preco[1])}</span>
                    </div>
                """, unsafe_allow_html=True)
            
            # Aplicar filtro de preço
            if col_precounitario and col_precounitario in dataframe.columns:
                dataframe = dataframe[
                    (dataframe[col_precounitario] >= faixa_preco[0]) &
                    (dataframe[col_precounitario] <= faixa_preco[1])
                ]
            
            # Mostrar estatísticas apenas dos dados filtrados
            if len(dataframe) > 0:
                # Seção de resultados
                st.markdown("""
                    <div style="color: #d4af37; font-size: 20px; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem;">
                        📊 Resultados
                    </div>
                """, unsafe_allow_html=True)
                
                # Mostrar contagem com destaque
                col_info_a, col_info_b, col_btn_excel, col_btn_pdf = st.columns([1, 1, 0.6, 0.6])
                with col_info_a:
                    st.markdown(f"""
                        <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                            <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">REGISTROS ENCONTRADOS</div>
                            <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">{len(dataframe)}</div>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Mostrar informações sobre outliers removidos
                if outliers_info:
                    with col_info_b:
                        st.markdown(f"""
                            <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                                <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">OUTLIERS REMOVIDOS</div>
                                <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">{outliers_info['removidos']}</div>
                            </div>
                        """, unsafe_allow_html=True)
                
                # Botões de download
                if col_precounitario and col_precounitario in dataframe.columns:
                    mean = dataframe[col_precounitario].mean()
                    median = dataframe[col_precounitario].median()
                    preco_min = dataframe[col_precounitario].min()
                    preco_max = dataframe[col_precounitario].max()
                    std = dataframe[col_precounitario].std()
                    cv = ((std / mean) * 100) if mean != 0 else 0
                    
                    estatisticas = {
                        'min': preco_min,
                        'mean': mean,
                        'median': median,
                        'max': preco_max,
                        'std': std,
                        'cv': cv
                    }
                    
                    # Gerar Excel
                    excel_bytes = gerar_relatorio_excel(dataframe, estatisticas, outliers_info, col_precounitario)
                    data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename_excel = f"relatorio_AtaCotada_{data_hora}.xlsx"
                    
                    with col_btn_excel:
                        st.download_button(
                            label="📊 Excel",
                            data=excel_bytes,
                            file_name=filename_excel,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # Gerar PDF
                    pdf_bytes = gerar_relatorio_pdf_simples(dataframe, estatisticas, outliers_info, col_precounitario)
                    filename_pdf = f"relatorio_AtaCotada_{data_hora}.pdf"
                    
                    with col_btn_pdf:
                        st.download_button(
                            label="📄 PDF",
                            data=pdf_bytes,
                            file_name=filename_pdf,
                            mime="application/pdf",
                            use_container_width=True
                        )
                    
                    # Botão centralizado para consultar fornecedores
                    st.markdown("<br>", unsafe_allow_html=True)
                    col_fornecedores = st.columns([1, 1, 1])
                    with col_fornecedores[1]:
                        if st.button("📞 Consultar Fornecedores", use_container_width=True, key="btn_fornecedores"):
                            with st.spinner("Consultando dados de fornecedores..."):
                                # Extrair descrição do item se disponível
                                col_descricao = encontrar_coluna(dataframe, ['descricaoItem', 'descricao', 'description', 'item_descricao'])
                                descricao_item = None
                                if col_descricao and len(dataframe) > 0:
                                    descricao_item = str(dataframe.iloc[0][col_descricao])
                                
                                html_fornecedores = gerar_html_fornecedores(dataframe, descricao_item)
                                html_bytes = html_fornecedores.encode('utf-8')
                                html_filename = f"fornecedores_AtaCotada_{data_hora}.html"
                                
                                st.download_button(
                                    label="📋 Abrir Relatório de Fornecedores",
                                    data=html_bytes,
                                    file_name=html_filename,
                                    mime="text/html",
                                    use_container_width=True,
                                    key="btn_download_fornecedores"
                                )
                
                # Mostrar informações sobre outliers removidos quando não houver
                if not outliers_info:
                    with col_info_b:
                        st.markdown(f"""
                            <div style="background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center;">
                                <div style="color: #d4af37; font-size: 14px; font-weight: bold; letter-spacing: 1px;">OUTLIERS REMOVIDOS</div>
                                <div style="color: #ffffff; font-size: 32px; font-weight: bold; margin-top: 0.5rem;">0</div>
                            </div>
                        """, unsafe_allow_html=True)
                
                # Mostrar faixa de outliers
                if outliers_info:
                    st.markdown(f"""
                        <div style="background-color: #0a2540; padding: 1rem; border-radius: 8px; border-left: 5px solid #d4af37; margin-bottom: 1.5rem;">
                            <span style="color: #d4af37; font-weight: bold;">📊 Faixa válida (sem outliers):</span><br>
                            <span style="color: #ffffff;">{formatar_preco_reais(outliers_info['limite_inf'])} até {formatar_preco_reais(outliers_info['limite_sup'])}</span>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Mostrar estatísticas se a coluna precoUnitario existir
                if col_precounitario and col_precounitario in dataframe.columns:
                    mean = dataframe[col_precounitario].mean()
                    median = dataframe[col_precounitario].median()
                    preco_min = dataframe[col_precounitario].min()
                    preco_max = dataframe[col_precounitario].max()
                    std = dataframe[col_precounitario].std()
                    cv = ((std / mean) * 100) if mean != 0 else 0
                    
                    st.markdown(
                        f"""
                            |Preço Mínimo|Preço Unitário Médio|Preço Unitário Mediano|Preço Máximo|Desvio Padrão|Coeficiente de Variação|
                            |:----------:|:------------------:|:--------------------:|:----------:|:-----------:|:---------------------:|
                            |**{formatar_preco_reais(preco_min)}**|**{formatar_preco_reais(mean)}**|**{formatar_preco_reais(median)}**|**{formatar_preco_reais(preco_max)}**|**{formatar_numero_br(std)}**|**{formatar_numero_br(cv)}%**|
                        """
                    )
                    
                    # Formatar preço para exibição
                    df_exibicao = dataframe.copy()
                    df_exibicao[col_precounitario] = df_exibicao[col_precounitario].apply(formatar_preco_reais)
                    st.write(df_exibicao)
                else:
                    st.write(dataframe)
            else:
                st.warning("Nenhum resultado encontrado com os filtros aplicados. Tente ajustar os critérios.")
                     
        else:
            st.error("Formato dos itens inválido para normalização.")
    except Exception as e:
        st.error(f"Erro ao processar os itens: {str(e)}")
        # Mostrar colunas disponíveis quando há erro
        if st.session_state.get('itens'):
            try:
                df_debug = pd.json_normalize(st.session_state['itens'])
                with st.expander("ℹ️ Colunas disponíveis (Debug)"):
                    st.write("Colunas encontradas no dataframe:")
                    st.write(list(df_debug.columns))
                    st.write("\nPrimeiras linhas dos dados:")
                    st.write(df_debug.head())
            except:
                pass

# Rodapé
st.markdown("""
    <div style="margin-top: 3rem; padding-top: 2rem; border-top: 1px solid #0033cc; text-align: center;">
        <p style="color: #d4af37; font-size: 12px; margin: 0; letter-spacing: 0.5px;">
            Idealizado e Desenvolvido por COpAb - Sobressalentes - V1.2026
        </p>
    </div>
""", unsafe_allow_html=True)
