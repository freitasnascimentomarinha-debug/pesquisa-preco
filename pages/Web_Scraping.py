import streamlit as st
import pandas as pd
import time
import random
import re
import os
import base64
import json
from datetime import datetime
from io import BytesIO
from urllib.parse import urlparse, quote_plus

# Configuração da página
st.set_page_config(
    page_title="AtaCotada - Web Scraping",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# CSS customizado (mesmo padrão das outras páginas)
st.markdown("""
    <style>
        html, body, [data-testid="stAppViewContainer"],
        .main, [data-testid="stApp"], .stApp {
            background-color: #001a4d !important;
            color: #ffffff !important;
        }

        .stApp { animation: fadeIn 0.3s ease-in; }
        @keyframes fadeIn { from { opacity: 0.7; } to { opacity: 1; } }

        .header-container {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            padding: 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
            text-align: center;
        }
        .logo-text { color: #ffffff; font-size: 14px; font-weight: 600; letter-spacing: 2px; margin-bottom: 0.5rem; }
        .sistema-nome {
            color: #d4af37; font-size: 48px; font-weight: bold; letter-spacing: 3px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5); margin: 0.5rem 0; font-family: 'Arial Black', sans-serif;
        }
        .subtitulo { color: #ffffff; font-size: 14px; margin-top: 0.5rem; letter-spacing: 1px; }

        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0a0a0a 0%, #111111 50%, #0a0a0a 100%) !important;
            border-right: 3px solid #d4af37 !important;
            box-shadow: 4px 0 15px rgba(0, 0, 0, 0.5);
        }
        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { background: transparent !important; }
        [data-testid="stSidebarNav"] { display: none !important; }

        [data-testid="stSidebar"] .stMarkdown h2 {
            color: #d4af37 !important;
            font-family: 'Arial Black', sans-serif;
            font-size: 22px;
            text-align: center;
            letter-spacing: 2px;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.8);
            border-bottom: 2px solid #d4af37;
            padding-bottom: 0.75rem;
            margin-bottom: 1.5rem;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] {
            background: linear-gradient(135deg, #1a1a1a 0%, #252525 100%) !important;
            color: #ffffff !important;
            border: 1px solid #333333 !important;
            border-radius: 10px !important;
            margin: 0.4rem 0 !important;
            padding: 0.85rem 1.2rem !important;
            font-weight: 600 !important;
            font-size: 15px !important;
            transition: all 0.3s ease !important;
            text-decoration: none !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] span { color: #ffffff !important; }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover {
            background: linear-gradient(135deg, #252525 0%, #353535 100%) !important;
            border: 1px solid #d4af37 !important;
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.25) !important;
        }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover span { color: #d4af37 !important; }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] {
            background: linear-gradient(135deg, #d4af37 0%, #c5a028 100%) !important;
            color: #0a0a0a !important;
            border: 1px solid #d4af37 !important;
            font-weight: bold !important;
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.4) !important;
        }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] span { color: #0a0a0a !important; }
        [data-testid="stSidebar"] label, [data-testid="stSidebar"] .stText, [data-testid="stSidebar"] p { color: #ffffff !important; }
        [data-testid="stSidebar"] hr { border-color: #333333 !important; margin: 1rem 0 !important; }

        .info-card {
            background: rgba(0, 26, 77, 0.6);
            border: 1px solid #d4af37;
            border-radius: 12px;
            padding: 1.5rem;
            box-shadow: 0 10px 30px rgba(0,0,0,0.35);
            margin-bottom: 1rem;
        }
        .info-title { color: #d4af37; font-size: 1.2rem; font-weight: bold; margin-bottom: 1rem; border-bottom: 1px solid rgba(212, 175, 55, 0.3); padding-bottom: 0.5rem; }

        .stButton > button[kind="primary"], .stButton > button[data-testid="stBaseButton-primary"] {
            background-color: #d4af37 !important; color: #ffffff !important; border: none !important; font-weight: bold !important;
        }
        .stButton > button[kind="primary"]:hover, .stButton > button[data-testid="stBaseButton-primary"]:hover {
            background-color: #c5a028 !important; color: #ffffff !important;
        }

        .stTabs [data-baseweb="tab-list"] { gap: 8px; }
        .stTabs [data-baseweb="tab"] {
            background-color: rgba(0, 26, 77, 0.6);
            border-radius: 8px 8px 0 0;
            padding: 10px 20px;
            color: #ffffff;
            font-weight: 600;
        }
        .stTabs [aria-selected="true"] {
            background-color: #d4af37 !important;
            color: #0a0a0a !important;
        }

        [data-testid="stDataFrame"] {
            background-color: #ffffff !important;
            border-radius: 8px;
            overflow: hidden;
        }
        th { background-color: #ffffff !important; color: #333333 !important; font-weight: bold; border-bottom: 2px solid #d4af37 !important; }
        td { background-color: #ffffff !important; color: #333333 !important; }

        .log-container {
            background: #0a0a0a;
            border: 1px solid #333;
            border-radius: 8px;
            padding: 1rem;
            font-family: 'Courier New', monospace;
            font-size: 13px;
            color: #00ff00;
            max-height: 400px;
            overflow-y: auto;
        }
        .log-info { color: #00ff00; }
        .log-warn { color: #ffaa00; }
        .log-error { color: #ff4444; }
        .log-success { color: #44ff44; font-weight: bold; }
        .log-orcamento { color: #1a1a1a; background: #87CEEB; font-weight: bold; padding: 2px 6px; border-radius: 4px; display: inline-block; margin: 1px 0; }

        .screenshot-container {
            border: 2px solid #d4af37;
            border-radius: 8px;
            overflow: hidden;
            margin: 0.5rem 0;
        }

        .sidebar-footer {
            color: #666666;
            font-size: 11px;
            text-align: center;
            padding: 1rem 0;
            border-top: 1px solid #333333;
            margin-top: 2rem;
        }
    </style>
""", unsafe_allow_html=True)


# ===================== CONSTANTES =====================

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
]

VARIANTES_BUSCA = [
    "{item} preço brasil",
    "{item} comprar brasil",
    "{item} fornecedor brasil",
    "comprar {item} online brasil",
    "{item} valor unitário loja brasileira",
    "{item} loja online brasil",
]

# Domínios a ignorar nos resultados
DOMINIOS_IGNORADOS = [
    "google.com", "google.com.br", "youtube.com", "facebook.com",
    "instagram.com", "twitter.com", "linkedin.com", "wikipedia.org",
    "gov.br", "reddit.com", "tiktok.com",
    # Marketplaces
    "mercadolivre.com.br", "mercadolivre.com", "lista.mercadolivre.com.br",
    "produto.mercadolivre.com.br", "mlstatic.com",
    "magazineluiza.com.br", "magalu.com.br",
    "shopee.com.br", "shopee.com",
    "amazon.com.br", "amazon.com",
    "aliexpress.com", "aliexpress.com.br",
    "casasbahia.com.br", "pontofrio.com.br",
    "submarino.com.br", "americanas.com.br",
    "kabum.com.br", "zoom.com.br",
    "buscape.com.br", "bondfaro.com.br",
    # Streamlit / app próprio
    "streamlit.app", "streamlit.io", "share.streamlit.io",
    # Blogs, notícias, comparadores, fóruns
    "blog.", "medium.com", "blogspot.com", "wordpress.com",
    "noticias.", "uol.com.br", "globo.com", "g1.globo.com",
    "folha.uol.com.br", "terra.com.br", "ig.com.br",
    "reclameaqui.com.br", "jusbrasil.com.br",
    "slideshare.net", "scribd.com", "pinterest.com",
    "quora.com", "stackoverflow.com", "github.com",
    "comparador.", "versus.com", "techtudo.com.br",
    "tudocelular.com", "canaltech.com.br", "tecmundo.com.br",
    "olx.com.br", "enjoei.com.br",
    # Específicos
    "forneceb2b.com", "forneceb2b.com.br",
    # Domínios de Portugal e Europa
    ".pt", ".es", ".fr", ".de", ".it", ".uk", ".eu",
]

MAX_FONTES_POR_ITEM = 3
MAX_RETRIES = 2
SCREENSHOT_DIR = "/tmp/scraping_screenshots"

# SearchAPI.io (Google Shopping) — fallback quando DDGS falha
SEARCHAPI_KEY = "wZb2W9zvLh3gPziTp2639VCr"
SEARCHAPI_URL = "https://www.searchapi.io/api/v1/search"


# ===================== FUNÇÕES AUXILIARES =====================

def gerar_delay(min_seg=2.0, max_seg=6.0):
    """Delay aleatório entre requisições para simular comportamento humano."""
    return random.uniform(min_seg, max_seg)


def gerar_delay_leitura(min_seg=1.5, max_seg=4.0):
    """Delay de permanência na página simulando leitura."""
    return random.uniform(min_seg, max_seg)


def escolher_user_agent():
    """Seleciona um User-Agent aleatório."""
    return random.choice(USER_AGENTS)


def gerar_headers(user_agent=None):
    """Gera headers realistas de navegador."""
    ua = user_agent or escolher_user_agent()
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.5,en;q=0.3",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
    }


def dominio_valido(url):
    """Verifica se o domínio não está na lista de ignorados e é brasileiro."""
    try:
        dominio = urlparse(url).netloc.lower()
        # Rejeitar domínios na lista de ignorados
        if any(d in dominio for d in DOMINIOS_IGNORADOS):
            return False
        # Aceitar apenas domínios brasileiros (.com.br, .br) ou .com genéricos
        if dominio.endswith('.br') or dominio.endswith('.com') or dominio.endswith('.net') or dominio.endswith('.org'):
            return True
        return False
    except Exception:
        return False


def extrair_dominio(url):
    """Extrai domínio limpo de uma URL."""
    try:
        return urlparse(url).netloc
    except Exception:
        return url


def log_msg(log_container, logs, msg, nivel="info"):
    """Adiciona mensagem de log e atualiza o container."""
    timestamp = datetime.now().strftime("%H:%M:%S")
    classe = f"log-{nivel}"
    logs.append(f'<span class="{classe}">[{timestamp}] {msg}</span>')
    log_container.markdown(
        '<div class="log-container">' + "<br>".join(logs[-50:]) + "</div>",
        unsafe_allow_html=True,
    )


def formatar_moeda_br(valor):
    """Formata valor numérico para moeda brasileira."""
    try:
        val = float(valor)
        return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "N/A"


# ===================== SCRAPING COM REQUESTS + BS4 =====================

def _dedup_urls(urls, num_results=8):
    """Remove duplicatas mantendo ordem e diversidade de domínios."""
    seen = set()
    unique = []
    for u in urls:
        dom = extrair_dominio(u)
        if dom not in seen:
            seen.add(dom)
            unique.append(u)
    return unique[:num_results]


def buscar_ddgs_api(query, num_results=8):
    """Busca usando o pacote ddgs (DuckDuckGo Search) — mais confiável em servidores."""
    try:
        from ddgs import DDGS
        results = list(DDGS().text(query, region="br-pt", max_results=num_results))
        urls = [r["href"] for r in results if r.get("href") and dominio_valido(r["href"])]
        return _dedup_urls(urls, num_results)
    except ImportError:
        try:
            from duckduckgo_search import DDGS
            with DDGS() as ddgs:
                results = list(ddgs.text(query, region="br-pt", max_results=num_results))
            urls = [r["href"] for r in results if r.get("href") and dominio_valido(r["href"])]
            return _dedup_urls(urls, num_results)
        except Exception:
            return []
    except Exception:
        return []


def buscar_duckduckgo(session, query, headers, num_results=8):
    """Busca no DuckDuckGo HTML usando requests e retorna lista de URLs."""
    from bs4 import BeautifulSoup
    from urllib.parse import unquote

    url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}"
    try:
        resp = session.get(url, headers=headers, timeout=15)
        if resp.status_code != 200:
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        urls = []

        for a_tag in soup.select("a.result__a"):
            href = a_tag.get("href", "")
            if href.startswith("//duckduckgo.com/l/?uddg="):
                real_url = unquote(href.split("uddg=")[1].split("&")[0])
            elif href.startswith("http"):
                real_url = href
            else:
                continue
            if real_url.startswith("http") and dominio_valido(real_url):
                urls.append(real_url)

        # Fallback: links dentro de .result__body ou .result
        if not urls:
            for a_tag in soup.select(".result a[href^='http']"):
                href = a_tag.get("href", "")
                if dominio_valido(href):
                    urls.append(href)

        return _dedup_urls(urls, num_results)

    except Exception:
        return []


def buscar_google_requests(session, query, headers, num_results=8):
    """Busca no Google usando requests (fallback)."""
    from bs4 import BeautifulSoup
    from urllib.parse import unquote

    url = f"https://www.google.com.br/search?q={quote_plus(query)}&hl=pt-BR&num={num_results}"
    google_headers = dict(headers)
    google_headers["Referer"] = "https://www.google.com.br/"
    google_headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    try:
        resp = session.get(url, headers=google_headers, timeout=15)
        if resp.status_code != 200:
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        urls = []

        for a_tag in soup.find_all("a", href=True):
            href = a_tag["href"]
            if href.startswith("/url?q="):
                real_url = unquote(href.split("/url?q=")[1].split("&")[0])
                if real_url.startswith("http") and dominio_valido(real_url):
                    urls.append(real_url)

        if not urls:
            for a_tag in soup.select("a[href^='http']"):
                href = a_tag.get("href", "")
                if href.startswith("http") and dominio_valido(href):
                    urls.append(href)

        return _dedup_urls(urls, num_results)

    except Exception:
        return []


def buscar_bing_requests(session, query, headers, num_results=8):
    """Busca no Bing como fallback adicional."""
    from bs4 import BeautifulSoup

    url = f"https://www.bing.com/search?q={quote_plus(query)}&setlang=pt-BR&count={num_results}"
    bing_headers = dict(headers)
    bing_headers["Referer"] = "https://www.bing.com/"
    try:
        resp = session.get(url, headers=bing_headers, timeout=15)
        if resp.status_code != 200:
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        urls = []

        for li in soup.select("li.b_algo"):
            a_tag = li.select_one("h2 a")
            if a_tag:
                href = a_tag.get("href", "")
                if href.startswith("http") and dominio_valido(href):
                    urls.append(href)

        if not urls:
            for a_tag in soup.select("#b_results a[href^='http']"):
                href = a_tag.get("href", "")
                if href.startswith("http") and dominio_valido(href):
                    urls.append(href)

        return _dedup_urls(urls, num_results)

    except Exception:
        return []


def buscar_searchapi(query, num_results=8):
    """Busca usando SearchAPI.io (Google Shopping) — fallback pago e confiável.
    Retorna resultados diretos com preços já extraídos."""
    import requests as req

    if not SEARCHAPI_KEY:
        return []

    try:
        params = {
            "engine": "google_shopping",
            "q": query,
            "api_key": SEARCHAPI_KEY,
            "location": "Brazil",
            "gl": "br",
            "hl": "pt",
            "num": num_results,
        }
        resp = req.get(SEARCHAPI_URL, params=params, timeout=20)
        if resp.status_code != 200:
            return []

        data = resp.json()
        shopping_results = data.get("shopping_results", [])

        resultados = []
        for item in shopping_results:
            preco = item.get("extracted_price")
            titulo = item.get("title", "")
            seller = item.get("seller", "Google Shopping")
            thumbnail = item.get("thumbnail", "")
            product_link = item.get("product_link", "")

            if preco and preco > 0 and titulo:
                resultados.append({
                    "titulo": titulo,
                    "preco": preco,
                    "url": product_link,
                    "dominio": seller,
                    "thumbnail": thumbnail,
                    "fonte": "Google Shopping (SearchAPI)",
                })

        return resultados[:num_results]

    except Exception:
        return []


def buscar_urls(session, query, headers, num_results=8):
    """Busca combinada: DDGS API > DuckDuckGo HTML > Google > Bing."""
    # 1. Tentar DDGS API (mais confiável em ambientes de servidor)
    urls = buscar_ddgs_api(query, num_results)
    if urls:
        return urls, "DDGS API"
    # 2. DuckDuckGo HTML scraping
    urls = buscar_duckduckgo(session, query, headers, num_results)
    if urls:
        return urls, "DuckDuckGo HTML"
    # 3. Google
    urls = buscar_google_requests(session, query, headers, num_results)
    if urls:
        return urls, "Google"
    # 4. Bing
    urls = buscar_bing_requests(session, query, headers, num_results)
    if urls:
        return urls, "Bing"
    return [], "nenhum"


def extrair_precos_pagina(html_content):
    """Extrai possíveis preços de uma página HTML usando múltiplas estratégias."""
    from bs4 import BeautifulSoup

    precos_estruturados = []  # Alta confiança (JSON-LD, meta, classes de preço)
    precos_regex = []         # Baixa confiança (regex genérico)

    soup = BeautifulSoup(html_content, "html.parser")

    # --- Estratégia 1: JSON-LD (schema.org) — máxima confiança ---
    for script_tag in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script_tag.string or "")
            _extrair_preco_jsonld(data, precos_estruturados)
        except (json.JSONDecodeError, TypeError):
            continue

    # --- Estratégia 2: Meta tags (og:price, product:price) ---
    for meta in soup.find_all("meta"):
        prop = (meta.get("property") or meta.get("name") or "").lower()
        if any(k in prop for k in ("price", "amount", "preco")):
            content = meta.get("content", "")
            val = _parse_valor_br(content)
            if val:
                precos_estruturados.append(val)

    # --- Estratégia 3: Elementos com classes/atributos de preço ---
    seletores_preco = [
        "[class*='price']", "[class*='preco']", "[class*='Price']",
        "[class*='valor']", "[class*='Valor']",
        "[itemprop='price']", "[data-price]",
        "[class*='sale']", "[class*='offer']",
        "[class*='product-price']", "[class*='finalPrice']",
    ]
    for sel in seletores_preco:
        for elem in soup.select(sel):
            # Priorizar atributo content/data-price sobre texto
            for attr in ("content", "data-price", "data-value"):
                attr_val = elem.get(attr)
                if attr_val:
                    val = _parse_valor_br(attr_val)
                    if val:
                        precos_estruturados.append(val)
            # Texto do elemento
            texto = elem.get_text(strip=True)
            val = _extrair_valor_texto(texto)
            if val:
                precos_estruturados.append(val)

    # --- Estratégia 4: Regex no HTML bruto (fallback) ---
    padroes = [
        r"R\$\s*(\d{1,3}(?:\.\d{3})*,\d{2})",
        r"R\$\s*(\d+,\d{2})",
    ]
    for padrao in padroes:
        matches = re.findall(padrao, html_content, re.IGNORECASE)
        for match in matches:
            val = _parse_valor_br(match)
            if val:
                precos_regex.append(val)

    # Se temos preços estruturados, preferir eles; senão usar regex
    if precos_estruturados:
        return sorted(set(precos_estruturados))
    return sorted(set(precos_regex))


def _extrair_preco_jsonld(data, out):
    """Extrai preços de dados JSON-LD recursivamente."""
    if isinstance(data, list):
        for item in data:
            _extrair_preco_jsonld(item, out)
        return
    if not isinstance(data, dict):
        return
    # offers.price / offers.lowPrice
    for key in ("price", "lowPrice", "highPrice"):
        if key in data:
            val = _parse_valor_br(str(data[key]))
            if val:
                out.append(val)
    # Recursão em sub-objetos relevantes
    for key in ("offers", "priceSpecification", "mainEntity"):
        if key in data:
            _extrair_preco_jsonld(data[key], out)


def _parse_valor_br(texto):
    """Converte texto de preço BR ou internacional para float. Retorna None se inválido."""
    if not texto:
        return None
    texto = texto.strip().replace("R$", "").replace("\xa0", "").strip()
    # Formato BR: 1.234,56
    m = re.match(r"^(\d{1,3}(?:\.\d{3})*),(\d{2})$", texto)
    if m:
        val = float(texto.replace(".", "").replace(",", "."))
        return val if 0.50 < val < 500_000 else None
    # Formato BR simples: 123,45
    m = re.match(r"^(\d+),(\d{2})$", texto)
    if m:
        val = float(texto.replace(",", "."))
        return val if 0.50 < val < 500_000 else None
    # Formato internacional: 1234.56
    m = re.match(r"^(\d+(?:\.\d+)?)$", texto)
    if m:
        val = float(texto)
        return val if 0.50 < val < 500_000 else None
    return None


def _extrair_valor_texto(texto):
    """Extrai primeiro valor monetário de um texto curto."""
    if not texto:
        return None
    m = re.search(r"R\$\s*(\d{1,3}(?:\.\d{3})*,\d{2})", texto)
    if not m:
        m = re.search(r"(\d{1,3}(?:\.\d{3})*,\d{2})", texto)
    if m:
        return _parse_valor_br(m.group(1))
    return None


def extrair_titulo_pagina(html_content):
    """Extrai o título da página."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html_content, "html.parser")
    title_tag = soup.find("title")
    if title_tag and title_tag.string:
        return title_tag.string.strip()[:120]
    h1_tag = soup.find("h1")
    if h1_tag:
        return h1_tag.get_text(strip=True)[:120]
    return "Sem título"


def _eh_pagina_produto(html, titulo):
    """Verifica se a página parece ser de um produto/fornecedor e não blog/notícia/comparador."""
    titulo_lower = (titulo or "").lower()
    html_lower = html[:5000].lower()
    # Rejeitar páginas que claramente não são de produto
    termos_rejeitar = [
        "notícia", "artigo", "blog post", "publicado em", "autor:",
        "cookie policy", "política de privacidade", "terms of service",
        "404 not found", "page not found", "página não encontrada",
        "error 404", "403 forbidden", "access denied",
    ]
    for termo in termos_rejeitar:
        if termo in titulo_lower or termo in html_lower:
            return False
    # Indicadores positivos de página de produto
    indicadores_produto = [
        "add to cart", "adicionar ao carrinho", "comprar", "buy now",
        "add-to-cart", "addtocart", "carrinho", "cart",
        'itemprop="price"', 'itemprop="offers"', "schema.org/Product",
        "schema.org/Offer", '"@type":"Product"', '"@type": "Product"',
        "product-price", "preco", "preço", "valor unitário",
    ]
    for ind in indicadores_produto:
        if ind in html_lower or ind in html[:10000]:
            return True
    # Se tem preços detectáveis, considerar válido
    return True


def scraping_requests(session, url, headers):
    """Acessa uma página via requests e extrai informações."""
    from bs4 import BeautifulSoup

    try:
        resp = session.get(url, headers=headers, timeout=15, allow_redirects=True)
        if resp.status_code != 200:
            return None

        html = resp.text
        titulo = extrair_titulo_pagina(html)

        # Verificar se é uma página de produto antes de gastar tempo extraindo preços
        if not _eh_pagina_produto(html, titulo):
            return None

        precos = extrair_precos_pagina(html)

        if not precos:
            return None

        # Pegar o preço mais provável (mediana dos encontrados)
        preco_medio = sorted(precos)[len(precos) // 2]

        # Gerar screenshot HTML como evidência (sem precisar de Playwright)
        screenshot_path = None
        try:
            soup = BeautifulSoup(html, "html.parser")
            # Remover scripts e styles para captura limpa
            for tag in soup(["script", "style", "noscript", "svg", "iframe"]):
                tag.decompose()

            # Extrair trecho de texto ao redor do preço para contexto
            body_text_full = soup.get_text(separator="\n", strip=True)
            preco_formatado = f"{preco_medio:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            trecho_preco = ""
            idx_preco = body_text_full.find(preco_formatado)
            if idx_preco == -1:
                # Tentar formato sem milhar
                preco_simples = f"{preco_medio:.2f}".replace(".", ",")
                idx_preco = body_text_full.find(preco_simples)
            if idx_preco >= 0:
                inicio = max(0, idx_preco - 200)
                fim = min(len(body_text_full), idx_preco + 200)
                trecho_preco = body_text_full[inicio:fim]
            else:
                trecho_preco = body_text_full[:600]

            # Listar todos os preços encontrados
            todos_precos_str = " | ".join([f"R$ {p:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") for p in precos[:8]])

            screenshot_dir = SCREENSHOT_DIR
            os.makedirs(screenshot_dir, exist_ok=True)
            safe_name = re.sub(r'[^a-zA-Z0-9]', '_', titulo[:40])
            snapshot_path = os.path.join(screenshot_dir, f"{safe_name}_{hash(url) % 10000}.html")
            with open(snapshot_path, "w", encoding="utf-8") as f:
                f.write(f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Evidência — {titulo}</title>
<style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background: #f8f9fa; color: #333; }}
.evidence-card {{ max-width: 900px; margin: auto; background: #fff; border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.1); overflow: hidden; }}
.evidence-header {{ background: linear-gradient(135deg, #001a4d 0%, #003399 100%); color: #fff; padding: 20px 24px; }}
.evidence-header h2 {{ margin: 0 0 8px 0; font-size: 18px; color: #d4af37; }}
.evidence-price {{ font-size: 28px; font-weight: bold; color: #4CAF50; margin: 12px 0; }}
.evidence-meta {{ display: flex; gap: 24px; flex-wrap: wrap; font-size: 12px; color: #ccc; margin-top: 8px; }}
.evidence-meta span {{ display: flex; align-items: center; gap: 4px; }}
.evidence-body {{ padding: 20px 24px; }}
.evidence-section {{ margin-bottom: 16px; }}
.evidence-section-title {{ font-weight: 600; color: #001a4d; font-size: 14px; margin-bottom: 6px; border-bottom: 2px solid #d4af37; padding-bottom: 4px; display: inline-block; }}
.evidence-url {{ color: #1a73e8; word-break: break-all; font-size: 13px; }}
.evidence-context {{ background: #f5f5f5; border-left: 4px solid #d4af37; padding: 12px 16px; border-radius: 0 8px 8px 0; font-size: 13px; line-height: 1.6; white-space: pre-wrap; word-wrap: break-word; max-height: 400px; overflow-y: auto; }}
.all-prices {{ background: #e8f5e9; padding: 8px 12px; border-radius: 6px; font-size: 13px; color: #2e7d32; }}
.evidence-footer {{ background: #f0f0f0; text-align: center; padding: 10px; font-size: 11px; color: #999; }}
</style></head><body>
<div class="evidence-card">
  <div class="evidence-header">
    <h2>{titulo}</h2>
    <div class="evidence-price">R$ {preco_medio:,.2f}</div>
    <div class="evidence-meta">
      <span>📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}</span>
      <span>🌐 {extrair_dominio(url)}</span>
      <span>📊 {len(precos)} preço(s) detectado(s)</span>
    </div>
  </div>
  <div class="evidence-body">
    <div class="evidence-section">
      <div class="evidence-section-title">🔗 Fonte</div>
      <div class="evidence-url"><a href="{url}" target="_blank">{url}</a></div>
    </div>
    <div class="evidence-section">
      <div class="evidence-section-title">💰 Todos os Preços Detectados</div>
      <div class="all-prices">{todos_precos_str}</div>
    </div>
    <div class="evidence-section">
      <div class="evidence-section-title">📄 Contexto Extraído</div>
      <div class="evidence-context">{trecho_preco}</div>
    </div>
  </div>
  <div class="evidence-footer">Evidência gerada automaticamente pelo AtaCotada — Marinha do Brasil</div>
</div>
</body></html>""")
            screenshot_path = snapshot_path
        except Exception:
            pass

        return {
            "titulo": titulo,
            "preco": preco_medio,
            "url": url,
            "dominio": extrair_dominio(url),
            "screenshot": screenshot_path,
        }
    except Exception:
        return None


# ===================== SCRAPING COM PLAYWRIGHT =====================

def _is_streamlit_cloud():
    """Detecta se está rodando no Streamlit Community Cloud."""
    # No Community Cloud, o HOME é /home/appuser e existe a variável STREAMLIT_SHARING_MODE
    return (
        os.environ.get("STREAMLIT_SHARING_MODE") is not None
        or os.environ.get("HOME", "") == "/home/appuser"
        or os.path.exists("/mount/src")
    )


def _ensure_playwright_installed():
    """Verifica se playwright está instalado (pacote + browser baixado)."""
    # Playwright não funciona no Streamlit Community Cloud (sem sudo, sem binários de browser)
    if _is_streamlit_cloud():
        return False
    try:
        from playwright.sync_api import sync_playwright
        import glob

        # Verificar em múltiplos caminhos possíveis (Codespaces podem variar o HOME)
        candidate_paths = []

        # PLAYWRIGHT_BROWSERS_PATH tem prioridade
        env_path = os.environ.get("PLAYWRIGHT_BROWSERS_PATH")
        if env_path:
            candidate_paths.append(env_path)

        # Caminho padrão baseado no HOME atual
        home = os.path.expanduser("~")
        candidate_paths.append(os.path.join(home, ".cache", "ms-playwright"))

        # Caminhos comuns em Codespaces / dev containers
        for user_dir in ["/home/codespace", "/home/vscode", "/root"]:
            p = os.path.join(user_dir, ".cache", "ms-playwright")
            if p not in candidate_paths:
                candidate_paths.append(p)

        for browsers_path in candidate_paths:
            if os.path.isdir(browsers_path) and glob.glob(os.path.join(browsers_path, "chromium*")):
                return True

        return False
    except ImportError:
        return False


def _fechar_popups(page):
    """Tenta fechar popups, modais, banners de cookies e propagandas."""
    # Seletores de botões de fechar (ordem: mais específicos primeiro)
    seletores_fechar = [
        # Banners de cookies / LGPD
        'button:has-text("Aceitar")', 'button:has-text("Aceito")',
        'button:has-text("Concordo")', 'button:has-text("Entendi")',
        'button:has-text("Prosseguir")', 'button:has-text("Accept")',
        'button:has-text("Got it")', 'button:has-text("I agree")',
        'button:has-text("continuar e fechar")', 'button:has-text("Continuar")',
        'a:has-text("continuar e fechar")', 'a:has-text("Continuar e fechar")',
        '[class*="cookie"] button', '[id*="cookie"] button',
        '[class*="lgpd"] button', '[id*="lgpd"] button',
        '[class*="consent"] button', '[id*="consent"] button',
        # Botões X de fechar (popups promocionais)
        '[class*="modal"] [class*="close"]', '[class*="modal"] button[class*="close"]',
        '[class*="popup"] [class*="close"]', '[class*="popup"] button[class*="close"]',
        '[class*="promo"] [class*="close"]', '[class*="banner"] [class*="close"]',
        '.modal .close', '.popup-close', '.btn-close',
        'button.close', '[data-dismiss="modal"]',
        '[aria-label="Close"]', '[aria-label="Fechar"]',
        '[aria-label="close"]', '[aria-label="fechar"]',
        'button[title="Close"]', 'button[title="Fechar"]',
        # Fechar genéricos (ícone X)
        '[class*="close-btn"]', '[class*="closebtn"]', '[class*="close_btn"]',
        '[class*="CloseButton"]', '[class*="closeButton"]',
        '[class*="icon-close"]', '[class*="icon_close"]',
        '[class*="dismiss"]', '[class*="Dismiss"]',
        # Overlays
        '[class*="overlay"] [class*="close"]',
        '[class*="lightbox"] [class*="close"]',
    ]
    for seletor in seletores_fechar:
        try:
            el = page.locator(seletor).first
            if el.is_visible(timeout=300):
                el.click(timeout=1000)
                time.sleep(0.3)
        except Exception:
            continue

    # Tentar fechar modais/overlays via tecla Escape
    try:
        page.keyboard.press("Escape")
        time.sleep(0.3)
    except Exception:
        pass

    # Remover overlays/modais via JavaScript (força bruta para popups resistentes)
    try:
        page.evaluate("""() => {
            // Remover elementos com z-index alto que cobrem a tela (modais/overlays)
            const all = document.querySelectorAll('*');
            for (const el of all) {
                const style = window.getComputedStyle(el);
                const zIndex = parseInt(style.zIndex) || 0;
                const pos = style.position;
                // Elementos fixos/absolutos com z-index alto = popup/modal/overlay
                if ((pos === 'fixed' || pos === 'absolute') && zIndex > 100) {
                    const rect = el.getBoundingClientRect();
                    // Se cobre boa parte da tela, remover
                    if (rect.width > window.innerWidth * 0.3 && rect.height > window.innerHeight * 0.3) {
                        el.remove();
                    }
                }
            }
            // Restaurar scroll no body (alguns modais bloqueiam)
            document.body.style.overflow = 'auto';
            document.documentElement.style.overflow = 'auto';
        }""")
        time.sleep(0.3)
    except Exception:
        pass


def _scrollar_ate_preco(page):
    """Tenta scrollar até o elemento que contém o preço e retorna a bounding box para clip."""
    # Primeiro tentar encontrar o container do produto (título + preço juntos)
    seletores_produto = [
        '[class*="product-info"]', '[class*="product-detail"]',
        '[class*="productInfo"]', '[class*="produto"]',
        '[class*="product-main"]', '[class*="product-summary"]',
        '[itemtype*="schema.org/Product"]',
        'main', '[role="main"]', '#product', '#produto',
    ]
    seletores_preco = [
        '[class*="price"]', '[class*="preco"]', '[class*="valor"]',
        '[class*="Price"]', '[class*="product-price"]',
        '[data-testid*="price"]', '[itemprop="price"]',
        '.price', '#price', '.product-price',
    ]
    # Tentar container do produto primeiro (captura título + preço)
    for seletor in seletores_produto:
        try:
            el = page.locator(seletor).first
            if el.is_visible(timeout=500):
                el.scroll_into_view_if_needed(timeout=2000)
                box = el.bounding_box()
                if box and box['height'] > 100:
                    return box
        except Exception:
            continue
    # Fallback: scrollar até o preço
    for seletor in seletores_preco:
        try:
            el = page.locator(seletor).first
            if el.is_visible(timeout=500):
                el.scroll_into_view_if_needed(timeout=2000)
                return None
        except Exception:
            continue
    # Se nada encontrado, voltar ao topo
    try:
        page.evaluate("window.scrollTo(0, 0)")
    except Exception:
        pass
    return None


def scraping_playwright(url, item_nome, screenshot_path=None):
    """Acessa uma página via Playwright (para sites dinâmicos) e extrai informações."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None

    resultado = None

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-blink-features=AutomationControlled",
                ],
            )

            context = browser.new_context(
                viewport={"width": 1366, "height": 768},
                user_agent=escolher_user_agent(),
                locale="pt-BR",
                timezone_id="America/Sao_Paulo",
            )

            page = context.new_page()

            # Navegar com timeout
            page.goto(url, wait_until="domcontentloaded", timeout=20000)

            # Aguardar possível carregamento dinâmico
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass  # Timeout de networkidle não é crítico

            # Fechar popups, modais e banners de cookies
            _fechar_popups(page)

            # Simular scroll humano
            page.evaluate("window.scrollBy(0, Math.random() * 400 + 200)")
            time.sleep(gerar_delay_leitura(1.0, 2.5))

            # Mais um scroll
            page.evaluate("window.scrollBy(0, Math.random() * 300 + 100)")
            time.sleep(random.uniform(0.5, 1.5))

            html = page.content()
            titulo = page.title() or extrair_titulo_pagina(html)

            # Verificar se é uma página de produto
            if not _eh_pagina_produto(html, titulo):
                browser.close()
                return None

            precos = extrair_precos_pagina(html)

            # Captura de tela real como PNG
            if screenshot_path and precos:
                os.makedirs(os.path.dirname(screenshot_path), exist_ok=True)
                # Tentar scrollar até o elemento com preço para capturar evidência clara
                box = _scrollar_ate_preco(page)
                time.sleep(0.5)
                # Se encontrou bounding box do container do produto, fazer clip
                if box:
                    # Expandir a área para dar contexto visual (mais espaço acima e abaixo)
                    clip_x = max(0, box['x'] - 30)
                    clip_y = max(0, box['y'] - 80)  # mais espaço acima para pegar título
                    clip_w = min(box['width'] + 60, 1366 - clip_x)
                    clip_h = min(box['height'] + 200, 768 * 2)  # mais espaço abaixo
                    clip_h = max(clip_h, 600)  # mínimo de 600px de altura
                    page.screenshot(path=screenshot_path, clip={'x': clip_x, 'y': clip_y, 'width': clip_w, 'height': clip_h})
                else:
                    # Sem box: scrollar um pouco abaixo do cabeçalho para evitar cortar produto
                    page.evaluate("window.scrollTo(0, 150)")
                    time.sleep(0.3)
                    page.screenshot(path=screenshot_path, full_page=False)

            browser.close()

            if precos:
                preco_medio = sorted(precos)[len(precos) // 2]
                resultado = {
                    "titulo": titulo,
                    "preco": preco_medio,
                    "url": url,
                    "dominio": extrair_dominio(url),
                    "screenshot": screenshot_path if screenshot_path and os.path.exists(screenshot_path) else None,
                }
    except Exception as e:
        # Logar o erro em vez de engolir silenciosamente
        import traceback
        traceback.print_exc()

    return resultado


# ===================== ORQUESTRADOR DE SCRAPING =====================

def executar_scraping(itens, usar_playwright, progress_bar, log_container, status_text):
    """Executa o scraping completo para todos os itens."""
    import requests as req

    logs = []
    resultados = []
    total_itens = len(itens)
    playwright_disponivel = usar_playwright and _ensure_playwright_installed()

    # Criar sessão reutilizável
    session = req.Session()
    ua = escolher_user_agent()
    headers = gerar_headers(ua)
    session.headers.update(headers)

    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    for idx, item in enumerate(itens):
        item = item.strip()
        if not item:
            continue

        log_msg(log_container, logs, f"━━━ Iniciando busca: <b>{item}</b> ({idx+1}/{total_itens}) ━━━", "info")
        status_text.text(f"Buscando: {item} ({idx+1}/{total_itens})")
        progress_bar.progress((idx) / total_itens)

        orcamentos_item = []
        dominios_usados = set()

        # Selecionar variantes de busca aleatoriamente (usar mais variantes para maximizar cobertura)
        variantes = random.sample(VARIANTES_BUSCA, min(5, len(VARIANTES_BUSCA)))

        for variante in variantes:
            if len(orcamentos_item) >= MAX_FONTES_POR_ITEM:
                log_msg(log_container, logs, f"✓ {MAX_FONTES_POR_ITEM} orçamentos encontrados para '{item}'. Avançando.", "success")
                break

            query = variante.format(item=item)
            log_msg(log_container, logs, f"🔍 Buscando: \"{query}\"", "info")

            # Delay antes da busca
            delay = gerar_delay(2.0, 5.0)
            log_msg(log_container, logs, f"⏳ Aguardando {delay:.1f}s...", "info")
            time.sleep(delay)

            # Buscar URLs (DDGS API > DuckDuckGo HTML > Google > Bing)
            urls, engine = buscar_urls(session, query, headers)

            if not urls:
                log_msg(log_container, logs, f"⚠ Nenhum resultado encontrado para \"{query}\"", "warn")
                continue

            log_msg(log_container, logs, f"📋 {len(urls)} resultados encontrados via {engine}", "info")

            for url in urls:
                if len(orcamentos_item) >= MAX_FONTES_POR_ITEM:
                    break

                dominio = extrair_dominio(url)
                if dominio in dominios_usados:
                    continue

                log_msg(log_container, logs, f"🌐 Acessando: {dominio}", "info")

                # Delay entre acessos a sites
                delay = gerar_delay(2.5, 6.0)
                log_msg(log_container, logs, f"⏳ Delay de navegação: {delay:.1f}s", "info")
                time.sleep(delay)

                resultado = None
                screenshot_path = os.path.join(
                    SCREENSHOT_DIR,
                    f"{re.sub(r'[^a-zA-Z0-9]', '_', item)}_{len(orcamentos_item)+1}.png",
                )

                # Se Playwright selecionado e disponível, usar primeiro
                if playwright_disponivel:
                    log_msg(log_container, logs, f"🎭 Tentando com navegador automatizado: {dominio}", "info")
                    time.sleep(gerar_delay(1.5, 3.5))
                    resultado = scraping_playwright(url, item, screenshot_path)

                # Se Playwright não disponível ou falhou, tentar com requests
                if not resultado:
                    for tentativa in range(MAX_RETRIES + 1):
                        # Simular tempo de leitura
                        time.sleep(gerar_delay_leitura())

                        resultado = scraping_requests(session, url, headers)
                        if resultado:
                            break

                        if tentativa < MAX_RETRIES:
                            retry_delay = gerar_delay(3.0, 7.0)
                            log_msg(log_container, logs, f"🔄 Retry {tentativa+1}/{MAX_RETRIES} em {retry_delay:.1f}s...", "warn")
                            time.sleep(retry_delay)
                            # Trocar User-Agent no retry
                            headers = gerar_headers()
                            session.headers.update(headers)

                if resultado:
                    resultado["item"] = item
                    resultado["data_coleta"] = datetime.now().strftime("%d/%m/%Y %H:%M")

                    # Capturar screenshot via Playwright se ainda não temos
                    if playwright_disponivel and not resultado.get("screenshot"):
                        try:
                            _resultado_pw = scraping_playwright(url, item, screenshot_path)
                            if _resultado_pw and _resultado_pw.get("screenshot"):
                                resultado["screenshot"] = _resultado_pw["screenshot"]
                        except Exception:
                            pass

                    orcamentos_item.append(resultado)
                    dominios_usados.add(dominio)
                    log_msg(
                        log_container,
                        logs,
                        f"💰 Orçamento [{len(orcamentos_item)}/{MAX_FONTES_POR_ITEM}] — {formatar_moeda_br(resultado['preco'])} em {dominio}",
                        "orcamento",
                    )
                else:
                    log_msg(log_container, logs, f"✗ Sem preço extraível de {dominio}", "error")

        # Complemento: se não atingiu o mínimo de fontes, usar SearchAPI (Google Shopping)
        faltam = MAX_FONTES_POR_ITEM - len(orcamentos_item)
        if faltam > 0:
            log_msg(log_container, logs, f"🛒 Faltam {faltam} orçamento(s) para '{item}'. Tentando Google Shopping (SearchAPI)...", "info")
            searchapi_results = buscar_searchapi(item, faltam + 2)  # pedir extras para compensar duplicados
            if searchapi_results:
                dominios_ja = {r['dominio'] for r in orcamentos_item}
                for sr in searchapi_results:
                    if len(orcamentos_item) >= MAX_FONTES_POR_ITEM:
                        break
                    if sr.get('dominio') in dominios_ja:
                        continue
                    sr["item"] = item
                    sr["data_coleta"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                    orcamentos_item.append(sr)
                    dominios_ja.add(sr.get('dominio', ''))
                    log_msg(
                        log_container,
                        logs,
                        f"💰 Orçamento [{len(orcamentos_item)}/{MAX_FONTES_POR_ITEM}] — Google Shopping: {formatar_moeda_br(sr['preco'])} em {sr['dominio']}",
                        "orcamento",
                    )
            if len(orcamentos_item) < MAX_FONTES_POR_ITEM:
                log_msg(log_container, logs, f"⚠ Apenas {len(orcamentos_item)} orçamento(s) encontrado(s) para '{item}'", "warn")

        resultados.extend(orcamentos_item)

        # Delay maior entre itens diferentes
        if idx < total_itens - 1:
            delay = gerar_delay(4.0, 8.0)
            log_msg(log_container, logs, f"⏳ Intervalo entre itens: {delay:.1f}s", "info")
            time.sleep(delay)

    progress_bar.progress(1.0)
    log_msg(log_container, logs, f"━━━ Scraping concluído! {len(resultados)} orçamentos coletados ━━━", "success")
    status_text.text("Scraping concluído!")

    return resultados


# ===================== GERAÇÃO DE RELATÓRIO =====================

def gerar_relatorio_excel(resultados):
    """Gera relatório Excel com os resultados do scraping."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Cotações"

    # Estilos
    header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="001A4D", end_color="001A4D", fill_type="solid")
    gold_font = Font(name="Arial", bold=True, size=14, color="D4AF37")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Título
    ws.merge_cells("A1:F1")
    titulo_cell = ws["A1"]
    titulo_cell.value = "RELATÓRIO DE COTAÇÕES - WEB SCRAPING"
    titulo_cell.font = gold_font
    titulo_cell.alignment = Alignment(horizontal="center")
    titulo_cell.fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")

    ws.merge_cells("A2:F2")
    ws["A2"].value = f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeçalhos
    headers = ["Item", "Fornecedor/Site", "Preço", "Link", "Data da Coleta", "Título da Página"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Dados
    for row_idx, r in enumerate(resultados, 5):
        ws.cell(row=row_idx, column=1, value=r.get("item", "")).border = border
        ws.cell(row=row_idx, column=2, value=r.get("dominio", "")).border = border
        preco_cell = ws.cell(row=row_idx, column=3, value=r.get("preco", 0))
        preco_cell.number_format = 'R$ #,##0.00'
        preco_cell.border = border
        ws.cell(row=row_idx, column=4, value=r.get("url", "")).border = border
        ws.cell(row=row_idx, column=5, value=r.get("data_coleta", "")).border = border
        ws.cell(row=row_idx, column=6, value=r.get("titulo", "")).border = border

    # Resumo por item
    ws_resumo = wb.create_sheet("Resumo por Item")
    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"].value = "RESUMO POR ITEM"
    ws_resumo["A1"].font = gold_font
    ws_resumo["A1"].fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    ws_resumo["A1"].alignment = Alignment(horizontal="center")

    resumo_headers = ["Item", "Menor Preço", "Maior Preço", "Preço Médio"]
    for col, h in enumerate(resumo_headers, 1):
        cell = ws_resumo.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    df = pd.DataFrame(resultados)
    if not df.empty and "preco" in df.columns:
        resumo = df.groupby("item")["preco"].agg(["min", "max", "mean"]).reset_index()
        for row_idx, (_, row) in enumerate(resumo.iterrows(), 4):
            ws_resumo.cell(row=row_idx, column=1, value=row["item"]).border = border
            for c, col_name in enumerate(["min", "max", "mean"], 2):
                cell = ws_resumo.cell(row=row_idx, column=c, value=row[col_name])
                cell.number_format = 'R$ #,##0.00'
                cell.border = border

    # Ajustar largura das colunas
    for ws_sheet in [ws, ws_resumo]:
        for col in ws_sheet.columns:
            max_len = 0
            try:
                col_letter = col[0].column_letter
            except AttributeError:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws_sheet.column_dimensions[col_letter].width = min(max_len + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_relatorio_csv(resultados):
    """Gera CSV simples dos resultados."""
    df = pd.DataFrame(resultados)
    if df.empty:
        return None
    colunas = ["item", "dominio", "preco", "url", "data_coleta", "titulo"]
    colunas_existentes = [c for c in colunas if c in df.columns]
    df = df[colunas_existentes]
    df.columns = ["Item", "Fornecedor", "Preço (R$)", "Link", "Data Coleta", "Título"][:len(colunas_existentes)]
    output = BytesIO()
    df.to_csv(output, index=False, encoding="utf-8-sig", sep=";")
    output.seek(0)
    return output


def gerar_pdf_evidencias(resultados):
    """Gera um PDF único com todas as evidências de preço coletadas."""
    from fpdf import FPDF

    class EvidenciaPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(0, 26, 77)
            self.cell(0, 8, "AtaCotada - Evidencias de Pesquisa de Precos", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_draw_color(212, 175, 55)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(4)

        def footer(self):
            self.set_y(-15)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f"Marinha do Brasil - Pagina {self.page_no()}/{{nb}}", align="C")

    pdf = EvidenciaPDF(orientation="P", unit="mm", format="A4")
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=20)

    screenshots = [r for r in resultados if r.get("screenshot") and os.path.exists(r.get("screenshot", ""))]

    if not screenshots:
        pdf.add_page()
        pdf.set_font("Helvetica", "", 14)
        pdf.cell(0, 40, "Nenhuma evidencia capturada.", align="C")
        buf = BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf

    # Capa
    pdf.add_page()
    pdf.ln(30)
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(0, 26, 77)
    pdf.cell(0, 15, "Relatorio de Evidencias", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 10, "Pesquisa de Precos - Web Scraping", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    pdf.set_draw_color(212, 175, 55)
    pdf.set_line_width(1)
    pdf.line(60, pdf.get_y(), 150, pdf.get_y())
    pdf.ln(10)
    pdf.set_font("Helvetica", "", 12)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 8, f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 8, f"Total de evidencias: {len(screenshots)}", align="C", new_x="LMARGIN", new_y="NEXT")
    itens_unicos = list(set(r.get("item", "") for r in screenshots))
    pdf.cell(0, 8, f"Itens pesquisados: {len(itens_unicos)}", align="C", new_x="LMARGIN", new_y="NEXT")

    # Páginas de evidência
    for i, r in enumerate(screenshots, 1):
        pdf.add_page()

        # Cabeçalho da evidência
        pdf.set_fill_color(0, 26, 77)
        pdf.rect(10, pdf.get_y(), 190, 28, "F")
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(212, 175, 55)
        y_start = pdf.get_y() + 3
        pdf.set_xy(14, y_start)
        pdf.cell(180, 6, f"Evidencia {i}/{len(screenshots)}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(14)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(255, 255, 255)
        titulo_clean = (r.get("titulo", "Sem titulo") or "Sem titulo").encode("latin-1", "replace").decode("latin-1")
        pdf.cell(180, 5, titulo_clean[:90], new_x="LMARGIN", new_y="NEXT")

        # Preço em destaque
        pdf.set_x(14)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(76, 175, 80)
        preco_str = f"R$ {r['preco']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        pdf.cell(180, 10, preco_str, new_x="LMARGIN", new_y="NEXT")

        pdf.ln(6)

        # Info do item
        pdf.set_text_color(60, 60, 60)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Item:")
        pdf.set_font("Helvetica", "", 10)
        item_clean = (r.get("item", "-") or "-").encode("latin-1", "replace").decode("latin-1")
        pdf.cell(0, 7, item_clean, new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Fornecedor:")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, r.get("dominio", "-"), new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "Data Coleta:")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, r.get("data_coleta", datetime.now().strftime("%d/%m/%Y %H:%M")), new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(30, 7, "URL:")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(26, 115, 232)
        url_str = (r.get("url", "-") or "-")[:120]
        pdf.cell(0, 7, url_str, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(60, 60, 60)

        # Separador
        pdf.ln(4)
        pdf.set_draw_color(212, 175, 55)
        pdf.set_line_width(0.3)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(4)

        # Conteúdo da evidência (trecho do HTML)
        sc_path = r["screenshot"]
        if sc_path.endswith(".html"):
            try:
                from bs4 import BeautifulSoup as BS4
                with open(sc_path, "r", encoding="utf-8") as f:
                    ev_html = f.read()
                ev_soup = BS4(ev_html, "html.parser")
                # Pegar o trecho de contexto
                contexto_div = ev_soup.select_one(".evidence-context")
                if contexto_div:
                    contexto_texto = contexto_div.get_text(strip=True)[:1500]
                else:
                    contexto_texto = ev_soup.get_text(separator="\n", strip=True)[:1500]
                contexto_texto = contexto_texto.encode("latin-1", "replace").decode("latin-1")

                pdf.set_font("Helvetica", "B", 10)
                pdf.set_text_color(0, 26, 77)
                pdf.cell(0, 7, "Contexto Extraido da Pagina:", new_x="LMARGIN", new_y="NEXT")
                pdf.set_font("Helvetica", "", 9)
                pdf.set_text_color(80, 80, 80)
                pdf.multi_cell(190, 5, contexto_texto)
            except Exception:
                pdf.set_font("Helvetica", "I", 9)
                pdf.cell(0, 7, f"Evidencia salva em: {sc_path}", new_x="LMARGIN", new_y="NEXT")

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


# ===================== SIDEBAR =====================

# Carregar imagem do acanto para a sidebar
_acanto_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Projeto Adesões", "acanto.png")
if os.path.exists(_acanto_path):
    with open(_acanto_path, "rb") as _f:
        _acanto_b64 = base64.b64encode(_f.read()).decode()
else:
    _acanto_b64 = None

with st.sidebar:
    if _acanto_b64:
        st.markdown(f'<div style="text-align:center;padding:1rem 0 0.5rem 0;"><img src="data:image/png;base64,{_acanto_b64}" style="max-width:70%;height:auto;"></div>', unsafe_allow_html=True)
    st.markdown("## MENU")
    st.markdown("---")
    st.page_link("streamlit_app.py", label="Cotação", icon="⚓")
    st.page_link("pages/Adesões.py", label="Adesões", icon="🤝")
    st.page_link("pages/Notas_Fiscais.py", label="Notas Fiscais", icon="📄")
    st.page_link("pages/Banco_de_Fornecedores.py", label="Fornecedores", icon="🏢")
    st.page_link("pages/Consulta.py", label="Consulta CNPJ", icon="💻")
    st.page_link("pages/Web_Scraping.py", label="Web Scraping", icon="🕷️")
    st.markdown("---")
    st.markdown("## LINKS ÚTEIS")
    st.markdown("""
    <div style="margin-bottom: 1rem;">
        <a href="https://freitasnascimentomarinha-debug.github.io/ShootMail/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            📧 Disparador de Emails
        </a>
    </div>
    <div style="margin-bottom: 1rem;">
        <a href="https://detetive-obtencao.vercel.app/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            🚨 Detetive Obtenção
        </a>
    </div>
    """, unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;color:#d4af37;font-size:10px;font-weight:600;padding:0.3rem 0;white-space:nowrap;">Centro de Operações do Abastecimento</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-footer">Marinha do Brasil<br>AtaCotada v1.0</div>', unsafe_allow_html=True)


# ===================== HEADER =====================

st.markdown("""
    <div class="header-container">
        <div class="logo-text">MARINHA DO BRASIL</div>
        <div class="sistema-nome">ATACOTADA</div>
        <div class="subtitulo">🕷️ Web Scraping — Pesquisa de Preços Automatizada</div>
    </div>
""", unsafe_allow_html=True)


# ===================== INTERFACE PRINCIPAL =====================

import streamlit.components.v1 as _components

_como_funciona_html = """
<div style="background: linear-gradient(135deg, #1a1a1a 0%, #252525 100%); border: 1px solid #333; border-radius: 12px; padding: 1.5rem; margin-bottom: 1.5rem; color: #e0e0e0; font-family: 'Segoe UI', sans-serif; font-size: 15px; line-height: 1.6;">
    <div style="font-size: 1.2rem; font-weight: bold; color: #d4af37; margin-bottom: 1rem;">⚙️ Como Funciona o Web Scraping</div>
    <div style="margin-bottom: 0.8rem;">Este módulo automatiza a pesquisa de preços na internet para fins de <b>cotação e estimativa de preços</b>,
    em conformidade com a IN 65/2021. O sistema busca preços diretamente em sites de fornecedores
    (ignorando marketplaces como Mercado Livre, Amazon, Shopee etc.) para obter valores mais próximos
    da realidade praticada no comércio direto.</div>

    <div style="font-weight:bold; color:#d4af37; margin-bottom: 0.5rem;">🔄 Fluxo de Execução:</div>
    <div style="margin-left: 1rem; margin-bottom: 1rem;">
        <div style="margin-bottom:0.3rem;">1. Você informa os itens que deseja pesquisar (um por linha)</div>
        <div style="margin-bottom:0.3rem;">2. O sistema gera até <b>5 variações de busca</b> para cada item (ex: "caneta preço", "comprar caneta online", "caneta fornecedor")</div>
        <div style="margin-bottom:0.3rem;">3. Para cada variação, busca URLs relevantes usando <b>4 mecanismos em cascata</b>:
            <br><b>DDGS API → DuckDuckGo HTML → Google → Bing</b></div>
        <div style="margin-bottom:0.3rem;">4. Acessa cada site encontrado e extrai preços usando <b>4 estratégias de detecção</b>:
            <br>dados estruturados (JSON-LD) → meta tags → classes de preço no HTML → regex em R$</div>
        <div style="margin-bottom:0.3rem;">5. Se ao final não atingir o mínimo de orçamentos por item, complementa automaticamente com o <b>Google Shopping (SearchAPI)</b>, que retorna preços de lojas cadastradas no Google</div>
        <div style="margin-bottom:0.3rem;">6. Salva uma evidência formatada (snapshot) de cada página com preço encontrado</div>
        <div style="margin-bottom:0.3rem;">7. Gera relatório exportável em <b>Excel, CSV, JSON ou PDF de evidências</b></div>
    </div>

    <div style="font-weight:bold; color:#d4af37; margin-bottom: 0.5rem;">⚙️ Configurações Disponíveis:</div>
    <div style="margin-left: 1rem; margin-bottom: 1rem;">
        <div style="margin-bottom:0.4rem;">• <b>Navegador Automatizado (Playwright):</b> Quando ativado, usa um navegador real (Chromium)
            para acessar sites que carregam preços via JavaScript. É mais lento, mas captura preços
            de sites dinâmicos que o modo padrão não consegue ler. <i>Recomendação: deixe desativado
            na maioria dos casos; ative apenas se estiver recebendo poucos resultados.</i></div>
        <div style="margin-bottom:0.4rem;">• <b>Máx. fontes por item:</b> Quantidade máxima de orçamentos diferentes que o sistema
            buscará para cada material. <i>Recomendação: <b>3</b> fontes é o ideal — já atende
            à IN 65/2021 e mantém a pesquisa rápida.</i></div>
        <div style="margin-bottom:0.4rem;">• <b>Delay mínimo / máximo (seg):</b> Intervalo de espera entre cada requisição,
            simulando comportamento humano. Evita bloqueios dos sites.
            <i>Recomendação: mínimo <b>2s</b> e máximo <b>6s</b> (padrão) —
            aumente para 4s/10s se pesquisar muitos itens de uma vez.</i></div>
    </div>
</div>
"""
with st.expander("⚙️ Como Funciona o Web Scraping", expanded=False):
    _components.html(_como_funciona_html, height=700, scrolling=True)

# Formulário de entrada
st.markdown("### 📝 Itens para Pesquisa")

col1, col2 = st.columns([3, 1])

with col1:
    itens_input = st.text_area(
        "Informe os itens (um por linha):",
        height=150,
        placeholder="Exemplo:\nArruela de pressão 1/4\nParafuso sextavado M10\nFita isolante 20m",
        help="Digite os nomes dos materiais que deseja pesquisar, um por linha.",
    )

with col2:
    st.markdown("#### ⚙️ Configurações")
    if _is_streamlit_cloud():
        st.info("ℹ️ Navegador automatizado indisponível na nuvem.", icon="☁️")
        usar_playwright = False
    else:
        usar_playwright = st.checkbox(
            "Usar navegador automatizado",
            value=False,
            help="Ativa o Playwright para sites com carregamento dinâmico. Mais lento, porém mais preciso.",
        )
        if usar_playwright:
            if _ensure_playwright_installed():
                st.success("✅ Playwright ativo", icon="🎭")
            else:
                st.error("❌ Playwright não disponível. Instale com: pip install playwright && playwright install chromium")
                usar_playwright = False
    max_fontes = st.number_input(
        "Máx. fontes por item",
        min_value=1,
        max_value=5,
        value=3,
        help="Número máximo de orçamentos por item.",
    )
    delay_min = st.number_input("Delay mín. (seg)", min_value=1.0, max_value=15.0, value=2.0, step=0.5)
    delay_max = st.number_input("Delay máx. (seg)", min_value=2.0, max_value=30.0, value=6.0, step=0.5)

# Validação
if delay_min >= delay_max:
    st.warning("O delay mínimo deve ser menor que o máximo.")

# Botão de execução
st.markdown("---")
col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])

with col_btn2:
    iniciar = st.button("🚀 Iniciar Scraping", type="primary", use_container_width=True)

if iniciar:
    itens = [i.strip() for i in itens_input.strip().split("\n") if i.strip()]

    if not itens:
        st.error("⚠️ Informe pelo menos um item para pesquisa.")
    else:
        # Atualizar constante global com a config do usuário
        MAX_FONTES_POR_ITEM_LOCAL = max_fontes

        st.markdown("### 📊 Execução do Scraping")

        # Área de logs
        st.markdown("#### 📜 Log de Execução")
        log_container = st.empty()

        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()

        # Executar scraping
        resultados = executar_scraping(
            itens=itens,
            usar_playwright=usar_playwright,
            progress_bar=progress_bar,
            log_container=log_container,
            status_text=status_text,
        )

        # Armazenar resultados no session_state
        st.session_state["scraping_resultados"] = resultados
        st.session_state["scraping_itens"] = itens

# ===================== EXIBIÇÃO DE RESULTADOS =====================

if "scraping_resultados" in st.session_state and st.session_state["scraping_resultados"]:
    resultados = st.session_state["scraping_resultados"]

    st.markdown("---")
    st.markdown("### 📋 Resultados")

    tab_tabela, tab_resumo, tab_evidencias, tab_export = st.tabs(
        ["📊 Tabela Completa", "📈 Resumo", "📸 Evidências", "📥 Exportar"]
    )

    with tab_tabela:
        df = pd.DataFrame(resultados)
        df_display = df[["item", "dominio", "preco", "url", "data_coleta", "titulo"]].copy()
        df_display.columns = ["Item", "Fornecedor", "Preço (R$)", "Link", "Data Coleta", "Título"]
        df_display["Preço (R$)"] = df_display["Preço (R$)"].apply(formatar_moeda_br)

        st.dataframe(
            df_display,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Link": st.column_config.LinkColumn("Link", display_text="Acessar"),
            },
        )

    with tab_resumo:
        df = pd.DataFrame(resultados)

        if not df.empty:
            # Métricas gerais
            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            with col_m1:
                st.metric("Total de Orçamentos", len(resultados))
            with col_m2:
                st.metric("Itens Pesquisados", len(st.session_state.get("scraping_itens", [])))
            with col_m3:
                st.metric("Fontes Distintas", df["dominio"].nunique())
            with col_m4:
                st.metric("Menor Preço", formatar_moeda_br(df["preco"].min()))

            st.markdown("#### Resumo por Item")
            resumo = df.groupby("item")["preco"].agg(["count", "min", "max", "mean"]).reset_index()
            resumo.columns = ["Item", "Qtd. Orçamentos", "Menor Preço", "Maior Preço", "Preço Médio"]
            resumo["Menor Preço"] = resumo["Menor Preço"].apply(formatar_moeda_br)
            resumo["Maior Preço"] = resumo["Maior Preço"].apply(formatar_moeda_br)
            resumo["Preço Médio"] = resumo["Preço Médio"].apply(formatar_moeda_br)

            st.dataframe(resumo, use_container_width=True, hide_index=True)

    with tab_evidencias:
        screenshots = [r for r in resultados if r.get("screenshot") and os.path.exists(r.get("screenshot", ""))]

        if screenshots:
            col_ev_info, col_ev_pdf = st.columns([3, 1])
            with col_ev_info:
                st.markdown(f"**{len(screenshots)} evidências visuais capturadas**")
            with col_ev_pdf:
                pdf_ev_data = gerar_pdf_evidencias(resultados)
                st.download_button(
                    label="📑 Exportar Evidências em PDF",
                    data=pdf_ev_data,
                    file_name=f"evidencias_scraping_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    type="primary",
                )
            for r in screenshots:
                with st.expander(f"📸 {r['item']} — {r['dominio']} — {formatar_moeda_br(r['preco'])}"):
                    sc_path = r["screenshot"]
                    if sc_path.endswith(".html"):
                        # Renderizar snapshot HTML como evidência
                        try:
                            with open(sc_path, "r", encoding="utf-8") as f:
                                html_content = f.read()
                            import streamlit.components.v1 as components
                            components.html(html_content, height=500, scrolling=True)
                            # Botão de download do HTML
                            st.download_button(
                                label="⬇️ Baixar evidência",
                                data=html_content,
                                file_name=f"evidencia_{r['item']}_{r['dominio']}.html",
                                mime="text/html",
                                key=f"dl_html_{r['item']}_{r['dominio']}_{id(r)}",
                            )
                        except Exception:
                            st.markdown(f"📄 Evidência salva em: `{sc_path}`")
                    else:
                        st.image(sc_path, caption=f"{r['dominio']} - {r['data_coleta']}")
                        # Botão de download da imagem PNG
                        try:
                            with open(sc_path, "rb") as img_file:
                                img_bytes = img_file.read()
                            st.download_button(
                                label="⬇️ Baixar screenshot",
                                data=img_bytes,
                                file_name=f"screenshot_{r['item']}_{r['dominio']}.png",
                                mime="image/png",
                                key=f"dl_img_{r['item']}_{r['dominio']}_{id(r)}",
                            )
                        except Exception:
                            pass
                    st.markdown(f"**Link:** [{r['url']}]({r['url']})")
        else:
            st.info(
                "Nenhuma evidência visual capturada nesta execução."
            )

    with tab_export:
        st.markdown("#### 📥 Exportar Relatório")

        col_exp1, col_exp2, col_exp3, col_exp4 = st.columns(4)

        with col_exp1:
            excel_data = gerar_relatorio_excel(resultados)
            st.download_button(
                label="📊 Baixar Excel",
                data=excel_data,
                file_name=f"relatorio_scraping_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        with col_exp2:
            csv_data = gerar_relatorio_csv(resultados)
            if csv_data:
                st.download_button(
                    label="📄 Baixar CSV",
                    data=csv_data,
                    file_name=f"relatorio_scraping_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

        with col_exp3:
            json_data = json.dumps(resultados, ensure_ascii=False, indent=2, default=str)
            st.download_button(
                label="🔗 Baixar JSON",
                data=json_data,
                file_name=f"relatorio_scraping_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                mime="application/json",
                use_container_width=True,
            )

        with col_exp4:
            pdf_data = gerar_pdf_evidencias(resultados)
            st.download_button(
                label="📑 Baixar Evidências PDF",
                data=pdf_data,
                file_name=f"evidencias_scraping_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

elif "scraping_resultados" in st.session_state and not st.session_state["scraping_resultados"]:
    st.warning("⚠️ O scraping foi executado mas nenhum orçamento foi encontrado. Tente com outros termos.")
