"""Sugestão automática de códigos CATMAT e CATSERV a partir de descrições."""

from __future__ import annotations

import io
import json
import os
import re
import unicodedata
import base64
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
import requests
import streamlit as st
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="CATMAT/CATSERV Automático",
    page_icon="🔎",
    layout="wide",
    initial_sidebar_state="expanded",
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CATALOGO_DIR = os.path.join(BASE_DIR, "Projeto Adesões")
PDM_PATH = os.path.join(CATALOGO_DIR, "catalogo_pdm.json")
CATSERV_PATH = os.path.join(CATALOGO_DIR, "catalogo_servicos.json")
CATMAT_API_URL = "https://dadosabertos.compras.gov.br/modulo-material/4_consultarItemMaterial"
STOP_WORDS = {"a", "as", "com", "da", "das", "de", "do", "dos", "e", "em", "para", "por", "sem", "um", "uma"}
TERMOS_RESTRITIVOS = {
    "automotivo", "cartucho", "descartavel", "hospitalar", "impressora", "industrial",
    "infantil", "medico", "odontologico", "recarga", "refil", "tinteiro", "toner",
}
EXPANSOES_DE_BUSCA = {
    "caneta": "caneta esferografica",
    "fita crepe": "fita crepe adesiva",
}
FALLBACKS_GENERICOS = {
    "caneta": {"codigo_pdm": "99", "termos_preferidos": {"esferografica"}},
    "fita crepe": {"codigo_pdm": "18071", "termos_preferidos": {"papel", "crepado"}},
}
LIMIAR_SIMILARIDADE = 45.0


st.markdown(
    """
    <style>
        html, body, [data-testid="stAppViewContainer"], .stApp {
            background: #001a4d !important;
            color: #f8fafc;
        }
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0a0a0a 0%, #111111 50%, #0a0a0a 100%) !important;
            border-right: 3px solid #d4af37 !important;
        }
        [data-testid="stSidebarNav"] { display: none !important; }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] {
            background: linear-gradient(135deg, #1a1a1a, #252525) !important;
            border: 1px solid #333 !important; border-radius: 8px !important;
            color: #fff !important; margin: .2rem 0 !important; padding: .45rem .7rem !important;
            font-size: 12.5px !important; font-weight: 600 !important; justify-content: center !important;
        }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] span { color: #fff !important; }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] {
            background: #d4af37 !important; border-color: #d4af37 !important;
        }
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] span { color: #0a0a0a !important; }
        .catalog-header {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            border: 1px solid rgba(96, 165, 250, .45); border-radius: 10px;
            padding: 1.35rem 1.55rem; margin-bottom: 1rem;
            box-shadow: 0 10px 28px rgba(0, 0, 0, .28);
        }
        .catalog-header-top { display: flex; align-items: center; gap: .85rem; flex-wrap: wrap; }
        .catalog-symbol { width: 2.8rem; height: 2.8rem; display: flex; justify-content: center; align-items: center; background: #d4af37; border-radius: 7px; font-size: 1.35rem; }
        .catalog-header h1 { color: #fff; margin: 0; font-size: 1.35rem; }
        .catalog-header p { color: #bfdbfe; margin: .18rem 0 0; font-size: .85rem; }
        .catalog-badge { color: #d4af37; font-size: .7rem; font-weight: 800; letter-spacing: .08em; margin-left: auto; }
        .input-panel, .result-panel {
            background: rgba(8, 26, 57, .78); border: 1px solid #1e5b9f; border-radius: 9px;
            padding: 1rem 1.1rem; margin: .5rem 0 1rem;
        }
        .input-panel h3, .result-panel h3 { color: #d4af37; font-size: .95rem; margin: 0 0 .25rem; }
        .input-panel p, .result-panel p { color: #b6cae2; font-size: .78rem; margin: 0; }
        .status-chip { display: inline-block; border-radius: 999px; padding: .2rem .55rem; font-size: .7rem; font-weight: 700; }
        .status-good { background: rgba(34, 197, 94, .18); color: #86efac; border: 1px solid rgba(34, 197, 94, .4); }
        .status-review { background: rgba(245, 158, 11, .15); color: #fcd34d; border: 1px solid rgba(245, 158, 11, .4); }
        .status-low { background: rgba(239, 68, 68, .14); color: #fca5a5; border: 1px solid rgba(239, 68, 68, .4); }
        [data-testid="stTextArea"] textarea, [data-testid="stSelectbox"] div[data-baseweb="select"] > div {
            background: rgba(7, 20, 42, .8) !important; border-color: #2b6cb0 !important; color: #f8fafc !important;
        }
        [data-testid="stDataFrame"] { border: 1px solid #1e5b9f; border-radius: 8px; overflow: hidden; }
        .sidebar-footer { color: #666; font-size: 11px; text-align: center; padding: 1rem 0; border-top: 1px solid #333; margin-top: 2rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

acanto_path = os.path.join(CATALOGO_DIR, "acanto.png")
if os.path.exists(acanto_path):
    with open(acanto_path, "rb") as arquivo_acanto:
        acanto_b64 = base64.b64encode(arquivo_acanto.read()).decode()
else:
    acanto_b64 = None

with st.sidebar:
    if acanto_b64:
        st.markdown(f'<div style="text-align:center;padding:1rem 0 0.5rem 0;"><img src="data:image/png;base64,{acanto_b64}" style="max-width:70%;height:auto;"></div>', unsafe_allow_html=True)
    st.markdown("## MENU")
    st.markdown("---")
    st.page_link("streamlit_app.py", label="Cotação", icon="⚓")
    st.page_link("pages/Detalhes_Compra.py", label="Detalhes Compra", icon="🔍")
    st.page_link("pages/Adesões.py", label="Adesões", icon="🤝")
    st.page_link("pages/Notas_Fiscais.py", label="Notas Fiscais", icon="📄")
    st.page_link("pages/Banco_de_Fornecedores.py", label="Fornecedores", icon="🏢")
    st.page_link("pages/Consulta.py", label="Consulta CNPJ", icon="💻")
    st.page_link("pages/Web_Scraping.py", label="Web Scraping", icon="🕷️")
    st.page_link("pages/O_Babilaca_(IA).py", label="O Babilaca (IA)", icon="🧠")
    st.page_link("pages/Calculo_IPCA.py", label="Cálculo IPCA", icon="📊")
    st.page_link("pages/CATMAT_CATSERV_Automatico.py", label="CATMAT/CATSERV", icon="🔎")
    st.markdown("---")
    st.markdown("## LINKS ÚTEIS")
    st.markdown("""<div style="margin-bottom:0.6rem;">
        <a href="https://detetive-obtencao.vercel.app/" target="_blank" style="color:#cbd5e1;text-decoration:none;font-size:0.9rem;display:flex;align-items:center;gap:0.5rem;">🚨 Detetive Obtenção</a>
    </div>
    <div style="margin-bottom:1rem;">
        <a href="https://depurador.streamlit.app/" target="_blank" style="color:#cbd5e1;text-decoration:none;font-size:0.9rem;display:flex;align-items:center;gap:0.5rem;">🧾 Depurador de Orçamentos</a>
    </div>""", unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;color:#d4af37;font-size:10px;font-weight:600;padding:0.3rem 0;white-space:nowrap;">Centro de Operações do Abastecimento</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-footer">Marinha do Brasil<br>AtaCotada v1.0</div>', unsafe_allow_html=True)


def _normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto)).encode("ASCII", "ignore").decode("ASCII")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", texto.lower())).strip()


def _tokens(texto: str) -> set[str]:
    return {token for token in _normalizar(texto).split() if len(token) > 1 and token not in STOP_WORDS}


def _termo_principal(texto: str) -> str:
    """Retorna o primeiro termo informativo, que representa a categoria principal do item."""
    return next((token for token in _normalizar(texto).split() if len(token) > 1 and token not in STOP_WORDS), "")


def _texto_pdf(texto: object) -> str:
    """Converte texto do catálogo para caracteres aceitos pela fonte padrão do PDF."""
    return unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode("ascii")


def _texto_pdf_quebravel(texto: object, tamanho_maximo: int = 45) -> str:
    """Insere espaços em tokens longos para que o FPDF consiga quebrar as linhas."""
    linhas = []
    for linha in _texto_pdf(texto).splitlines() or [""]:
        palavras = []
        for palavra in linha.split(" "):
            partes = [palavra[indice:indice + tamanho_maximo] for indice in range(0, len(palavra), tamanho_maximo)]
            palavras.append(" ".join(partes))
        linhas.append(" ".join(palavras))
    return "\n".join(linhas)


@st.cache_data(show_spinner=False)
def carregar_catalogo(caminho: str) -> list[dict[str, object]]:
    with open(caminho, "r", encoding="utf-8") as arquivo:
        catalogo = json.load(arquivo)
    return [
        {"codigo": str(codigo), "descricao": descricao, "tokens": sorted(_tokens(descricao))}
        for descricao, codigo in catalogo.items()
    ]


@st.cache_data(ttl=3600, show_spinner=False)
def buscar_catmat_api(consulta: str) -> list[dict[str, str]]:
    """Busca candidatos do catálogo público do Compras.gov por descrição."""
    termos = list(_tokens(consulta))
    consultas = [consulta, EXPANSOES_DE_BUSCA.get(_normalizar(consulta), "")]
    consultas.extend(termo for termo in termos if len(termo) >= 4)
    candidatos = {}
    for texto_consulta in dict.fromkeys(texto for texto in consultas if texto):
        try:
            resposta = requests.get(
                CATMAT_API_URL,
                params={"pagina": 1, "tamanhoPagina": 100, "descricaoItem": texto_consulta[:180], "statusItem": "true"},
                timeout=15,
            )
            if resposta.status_code != 200:
                continue
            for item in resposta.json().get("resultado", []):
                codigo = str(item.get("codigoItem", ""))
                descricao = item.get("descricaoItem", "")
                if codigo and descricao:
                    candidatos[codigo] = {
                        "codigo": codigo,
                        "descricao": descricao,
                        "codigo_pdm": str(item.get("codigoPdm", "")),
                        "descricao_pdm": item.get("nomePdm", ""),
                    }
        except requests.RequestException:
            continue
    return list(candidatos.values())


@st.cache_data(ttl=3600, show_spinner=False)
def buscar_itens_catmat_por_pdm(codigo_pdm: str) -> list[dict[str, str]]:
    """Retorna itens CATMAT oficiais associados a um PDM, sem expor o PDM como resultado."""
    try:
        resposta = requests.get(
            CATMAT_API_URL,
            params={"pagina": 1, "tamanhoPagina": 100, "codigoPdm": codigo_pdm, "statusItem": "true"},
            timeout=15,
        )
        if resposta.status_code != 200:
            return []
        return [
            {
                "codigo": str(item.get("codigoItem", "")),
                "descricao": item.get("descricaoItem", ""),
                "codigo_pdm": str(item.get("codigoPdm", "")),
                "descricao_pdm": item.get("nomePdm", ""),
            }
            for item in resposta.json().get("resultado", [])
            if item.get("codigoItem") and item.get("descricaoItem")
        ]
    except requests.RequestException:
        return []


@st.cache_data(ttl=3600, show_spinner=False)
def buscar_unidade_fornecimento(codigo_pdm: str) -> str:
    """Retorna a primeira unidade de fornecimento ativa associada ao PDM informado."""
    try:
        resposta = requests.get(
            "https://dadosabertos.compras.gov.br/modulo-material/6_consultarMaterialUnidadeFornecimento",
            params={"pagina": 1, "tamanhoPagina": 100, "codigoPdm": codigo_pdm, "statusUnidadeFornecimentoPdm": "true"},
            timeout=15,
        )
        if resposta.status_code != 200:
            return ""
        unidades = resposta.json().get("resultado", [])
        if not unidades:
            return ""
        unidade = unidades[0]
        sigla = unidade.get("siglaUnidadeFornecimento", "")
        nome = unidade.get("nomeUnidadeFornecimento", "")
        return f"{sigla} - {nome}".strip(" -")
    except requests.RequestException:
        return ""


def calcular_similaridade(descricao: str, candidato: str) -> float:
    origem = _tokens(descricao)
    destino = _tokens(candidato)
    if not origem or not destino:
        return 0.0
    em_comum = origem & destino
    cobertura = len(em_comum) / len(origem)
    precisao = len(em_comum) / len(destino)
    sequencia = SequenceMatcher(None, _normalizar(descricao), _normalizar(candidato)).ratio()
    termo_principal = _termo_principal(descricao)
    restritivos_ausentes = (destino - origem) & TERMOS_RESTRITIVOS
    termos_extras = len(destino - origem) / len(destino)
    penalidade = min(36, len(restritivos_ausentes) * 18) + (termos_extras * 12)
    bonus_principal = 25 if termo_principal in destino else -50
    pontuacao = cobertura * 65 + precisao * 25 + sequencia * 10 + bonus_principal - penalidade
    return round(max(0, min(100, pontuacao)), 1)


def melhores_do_catalogo(descricao: str, catalogo: list[dict[str, object]], limite: int = 40) -> list[dict[str, object]]:
    tokens_origem = _tokens(descricao)
    candidatos = []
    for item in catalogo:
        tokens_destino = set(item["tokens"])
        em_comum = len(tokens_origem & tokens_destino)
        if em_comum:
            candidatos.append((em_comum / len(tokens_origem), item))
    candidatos.sort(key=lambda candidato: candidato[0], reverse=True)
    return [item for _, item in candidatos[:limite]]


def _pdm_generico(descricao: str, catalogo_pdm: list[dict[str, object]]) -> dict[str, object] | None:
    descricao_normalizada = _normalizar(descricao)
    termo_principal = _termo_principal(descricao)
    tokens_descricao = set(descricao_normalizada.split())
    perfis_compativeis = [
        (set(chave.split()), fallback)
        for chave, fallback in FALLBACKS_GENERICOS.items()
        if set(chave.split()).issubset(tokens_descricao)
    ]
    if perfis_compativeis:
        _, fallback = max(perfis_compativeis, key=lambda perfil: len(perfil[0]))
        return {
            "codigo": fallback["codigo_pdm"],
            "descricao": descricao_normalizada,
            "termos_preferidos": fallback["termos_preferidos"],
        }
    pdm_exato = next(
        (item for item in catalogo_pdm if _normalizar(str(item["descricao"])) == termo_principal),
        None,
    )
    if pdm_exato:
        return {**pdm_exato, "termos_preferidos": {termo_principal}}
    if len(_tokens(descricao)) > 3:
        return None
    candidatos = melhores_do_catalogo(descricao, catalogo_pdm, limite=30)
    if not candidatos:
        return None
    return max(candidatos, key=lambda candidato: calcular_similaridade(descricao, str(candidato["descricao"])))


def sugerir_codigo(descricao: str, catalogo_pdm: list[dict[str, object]], catalogo_servico: list[dict[str, object]], tipo: str) -> dict[str, object]:
    opcoes: list[dict[str, object]] = []
    tipos_consulta = [tipo] if tipo != "Automático" else ["Material", "Serviço"]
    for tipo_atual in tipos_consulta:
        origem = "CATMAT específico"
        termos_preferidos = set()
        if tipo_atual == "Material":
            candidatos = buscar_catmat_api(descricao)
            termo_principal = _termo_principal(descricao)
            candidatos = [
                candidato for candidato in candidatos
                if termo_principal in _tokens(str(candidato["descricao"]))
            ]
            pdm = _pdm_generico(descricao, catalogo_pdm) if not candidatos else None
            if pdm:
                candidatos = buscar_itens_catmat_por_pdm(str(pdm["codigo"]))
                origem = "CATMAT genérico"
                termos_preferidos = set(pdm.get("termos_preferidos", []))
        else:
            candidatos = melhores_do_catalogo(descricao, catalogo_servico)
            origem = "CATSERV"
        vistos = set()
        for candidato in candidatos:
            codigo = str(candidato["codigo"])
            if codigo in vistos:
                continue
            vistos.add(codigo)
            tokens_candidato = _tokens(str(candidato["descricao"]))
            termos_restritivos_ausentes = (tokens_candidato - _tokens(descricao)) & TERMOS_RESTRITIVOS
            if origem == "CATMAT genérico" and termos_restritivos_ausentes:
                continue
            similaridade = calcular_similaridade(descricao, str(candidato["descricao"]))
            if termos_preferidos:
                similaridade = min(100, similaridade + 18 * len(tokens_candidato & termos_preferidos))
            codigo_pdm = str(candidato.get("codigo_pdm", "")) if tipo_atual == "Material" else ""
            opcoes.append(
                {
                    "tipo": tipo_atual,
                    "codigo": codigo,
                    "descricao_catalogo": str(candidato["descricao"]),
                    "similaridade": similaridade,
                    "origem": origem,
                    "unidade_fornecimento": buscar_unidade_fornecimento(codigo_pdm) if codigo_pdm else "",
                    "codigo_pdm": codigo_pdm,
                    "descricao_pdm": str(candidato.get("descricao_pdm", "")) if codigo_pdm else "",
                }
            )
    if not opcoes:
        return {"tipo": "-", "codigo": "-", "descricao_catalogo": "Nenhuma correspondência encontrada", "similaridade": 0.0, "origem": "-", "unidade_fornecimento": "", "codigo_pdm": "", "descricao_pdm": ""}
    melhor_opcao = max(opcoes, key=lambda opcao: opcao["similaridade"])
    limiar = 35.0 if melhor_opcao["origem"] == "CATMAT genérico" else LIMIAR_SIMILARIDADE
    if melhor_opcao["similaridade"] < limiar:
        return {"tipo": "-", "codigo": "-", "descricao_catalogo": "Descrição insuficiente para sugerir um código com segurança", "similaridade": melhor_opcao["similaridade"], "origem": "-", "unidade_fornecimento": "", "codigo_pdm": "", "descricao_pdm": ""}
    return melhor_opcao


def gerar_excel(resultados: pd.DataFrame) -> bytes:
    saida = io.BytesIO()
    with pd.ExcelWriter(saida, engine="openpyxl") as escritor:
        resultados.to_excel(escritor, index=False, sheet_name="Correlação")
        planilha = escritor.book["Correlação"]
        cabecalho = PatternFill("solid", fgColor="001A4D")
        for celula in planilha[1]:
            celula.font = Font(color="FFFFFF", bold=True)
            celula.fill = cabecalho
            celula.alignment = Alignment(horizontal="center", vertical="center")
        for coluna in planilha.columns:
            indice = coluna[0].column
            maior = max(len(str(celula.value or "")) for celula in coluna)
            planilha.column_dimensions[get_column_letter(indice)].width = min(max(maior + 2, 14), 62)
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions
    return saida.getvalue()


class RelatorioPDF(FPDF):
    def header(self) -> None:
        self.set_fill_color(0, 26, 77)
        self.rect(0, 0, 210, 22, "F")
        self.set_text_color(212, 175, 55)
        self.set_font("Helvetica", "B", 14)
        self.set_xy(12, 8)
        self.cell(0, 7, "CATMAT/CATSERV - Correlacao Automatica")
        self.ln(20)


def gerar_pdf(resultados: pd.DataFrame) -> bytes:
    pdf = RelatorioPDF()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.add_page()
    pdf.set_text_color(40, 40, 40)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(0, 5, "Sugestoes calculadas a partir do catalogo publico do Compras.gov. Revise a descricao e o codigo antes de utilizar no processo.")
    pdf.ln(3)
    for indice, linha in resultados.iterrows():
        pdf.set_fill_color(235, 242, 252)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 7, _texto_pdf_quebravel(f"{indice + 1}. {linha['Descrição informada']}"), fill=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 5, _texto_pdf_quebravel(f"{linha['Tipo']} {linha['Código']}  |  Similaridade: {linha['Similaridade (%)']}%"))
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 5, _texto_pdf_quebravel(f"Sugestao: {linha['Descrição sugerida']}"))
        if linha.get("Código PDM", ""):
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(0, 5, _texto_pdf_quebravel(f"PDM: {linha['Código PDM']} - {linha['Descrição PDM']}"))
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(0, 5, _texto_pdf_quebravel(f"Unidade de fornecimento: {linha['Unidade de fornecimento']}"))
        pdf.ln(3)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - AtaCotada", ln=True)
    return bytes(pdf.output())


def ler_lista_enviada(arquivo: object) -> tuple[list[str], list[str]]:
    nome = getattr(arquivo, "name", "").lower()
    if nome.endswith(".csv"):
        dados = pd.read_csv(arquivo)
    else:
        dados = pd.read_excel(arquivo)
    colunas = [str(coluna) for coluna in dados.columns]
    return colunas, dados.astype(str).fillna("").to_dict("list")


st.markdown(
    """
    <div class="catalog-header">
        <div class="catalog-header-top">
            <div class="catalog-symbol">🔎</div>
            <div><h1>CATMAT/CATSERV Automático</h1><p>Encontre sugestões de classificação para materiais e serviços a partir das descrições do seu processo.</p></div>
            <span class="catalog-badge">CATÁLOGO COMPRAS.GOV</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

catalogo_pdm = carregar_catalogo(PDM_PATH)
catalogo_servico = carregar_catalogo(CATSERV_PATH)

st.markdown('<div class="input-panel"><h3>Lista de itens</h3><p>Digite ou cole uma descrição por linha. Você também pode importar uma planilha CSV ou Excel.</p></div>', unsafe_allow_html=True)
entrada_manual = st.text_area(
    "Descrições dos itens",
    placeholder="Ex: Papel sulfite A4 75 g/m², resma com 500 folhas\nManutenção preventiva de aparelhos de ar-condicionado\nCadeira ergonômica para escritório",
    height=190,
)

col_tipo, col_arquivo = st.columns([1, 2])
with col_tipo:
    tipo_busca = st.selectbox("Classificar como", ["Automático", "Material", "Serviço"])
with col_arquivo:
    arquivo_lista = st.file_uploader("Importar lista (CSV ou Excel)", type=["csv", "xlsx", "xls"])

itens_arquivo: list[str] = []
if arquivo_lista:
    try:
        colunas_arquivo, dados_arquivo = ler_lista_enviada(arquivo_lista)
        coluna_descricao = st.selectbox("Coluna com as descrições", colunas_arquivo)
        itens_arquivo = [valor.strip() for valor in dados_arquivo[coluna_descricao] if valor and valor.strip().lower() != "nan"]
        st.caption(f"{len(itens_arquivo)} item(ns) identificado(s) no arquivo.")
    except Exception as erro:
        st.error(f"Não foi possível ler o arquivo: {erro}")

if st.button("🔎 Encontrar códigos sugeridos", type="primary", use_container_width=True):
    itens_manuais = [linha.strip(" -•\t") for linha in entrada_manual.splitlines() if linha.strip()]
    itens = list(dict.fromkeys(itens_manuais + itens_arquivo))
    if not itens:
        st.warning("Informe ao menos uma descrição ou envie uma lista de itens.")
    else:
        with st.spinner(f"Analisando {len(itens)} item(ns) nos catálogos oficiais..."):
            resultados_brutos = []
            for item in itens:
                sugestao = sugerir_codigo(item, catalogo_pdm, catalogo_servico, tipo_busca)
                resultados_brutos.append(
                    {
                        "Descrição informada": item,
                        "Tipo": "CATMAT" if sugestao["tipo"] == "Material" else "CATSERV" if sugestao["tipo"] == "Serviço" else "-",
                        "Código": sugestao["codigo"],
                        "Descrição sugerida": sugestao["descricao_catalogo"],
                        "Similaridade (%)": sugestao["similaridade"],
                        "Unidade de fornecimento": sugestao["unidade_fornecimento"],
                        "Código PDM": sugestao["codigo_pdm"],
                        "Descrição PDM": sugestao["descricao_pdm"],
                    }
                )
        st.session_state["catmat_catserv_resultados"] = resultados_brutos

resultados_salvos = st.session_state.get("catmat_catserv_resultados")
if resultados_salvos:
    resultados = pd.DataFrame(resultados_salvos)
    alta = int((resultados["Similaridade (%)"] >= 70).sum())
    media = int(((resultados["Similaridade (%)"] >= 45) & (resultados["Similaridade (%)"] < 70)).sum())
    baixa = len(resultados) - alta - media
    st.markdown('<div class="result-panel"><h3>Resultado da correlação</h3><p>Use a similaridade como apoio à decisão e confira as especificações do item no catálogo antes de utilizar o código.</p></div>', unsafe_allow_html=True)
    metricas = st.columns(4)
    metricas[0].metric("Itens analisados", len(resultados))
    metricas[1].markdown(f'<span class="status-chip status-good">{alta} alta similaridade</span>', unsafe_allow_html=True)
    metricas[2].markdown(f'<span class="status-chip status-review">{media} para revisar</span>', unsafe_allow_html=True)
    metricas[3].markdown(f'<span class="status-chip status-low">{baixa} baixa similaridade</span>', unsafe_allow_html=True)
    st.dataframe(resultados, use_container_width=True, hide_index=True, column_config={"Similaridade (%)": st.column_config.NumberColumn(format="%.1f%%")})
    excel, pdf = st.columns(2)
    with excel:
        st.download_button("⬇️ Baixar correlação em Excel", gerar_excel(resultados), "correlacao_catmat_catserv.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with pdf:
        st.download_button("⬇️ Baixar correlação em PDF", gerar_pdf(resultados), "correlacao_catmat_catserv.pdf", "application/pdf", use_container_width=True)