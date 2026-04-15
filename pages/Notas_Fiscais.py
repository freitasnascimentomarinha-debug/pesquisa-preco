import base64
import streamlit as st
import requests
import pandas as pd
import os
import re
import io
import tempfile
from datetime import datetime

# Configuração da página
st.set_page_config(
    page_title="AtaCotada - Notas Fiscais",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado
st.markdown("""
    <style>
        /* Anti-flash: forçar fundo escuro desde o início */
        html, body, [data-testid="stAppViewContainer"],
        .main, [data-testid="stApp"], .stApp {
            background-color: #001a4d !important;
            color: #ffffff !important;
        }
        
        /* Transição suave ao carregar */
        .stApp {
            animation: fadeIn 0.3s ease-in;
        }
        
        @keyframes fadeIn {
            from { opacity: 0.7; }
            to { opacity: 1; }
        }
        
        /* Header customizado */
        .header-container {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            padding: 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
            text-align: center;
        }
        
        .logo-text {
            color: #ffffff;
            font-size: 14px;
            font-weight: 600;
            letter-spacing: 2px;
            margin-bottom: 0.5rem;
        }
        
        .sistema-nome {
            color: #d4af37;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 3px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);
            margin: 0.5rem 0;
            font-family: 'Arial Black', sans-serif;
        }
        
        .subtitulo {
            color: #ffffff;
            font-size: 14px;
            margin-top: 0.5rem;
            letter-spacing: 1px;
        }

        /* Filtros container */
        .filtros-container {
            background: linear-gradient(135deg, #0a2540 0%, #164863 100%);
            padding: 1.5rem;
            border-radius: 8px;
            margin-bottom: 1.5rem;
            border-left: 5px solid #d4af37;
        }

        /* Cards de estatísticas */
        .stats-card {
            background: linear-gradient(135deg, #0a2540 0%, #0f4c75 100%);
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 5px solid #d4af37;
            margin-bottom: 1rem;
        }

        /* Botões */
        .stButton > button {
            background-color: #d4af37;
            color: #ffffff;
            border: none;
            border-radius: 6px;
            padding: 0.75rem 1.5rem;
            font-weight: bold;
            font-size: 16px;
            cursor: pointer;
            transition: all 0.3s ease;
        }

        .stButton > button:hover {
            background-color: #ffd700;
            color: #ffffff;
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(212, 175, 55, 0.3);
        }

        /* Botões de download */
        .stDownloadButton > button {
            background-color: #d4af37 !important;
            color: #ffffff !important;
            border: none !important;
        }

        .stDownloadButton > button:hover {
            background-color: #ffd700 !important;
            color: #ffffff !important;
        }

        /* Inputs */
        .stTextInput > div > div > input,
        .stNumberInput > div > div > input {
            background-color: #0a2540 !important;
            color: #ffffff !important;
            border: 2px solid #0033cc !important;
            border-radius: 6px !important;
        }

        label {
            color: #ffffff !important;
        }

        /* Títulos */
        h1, h2, h3 {
            color: #d4af37;
            text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.5);
        }

        /* Tabelas */
        [data-testid="stDataFrame"] {
            background-color: #ffffff !important;
            border-radius: 8px;
            overflow: hidden;
        }

        th {
            background-color: #ffffff !important;
            color: #333333 !important;
            font-weight: bold;
            border-bottom: 2px solid #d4af37 !important;
        }

        td {
            background-color: #ffffff !important;
            color: #333333 !important;
        }

        tr:hover {
            background-color: #f5f5f5 !important;
        }
        
        /* ===== SIDEBAR MODERNA - PRETA COM BORDA DOURADA ===== */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0a0a0a 0%, #111111 50%, #0a0a0a 100%) !important;
            border-right: 3px solid #d4af37 !important;
            box-shadow: 4px 0 15px rgba(0, 0, 0, 0.5);
        }

        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            background: transparent !important;
        }

        /* Esconder navegação padrão do Streamlit */
        [data-testid="stSidebarNav"] {
            display: none !important;
        }

        /* Título da sidebar */
        [data-testid="stSidebar"] .stMarkdown h2 {
            color: #d4af37 !important;
            font-family: 'Arial Black', sans-serif;
            font-size: 22px;
            text-align: center;
            letter-spacing: 2px;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.8);
            border-bottom: 2px solid #d4af37;
            padding-bottom: 0.75rem;
            margin-bottom: 1.5rem;
        }

        /* Links de navegação na sidebar */
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] {
            background: linear-gradient(135deg, #1a1a1a 0%, #252525 100%) !important;
            color: #ffffff !important;
            border: 1px solid #333333 !important;
            border-radius: 10px !important;
            margin: 0.4rem 0 !important;
            padding: 0.85rem 1.2rem !important;
            font-weight: 600 !important;
            font-size: 15px !important;
            transition: all 0.3s ease !important;
            text-decoration: none !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] span {
            color: #ffffff !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover {
            background: linear-gradient(135deg, #252525 0%, #353535 100%) !important;
            border: 1px solid #d4af37 !important;
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.25) !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover span {
            color: #d4af37 !important;
        }

        /* Link ativo / página atual */
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] {
            background: linear-gradient(135deg, #d4af37 0%, #c5a028 100%) !important;
            color: #0a0a0a !important;
            border: 1px solid #d4af37 !important;
            font-weight: bold !important;
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.4) !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] span {
            color: #0a0a0a !important;
        }

        /* Texto e labels na sidebar */
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stText,
        [data-testid="stSidebar"] p {
            color: #ffffff !important;
        }

        /* Separadores na sidebar */
        [data-testid="stSidebar"] hr {
            border-color: #333333 !important;
            margin: 1rem 0 !important;
        }

        /* Rodapé da sidebar */
        .sidebar-footer {
            color: #666666;
            font-size: 11px;
            text-align: center;
            padding: 1rem 0;
            border-top: 1px solid #333333;
            margin-top: 2rem;
        }
    </style>
""", unsafe_allow_html=True)

# ===== SIDEBAR - Navegação customizada =====
# Carregar imagem do acanto para a sidebar
_acanto_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Projeto Adesões", "acanto.png")
if os.path.exists(_acanto_path):
    with open(_acanto_path, "rb") as _f:
        _acanto_b64 = base64.b64encode(_f.read()).decode()
else:
    _acanto_b64 = None

with st.sidebar:
    if _acanto_b64:
        st.markdown(f'<div style="text-align:center;padding:1rem 0 0.5rem 0;"><img src="data:image/png;base64,{_acanto_b64}" style="max-width:70%;height:auto;"></div>', unsafe_allow_html=True)
    st.markdown("## MENU")
    st.markdown("---")
    st.page_link("streamlit_app.py", label="Cotação", icon="⚓")
    st.page_link("pages/Detalhes_Compra.py", label="Detalhes Compra", icon="🔍")
    st.page_link("pages/Adesões.py", label="Adesões", icon="🤝")
    st.page_link("pages/Notas_Fiscais.py", label="Notas Fiscais", icon="📄")
    st.page_link("pages/Banco_de_Fornecedores.py", label="Fornecedores", icon="🏢")
    st.page_link("pages/Consulta.py", label="Consulta CNPJ", icon="💻")
    st.page_link("pages/Web_Scraping.py", label="Web Scraping", icon="🕷️")
    st.page_link("pages/O_Babilaca_(IA).py", label="O Babilaca (IA)", icon="🧠")
    st.markdown("---")
    st.markdown("## LINKS ÚTEIS")
    st.markdown("""
    <div style="margin-bottom: 1rem;">
        <a href="https://freitasnascimentomarinha-debug.github.io/ShootMail/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            📧 Disparador de Emails
        </a>
    </div>
    <div style="margin-bottom: 1rem;">
        <a href="https://detetive-obtencao.vercel.app/" target="_blank" style="color: #cbd5e1; text-decoration: none; font-size: 0.9rem; display: flex; align-items: center; gap: 0.5rem;">
            🚨 Detetive Obtenção
        </a>
    </div>
    """, unsafe_allow_html=True)
    st.markdown('<div style="text-align:center;color:#d4af37;font-size:10px;font-weight:600;padding:0.3rem 0;white-space:nowrap;">Centro de Operações do Abastecimento</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-footer">Marinha do Brasil<br>AtaCotada v1.0</div>', unsafe_allow_html=True)

# Header
st.markdown("""
    <div class="header-container">
        <div class="logo-text">⚓ MARINHA DO BRASIL ⚓</div>
        <div class="sistema-nome">AtaCotada</div>
        <div class="subtitulo">Notas Fiscais</div>
    </div>
""", unsafe_allow_html=True)

# ===== CONTEÚDO - NOTAS FISCAIS =====
st.title("📄 Notas Fiscais")
st.markdown("Consulta de notas fiscais eletrônicas a partir dos dados abertos do **Portal da Transparência**.")

# ===== ABAS PRINCIPAIS =====
tab_pesquisa, tab_consulta_nfe = st.tabs([
    "🔍 Pesquisa de NF (Portal da Transparência)",
    "📋 Consulta / Download NFe (Receita Federal)"
])

# ==================== ABA 2: CONSULTA NFe (RECEITA FEDERAL) ====================
NFE_URL = "https://www.nfe.fazenda.gov.br/portal/consultaRecaptcha.aspx?tipoConsulta=resumo&tipoConteudo=7PhJ+gAVw2g="

with tab_consulta_nfe:
    st.markdown("### 📋 Consulta de NFe — Receita Federal")

    st.markdown("")
    st.link_button("🌐 Acessar Portal da NFe (Receita Federal)", NFE_URL, use_container_width=True)
    st.markdown("")

    st.markdown(f"""
    <div style="background: #0a2540; border: 1px solid #333; border-radius: 8px; padding: 1.2rem; margin-top: 0.5rem;">
        <p style="color: #d4af37; font-weight: bold; font-size: 15px; margin-bottom: 0.8rem;">📌 Como consultar e baixar uma NFe:</p>
        <ol style="color: #cccccc; font-size: 13px; line-height: 2;">
            <li>Clique no botão acima — o portal da Receita Federal abrirá em uma <b>nova aba</b> do navegador.</li>
            <li>No portal, insira a <b>chave de acesso</b> da NFe (44 dígitos numéricos, encontrada no DANFE ou no corpo da nota).</li>
            <li>Resolva o <b>captcha</b> e clique em <b>"Continuar"</b>.</li>
            <li>Na tela de resultado, você poderá visualizar o resumo da nota e baixar o <b>DANFE (PDF)</b> ou o <b>XML</b>.</li>
            <li>Para consultar outra NFe, basta repetir o processo na mesma aba.</li>
        </ol>
        <hr style="border-color: #333; margin: 1rem 0;">
        <p style="color: #aaa; font-size: 12px; margin: 0;">
            💡 <b>Dica:</b> A chave de acesso possui 44 dígitos e identifica unicamente cada Nota Fiscal Eletrônica.
            Ela pode ser encontrada no campo "Chave de Acesso" do DANFE impresso ou nos dados da nota no sistema emissor.
        </p>
    </div>
    """, unsafe_allow_html=True)

# --- Configuração da fonte de dados ---
PASTA_ID = "1369rEJAqpprCP3dZp55eXaTcQRU9D5Ol"
DOWNLOAD_BASE = "https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
CACHE_DIR = os.path.join(tempfile.gettempdir(), "atacotada_nf")
os.makedirs(CACHE_DIR, exist_ok=True)

# Fallback caso a listagem falhe
ARQUIVOS_FALLBACK = {
    "1HwHmY16I7OXmhdRqhaBbLY_tuyLe3plx": "202503_NFe_NotaFiscalItem.csv",
    "1vYWRVtDFCklm2o2TQJPbFbVzEGUZBKYt": "202504_NFe_NotaFiscalItem.csv",
    "13JjeGhNsIoUlfZH8ZnNOtB_xa3aCAyVJ": "202505_NFe_NotaFiscalItem.csv",
    "1j1y5PgaxbgRWbPkwymRBYE6kNNDSJeM6": "202506_NFe_NotaFiscalItem.csv",
    "1ibH9e3GRS638eLDoMyWsxcckKW4RYYFR": "202507_NFe_NotaFiscalItem.csv",
    "1wTiEvuD0NgSGXTbPB9LSlrFgqmXnZDUa": "202508_NFe_NotaFiscalItem.csv",
    "1ZJJtcfFpkCtQBfxUB-iOowv0buFq01VX": "202509_NFe_NotaFiscalItem.csv",
    "1sTDH3Zi38dZmsL3NOb1pcbV9WEDxn6yh": "202510_NFe_NotaFiscalItem.csv",
    "1jD1NLznnwvdHhWcr3NSBGeIzrRjMxX3g": "202511_NFe_NotaFiscalItem.csv",
    "1Ye9GhANeEErRuV4GjC3Y-wyn6HKUTSDh": "202512_NFe_NotaFiscalItem.csv",
    "1tSqz-nIiM_uDZW38GdWeRboNC7nwq3jH": "202601_NFe_NotaFiscalItem.csv",
    "1tgekwOo8__NZZSs2OdFMS6xT6pS3LXVn": "202602_NFe_NotaFiscalItem.csv",
}


# --- Funções auxiliares ---
@st.cache_data(ttl=600, show_spinner=False)
def listar_arquivos_disponiveis(folder_id):
    """Descobre os arquivos CSV disponíveis no Portal da Transparência."""
    url = f"https://drive.google.com/drive/folders/{folder_id}"
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        # Extrair data-id dos elementos do HTML
        ids_encontrados = re.findall(r'data-id="([a-zA-Z0-9_-]{20,})"', resp.text)
        ids_encontrados = list(dict.fromkeys(ids_encontrados))  # Remove duplicatas

        if not ids_encontrados:
            return None

        # Obter nomes dos arquivos via título da página do Google Drive
        arquivos = {}
        for fid in ids_encontrados:
            try:
                view_url = f"https://drive.google.com/file/d/{fid}/view"
                view_resp = requests.get(view_url, timeout=15)
                title_match = re.search(r"<title>([^<]+)</title>", view_resp.text)
                if title_match:
                    nome = title_match.group(1).replace(" - Google Drive", "").strip()
                    if nome and nome != "Google Drive":
                        arquivos[fid] = nome
                    else:
                        arquivos[fid] = f"Arquivo {fid[:10]}"
                else:
                    arquivos[fid] = f"Arquivo {fid[:10]}"
            except Exception:
                arquivos[fid] = f"Arquivo {fid[:10]}"

        return arquivos if arquivos else None
    except Exception:
        return None


def baixar_arquivo_csv(file_id, progress_bar=None):
    """Baixa o arquivo CSV para cache local."""
    cache_path = os.path.join(CACHE_DIR, f"{file_id}.csv")

    # Verificar cache existente — descartar se for muito pequeno (provável HTML de erro)
    if os.path.exists(cache_path):
        if os.path.getsize(cache_path) > 10000:
            return cache_path
        else:
            os.remove(cache_path)

    url = DOWNLOAD_BASE.format(file_id=file_id)
    response = requests.get(url, stream=True, timeout=600)
    response.raise_for_status()

    # Detectar resposta HTML (quota excedida, página de confirmação, etc.)
    content_type = response.headers.get("Content-Type", "")
    if "text/html" in content_type:
        response.close()
        raise Exception(
            "O Google Drive retornou uma página de erro (quota de downloads excedida). "
            "Tente novamente em alguns minutos."
        )

    total = int(response.headers.get("content-length", 0))
    downloaded = 0
    tmp_path = cache_path + ".tmp"

    with open(tmp_path, "wb") as f:
        for data in response.iter_content(chunk_size=131072):
            f.write(data)
            downloaded += len(data)
            if progress_bar and total:
                pct = min(downloaded / total, 1.0)
                mb_down = downloaded / (1024 * 1024)
                mb_total = total / (1024 * 1024)
                progress_bar.progress(pct, text=f"Baixando... {mb_down:.0f} / {mb_total:.0f} MB ({pct*100:.0f}%)")

    # Validar que o arquivo baixado é CSV (não HTML de erro)
    with open(tmp_path, "rb") as f:
        head = f.read(200)
    if b"<!DOCTYPE" in head or b"<html" in head.lower():
        os.remove(tmp_path)
        raise Exception(
            "O arquivo baixado não é um CSV válido (quota de downloads do Google Drive excedida). "
            "Tente novamente em alguns minutos."
        )

    os.rename(tmp_path, cache_path)
    return cache_path


@st.cache_data(show_spinner=False, ttl=1800)
def pesquisar_notas(file_path, filtro_produto="", filtro_nome_dest="", filtro_uf_dest="", filtro_uf_emit="", max_resultados=1000):
    """Pesquisa nos dados CSV filtrando por produto, nome destinatário, UF destinatário e UF emitente."""
    resultados = []
    total_linhas = 0

    for chunk in pd.read_csv(
        file_path,
        sep=";",
        encoding="latin-1",
        chunksize=50000,
        dtype=str,
        on_bad_lines="skip",
    ):
        mask = pd.Series(True, index=chunk.index)

        if filtro_produto and filtro_produto.strip():
            col_prod = "DESCRIÇÃO DO PRODUTO/SERVIÇO"
            if col_prod in chunk.columns:
                mask &= chunk[col_prod].astype(str).str.contains(
                    filtro_produto.strip(), case=False, na=False
                )

        if filtro_nome_dest and filtro_nome_dest.strip():
            col_nome = "NOME DESTINATÁRIO"
            if col_nome in chunk.columns:
                mask &= chunk[col_nome].astype(str).str.contains(
                    filtro_nome_dest.strip(), case=False, na=False
                )

        if filtro_uf_dest and filtro_uf_dest.strip():
            col_uf_dest = "UF DESTINATÁRIO"
            if col_uf_dest in chunk.columns:
                mask &= chunk[col_uf_dest].astype(str).str.upper().eq(
                    filtro_uf_dest.strip().upper()
                )

        if filtro_uf_emit and filtro_uf_emit.strip():
            col_uf_emit = "UF EMITENTE"
            if col_uf_emit in chunk.columns:
                mask &= chunk[col_uf_emit].astype(str).str.upper().eq(
                    filtro_uf_emit.strip().upper()
                )

        filtered = chunk[mask]
        if len(filtered) > 0:
            resultados.append(filtered)

        total_linhas += len(chunk)

        total_encontrados = sum(len(r) for r in resultados)
        if total_encontrados >= max_resultados:
            break

    if resultados:
        df = pd.concat(resultados, ignore_index=True).head(max_resultados)
        return df, total_linhas
    return pd.DataFrame(), total_linhas


# --- Funções de consulta de fornecedores via API OpenCNPJ ---
def buscar_dados_fornecedor(cnpj):
    """Busca dados do fornecedor na API OpenCNPJ"""
    if not cnpj or pd.isna(cnpj):
        return None
    try:
        cnpj_limpo = str(cnpj).replace('.', '').replace('/', '').replace('-', '').strip()
        if len(cnpj_limpo) != 14:
            return None
        api_url = f'https://api.opencnpj.org/{cnpj_limpo}'
        response = requests.get(api_url, timeout=5)
        if response.status_code == 200:
            return response.json()
        return None
    except Exception:
        return None


def formatar_telefone(telefone):
    """Formata telefone para um padrão legível"""
    if isinstance(telefone, dict):
        ddd = telefone.get('ddd', '').strip() if telefone.get('ddd') else ''
        numero = telefone.get('numero', '').strip() if telefone.get('numero') else ''
        if ddd and numero:
            numero_fmt = f"{numero[:4]}-{numero[4:]}" if len(numero) >= 8 else numero
            return f"({ddd}) {numero_fmt}"
        return numero or ddd or str(telefone)
    return str(telefone).strip()


def extrair_contatos_fornecedor(dados):
    """Extrai email, telefones, estado e cidade do retorno da API"""
    if not dados:
        return None
    try:
        email = ''
        for c in ['email', 'correio_eletronico', 'mail', 'e_mail', 'emailComercial']:
            email = dados.get(c, '')
            if email:
                break

        estado = ''
        for c in ['state', 'estado', 'uf', 'UF', 'state_code', 'sigla_estado']:
            v = dados.get(c, '')
            if v:
                estado = str(v).upper().strip()
                break

        cidade = ''
        for c in ['city', 'cidade', 'municipio', 'municipality', 'city_name']:
            v = dados.get(c, '')
            if v:
                cidade = str(v).strip()
                break

        telefones = []
        for c in ['phone', 'telefone', 'phonePrimary', 'phoneSecondary',
                   'telefone_comercial', 'telefone_principal', 'ddd_telefone',
                   'ddd_fax', 'fone', 'telephone']:
            v = dados.get(c, '')
            if v:
                t = formatar_telefone(v)
                if t and t not in telefones:
                    telefones.append(t)
        for c in ['phones', 'telefones', 'phone_numbers']:
            vals = dados.get(c, [])
            if isinstance(vals, list):
                for v in vals:
                    if v:
                        t = formatar_telefone(v)
                        if t and t not in telefones:
                            telefones.append(t)

        loc = f"{cidade}, {estado}" if cidade and estado else cidade or estado or ''

        return {
            'email': email or 'Não informado',
            'telefones': ', '.join(telefones) if telefones else 'Não informado',
            'localizacao': loc or 'Não informado',
        }
    except Exception:
        return None


def gerar_html_fornecedores_nf(df_resultado):
    """Gera relatório HTML com dados dos fornecedores (CNPJs) encontrados nas notas fiscais."""
    col_cnpj = "CPF/CNPJ Emitente"
    col_razao = "RAZÃO SOCIAL EMITENTE"
    col_uf = "UF EMITENTE"
    col_mun = "MUNICÍPIO EMITENTE"

    # Obter CNPJs únicos
    if col_cnpj not in df_resultado.columns:
        return "<h3 style='color:red;'>Coluna CNPJ Emitente não encontrada.</h3>"

    fornecedores_unicos = df_resultado.drop_duplicates(subset=[col_cnpj])

    css = """
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); font-family: Arial, sans-serif; padding: 2rem; min-height: 100vh; }
    .container { max-width: 1200px; margin: 0 auto; background: #fff; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.3); overflow: hidden; }
    .header { background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%); color: #fff; padding: 2rem; text-align: center; }
    .header h1 { color: #d4af37; font-size: 36px; margin-bottom: 0.5rem; letter-spacing: 2px; }
    .header p { font-size: 14px; color: #fff; }
    .content { padding: 2rem; }
    .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 1rem; margin-bottom: 2rem; }
    .stat-box { background: linear-gradient(135deg, #0a2540 0%, #164863 100%); padding: 1.5rem; border-radius: 8px; border-left: 5px solid #d4af37; text-align: center; }
    .stat-box label { color: #d4af37; font-weight: bold; font-size: 12px; letter-spacing: 1px; }
    .stat-box .value { color: #fff; font-size: 28px; font-weight: bold; margin-top: 0.5rem; }
    table { width: 100%; border-collapse: collapse; margin-top: 1rem; }
    thead { background: #001a4d; color: #fff; }
    th { padding: 1rem; text-align: left; font-weight: bold; border-bottom: 2px solid #d4af37; color: #fff; }
    td { padding: 0.75rem 1rem; border-bottom: 1px solid #e0e0e0; color: #333; }
    tbody tr:hover { background: #f5f5f5; }
    tbody tr:nth-child(even) { background: #f9f9f9; }
    .email { color: #0033cc; text-decoration: none; }
    .email:hover { text-decoration: underline; }
    .phone { color: #006600; }
    .footer { background: #f0f0f0; padding: 1rem 2rem; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }
    """

    linhas = []
    com_email = 0
    com_telefone = 0

    for _, row in fornecedores_unicos.iterrows():
        cnpj = row.get(col_cnpj, '')
        razao = row.get(col_razao, '') if col_razao in fornecedores_unicos.columns else ''
        uf_csv = row.get(col_uf, '') if col_uf in fornecedores_unicos.columns else ''
        mun_csv = row.get(col_mun, '') if col_mun in fornecedores_unicos.columns else ''

        dados_api = buscar_dados_fornecedor(cnpj)
        contatos = extrair_contatos_fornecedor(dados_api)

        if contatos:
            email = contatos['email']
            telefones = contatos['telefones']
            localizacao = contatos['localizacao']
            if email != 'Não informado':
                com_email += 1
            if telefones != 'Não informado':
                com_telefone += 1
        else:
            email = 'Não informado'
            telefones = 'Não informado'
            loc_parts = [p for p in [mun_csv, uf_csv] if p]
            localizacao = ', '.join(loc_parts) if loc_parts else 'Não informado'

        email_html = f'<a href="mailto:{email}" class="email">{email}</a>' if email != 'Não informado' else 'Não informado'

        linhas.append(
            f"<tr><td>{cnpj}</td><td>{razao}</td><td>{localizacao}</td>"
            f"<td>{email_html}</td><td><span class='phone'>{telefones}</span></td></tr>"
        )

    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Contatos de Fornecedores - Notas Fiscais - AtaCotada</title>
    <style>{css}</style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>AtaCotada</h1>
            <p>Contatos de Fornecedores — Notas Fiscais</p>
        </div>
        <div class="content">
            <div class="stats">
                <div class="stat-box">
                    <label>TOTAL DE FORNECEDORES</label>
                    <div class="value">{len(fornecedores_unicos)}</div>
                </div>
                <div class="stat-box">
                    <label>COM EMAIL</label>
                    <div class="value">{com_email}</div>
                </div>
                <div class="stat-box">
                    <label>COM TELEFONE</label>
                    <div class="value">{com_telefone}</div>
                </div>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>CNPJ</th>
                        <th>Razão Social</th>
                        <th>Localização</th>
                        <th>Email</th>
                        <th>Telefones</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(linhas)}
                </tbody>
            </table>
        </div>
        <div class="footer">
            <p>Relatório gerado por AtaCotada em {data_hora}</p>
        </div>
    </div>
</body>
</html>"""


# --- Funções de geração de relatórios formatados ---

def gerar_nome_pesquisa(filtro_produto="", filtro_nome_dest="", filtro_uf_dest="", filtro_uf_emit=""):
    """Gera nome descritivo da pesquisa baseado nos filtros"""
    partes = []
    if filtro_produto and filtro_produto.strip():
        partes.append(filtro_produto.strip().upper())
    if filtro_nome_dest and filtro_nome_dest.strip():
        partes.append(filtro_nome_dest.strip())
    if filtro_uf_dest and filtro_uf_dest.strip():
        partes.append(f"Dest:{filtro_uf_dest.strip().upper()}")
    if filtro_uf_emit and filtro_uf_emit.strip():
        partes.append(f"Emit:{filtro_uf_emit.strip().upper()}")
    return " | ".join(partes) if partes else "Consulta Geral"


def gerar_filtros_texto(filtro_produto="", filtro_nome_dest="", filtro_uf_dest="", filtro_uf_emit=""):
    """Gera texto descritivo dos filtros aplicados"""
    partes = []
    if filtro_produto and filtro_produto.strip():
        partes.append(f"Produto: {filtro_produto.strip()}")
    if filtro_nome_dest and filtro_nome_dest.strip():
        partes.append(f"Destinatário: {filtro_nome_dest.strip()}")
    if filtro_uf_dest and filtro_uf_dest.strip():
        partes.append(f"UF Dest.: {filtro_uf_dest.strip().upper()}")
    if filtro_uf_emit and filtro_uf_emit.strip():
        partes.append(f"UF Emit.: {filtro_uf_emit.strip().upper()}")
    return " | ".join(partes) if partes else "Sem filtros específicos"


def sanitizar_nome_arquivo(nome):
    """Remove caracteres inválidos para nomes de arquivo"""
    nome = re.sub(r'[<>:"/\\|?*]', '', str(nome))
    nome = re.sub(r'\s+', '_', nome.strip())
    return nome[:60] if nome else "Geral"


def _pdf_safe(text):
    """Garante texto seguro para renderização PDF com fontes Latin-1"""
    if not text or str(text) == 'nan' or str(text) == 'None':
        return ''
    try:
        return str(text).encode('latin-1', 'replace').decode('latin-1')
    except Exception:
        return str(text)


def _fmt_brl(valor):
    """Formata valor numérico para R$ no padrão brasileiro"""
    try:
        return (f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    except Exception:
        return "—"


JUSTIFICATIVA_NF = (
    "Justificativa quanto à utilização de Notas Fiscais como parâmetro de pesquisa de preços\n\n"
    "Considerando a necessidade de obtenção de parâmetros idôneos para a formação do preço "
    "estimado da contratação, e diante da inexistência ou insuficiência de registros plenamente "
    "compatíveis nos demais parâmetros prioritários, foram utilizadas Notas Fiscais recentes como "
    "elemento complementar de pesquisa de mercado.\n\n"
    "As referidas Notas Fiscais comprovam preços efetivamente praticados em contratações reais, "
    "observando-se a similaridade do objeto e a compatibilidade técnica com o item em análise. "
    "Ressalta-se que os documentos considerados encontram-se dentro do intervalo temporal admitido "
    "para fins de aferição de atualidade de preços, atendendo ao critério de contemporaneidade da pesquisa.\n\n"
    "A utilização dessas Notas Fiscais está devidamente fundamentada no parâmetro relativo a "
    "aquisições e contratações similares, constituindo meio legítimo de comprovação de valores "
    "praticados no mercado.\n\n"
    "Adicionalmente, informa-se que a planilha que acompanha o presente relatório contém a chave "
    "de acesso de cada Nota Fiscal utilizada, possibilitando a verificação de autenticidade e validade "
    "diretamente no Portal Nacional da NF-e, garantindo transparência, rastreabilidade e segurança "
    "documental ao procedimento.\n\n"
    "Dessa forma, entende-se que a metodologia adotada atende aos princípios da razoabilidade, "
    "economicidade e motivação do ato administrativo, conferindo robustez à formação do preço estimado."
)


def gerar_pdf_notas(df_exib, item_pesquisado, filtros_texto, stats_info, file_name):
    """Gera PDF formatado e padronizado com os resultados das notas fiscais"""
    from fpdf import FPDF

    # Colunas principais (sem Chave de Acesso) — landscape A4 (largura útil ~277mm)
    colunas_pdf_def = [
        ("Data", 22),
        ("Produto", 55),
        ("Unidade", 13),
        ("Quantidade", 17),
        ("Valor Unitário", 22),
        ("Valor Total", 22),
        ("Razão Social Emitente", 50),
        ("CNPJ Emitente", 28),
        ("UF Emitente", 12),
        ("Nome Destinatário", 36),
    ]

    # Chave de Acesso será exibida em sub-linha separada
    tem_chave = "Chave de Acesso" in df_exib.columns

    colunas_pdf = [(n, w) for n, w in colunas_pdf_def if n in df_exib.columns]
    if not colunas_pdf:
        cols = list(df_exib.columns)[:8]
        w_each = 277 / max(len(cols), 1)
        colunas_pdf = [(c, w_each) for c in cols]

    total_w = sum(w for _, w in colunas_pdf)
    effective_w = 277
    if abs(total_w - effective_w) > 1:
        factor = effective_w / total_w
        colunas_pdf = [(n, round(w * factor, 1)) for n, w in colunas_pdf]

    col_names = [n for n, _ in colunas_pdf]
    col_widths = [w for _, w in colunas_pdf]
    x_start = 10

    # Altura por registro: linha principal + sub-linha chave de acesso
    row_h = 5
    chave_h = 4 if tem_chave else 0
    block_h = row_h + chave_h

    class PDFNotas(FPDF):
        def header(self):
            self.set_fill_color(0, 26, 77)
            self.rect(0, 0, self.w, 30, 'F')
            self.set_fill_color(212, 175, 55)
            self.rect(0, 30, self.w, 1.2, 'F')
            self.set_font('Helvetica', 'B', 20)
            self.set_text_color(212, 175, 55)
            self.set_xy(10, 4)
            self.cell(60, 10, 'AtaCotada')
            self.set_font('Helvetica', 'B', 11)
            self.set_text_color(255, 255, 255)
            self.set_xy(10, 16)
            self.cell(self.w - 80, 8, _pdf_safe(f'Notas Fiscais - {item_pesquisado}')[:90])
            self.set_font('Helvetica', '', 7)
            self.set_text_color(180, 180, 180)
            self.set_xy(self.w - 70, 6)
            self.cell(60, 5, 'Marinha do Brasil', align='R')
            self.set_y(35)

        def footer(self):
            self.set_y(-10)
            self.set_fill_color(0, 26, 77)
            self.rect(0, self.h - 10, self.w, 10, 'F')
            self.set_font('Helvetica', 'I', 6.5)
            self.set_text_color(160, 160, 160)
            dh = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.cell(0, 8, _pdf_safe(f'AtaCotada - Gerado em {dh}'), align='L')
            self.set_x(-30)
            self.cell(20, 8, f'{self.page_no()}/{{nb}}', align='R')

    pdf = PDFNotas(orientation='L', unit='mm', format='A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Caixa de filtros
    y0 = pdf.get_y()
    pdf.set_fill_color(245, 248, 255)
    pdf.set_draw_color(0, 51, 204)
    pdf.rect(x_start, y0, effective_w, 12, 'DF')
    pdf.set_xy(x_start + 3, y0 + 1.5)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(0, 26, 77)
    pdf.cell(18, 4, 'Filtros:')
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(51, 51, 51)
    pdf.cell(0, 4, _pdf_safe(filtros_texto))
    pdf.set_xy(x_start + 3, y0 + 6.5)
    pdf.set_font('Helvetica', 'I', 6.5)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 4, _pdf_safe(f'Arquivo: {file_name}'))
    pdf.set_y(y0 + 15)

    # Cards de estatísticas — dividir em 2 linhas se houver mais de 4
    stats_items = list(stats_info.items())
    row1 = stats_items[:4]
    row2 = stats_items[4:]

    def _draw_stats_row(items, y_pos):
        n = len(items)
        sw = effective_w / max(n, 1)
        for i, (label, value) in enumerate(items):
            sx = x_start + i * sw
            pdf.set_fill_color(10, 37, 64)
            pdf.rect(sx + 0.5, y_pos, sw - 1, 13, 'F')
            pdf.set_fill_color(212, 175, 55)
            pdf.rect(sx + 0.5, y_pos, 1.2, 13, 'F')
            pdf.set_xy(sx + 4, y_pos + 1)
            pdf.set_font('Helvetica', '', 5.5)
            pdf.set_text_color(212, 175, 55)
            pdf.cell(sw - 6, 4, _pdf_safe(str(label)))
            pdf.set_xy(sx + 4, y_pos + 6)
            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(sw - 6, 5, _pdf_safe(str(value)))

    y_s = pdf.get_y()
    _draw_stats_row(row1, y_s)
    if row2:
        y_s2 = y_s + 15
        _draw_stats_row(row2, y_s2)
        pdf.set_y(y_s2 + 17)
    else:
        pdf.set_y(y_s + 17)

    # Colunas numéricas alinhadas à direita
    _right_cols = {'Quantidade', 'Valor Unitário', 'Valor Total'}

    def draw_table_header():
        pdf.set_fill_color(0, 26, 77)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 6)
        pdf.set_draw_color(30, 30, 80)
        pdf.set_x(x_start)
        for name, w in zip(col_names, col_widths):
            h_align = 'R' if name in _right_cols else 'L'
            pdf.cell(w, 6.5, _pdf_safe(name), border=1, fill=True, align=h_align)
        pdf.ln()
        # Sub-cabeçalho para Chave de Acesso
        if tem_chave:
            pdf.set_x(x_start)
            pdf.set_font('Helvetica', 'B', 5)
            pdf.set_fill_color(0, 26, 77)
            pdf.cell(sum(col_widths), 4, 'Chave de Acesso', border=1, fill=True, align='L')
            pdf.ln()
        pdf.set_fill_color(212, 175, 55)
        pdf.rect(x_start, pdf.get_y(), sum(col_widths), 0.4, 'F')
        pdf.ln(0.6)

    draw_table_header()

    pdf.set_font('Helvetica', '', 5.5)
    max_rows_pdf = min(len(df_exib), 800)

    for idx in range(max_rows_pdf):
        row = df_exib.iloc[idx]
        if pdf.get_y() + block_h + 1 > pdf.h - 15:
            pdf.add_page()
            draw_table_header()
            pdf.set_font('Helvetica', '', 5.5)

        if idx % 2 == 0:
            fill_color = (255, 255, 255)
        else:
            fill_color = (243, 246, 252)

        pdf.set_fill_color(*fill_color)
        pdf.set_text_color(51, 51, 51)
        pdf.set_draw_color(230, 230, 230)
        pdf.set_x(x_start)

        # Linha principal
        pdf.set_font('Helvetica', '', 5.5)
        for name, w in zip(col_names, col_widths):
            val = str(row.get(name, ''))
            if val in ('nan', 'None'):
                val = ''
            max_chars = int(w / 1.6)
            if len(val) > max_chars:
                val = val[:max_chars - 2] + '..'
            align = 'R' if name in ('Quantidade', 'Valor Unitário', 'Valor Total') else 'L'
            pdf.cell(w, row_h, _pdf_safe(val), border='', fill=True, align=align)
        pdf.ln()

        # Sub-linha: Chave de Acesso completa
        if tem_chave:
            chave = str(row.get('Chave de Acesso', ''))
            if chave in ('nan', 'None'):
                chave = ''
            pdf.set_fill_color(*fill_color)
            pdf.set_x(x_start)
            pdf.set_font('Helvetica', '', 4.5)
            pdf.set_text_color(100, 100, 100)
            pdf.cell(sum(col_widths), chave_h, _pdf_safe(f'  Chave: {chave}'), border='B', fill=True, align='L')
            pdf.ln()
        else:
            # Borda inferior quando não tem chave
            pdf.set_x(x_start)
            pdf.cell(sum(col_widths), 0, '', border='B')
            pdf.ln()

    pdf.set_draw_color(0, 26, 77)
    pdf.set_line_width(0.3)
    pdf.line(x_start, pdf.get_y(), x_start + sum(col_widths), pdf.get_y())

    if len(df_exib) > max_rows_pdf:
        pdf.ln(3)
        pdf.set_font('Helvetica', 'I', 7)
        pdf.set_text_color(130, 130, 130)
        pdf.cell(0, 4, _pdf_safe(f'... e mais {len(df_exib) - max_rows_pdf} registros (limitado a {max_rows_pdf} no PDF)'), align='C')

    # --- Justificativa ---
    pdf.add_page()
    pdf.ln(5)
    # Título centralizado e em negrito
    just_paragraphs = JUSTIFICATIVA_NF.split('\n\n')
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_text_color(0, 26, 77)
    pdf.multi_cell(effective_w, 6, _pdf_safe(just_paragraphs[0]), align='C')
    pdf.ln(4)
    # Parágrafos do corpo
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(51, 51, 51)
    for par in just_paragraphs[1:]:
        pdf.set_x(x_start)
        pdf.multi_cell(effective_w, 5, _pdf_safe(par), align='J')
        pdf.ln(3)

    return bytes(pdf.output())


def gerar_excel_formatado(df_exib, item_pesquisado, filtros_texto, stats_info, file_name):
    """Gera Excel formatado com cabeçalho bonito, cores padronizadas e nome do item"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_exib.to_excel(writer, index=False, sheet_name='Notas Fiscais', startrow=7)
        ws = writer.sheets['Notas Fiscais']

        max_col = len(df_exib.columns)
        last_col = get_column_letter(max_col)

        # Estilos
        navy = PatternFill(start_color='001A4D', end_color='001A4D', fill_type='solid')
        dark_navy = PatternFill(start_color='0A2540', end_color='0A2540', fill_type='solid')
        light_bg = PatternFill(start_color='F5F8FF', end_color='F5F8FF', fill_type='solid')
        alt_row = PatternFill(start_color='EDF2FA', end_color='EDF2FA', fill_type='solid')
        white_bg = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        gold_title = Font(name='Calibri', bold=True, color='D4AF37', size=20)
        white_subtitle = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
        gray_info = Font(name='Calibri', color='B0B0B0', size=9, italic=True)
        filter_font = Font(name='Calibri', color='333333', size=9)
        date_font = Font(name='Calibri', color='888888', size=8)
        gold_stat = Font(name='Calibri', bold=True, color='D4AF37', size=10)
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        data_font = Font(name='Calibri', size=10, color='333333')

        gold_bottom = Border(bottom=Side(style='medium', color='D4AF37'))
        thin_bottom = Border(bottom=Side(style='thin', color='E8E8E8'))

        center = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        wrap_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Linha 1: Título AtaCotada
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value = 'AtaCotada'
        c.font = gold_title
        c.fill = navy
        c.alignment = center
        ws.row_dimensions[1].height = 40
        for col in range(2, max_col + 1):
            ws.cell(row=1, column=col).fill = navy

        # Linha 2: Subtítulo com item pesquisado
        ws.merge_cells(f'A2:{last_col}2')
        c = ws['A2']
        c.value = f'Notas Fiscais — {item_pesquisado}'
        c.font = white_subtitle
        c.fill = navy
        c.alignment = center
        ws.row_dimensions[2].height = 26
        for col in range(2, max_col + 1):
            ws.cell(row=2, column=col).fill = navy

        # Linha 3: Marinha do Brasil
        ws.merge_cells(f'A3:{last_col}3')
        c = ws['A3']
        c.value = 'Marinha do Brasil — Centro de Operações do Abastecimento'
        c.font = gray_info
        c.fill = navy
        c.alignment = center
        ws.row_dimensions[3].height = 18
        for col in range(2, max_col + 1):
            ws.cell(row=3, column=col).fill = navy

        # Linha 4: Filtros
        ws.merge_cells(f'A4:{last_col}4')
        c = ws['A4']
        c.value = f'Filtros aplicados:  {filtros_texto}'
        c.font = filter_font
        c.fill = light_bg
        c.alignment = left_align
        ws.row_dimensions[4].height = 20
        for col in range(2, max_col + 1):
            ws.cell(row=4, column=col).fill = light_bg

        # Linha 5: Arquivo e data
        ws.merge_cells(f'A5:{last_col}5')
        c = ws['A5']
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        c.value = f'Arquivo: {file_name}  |  Gerado em: {data_hora}'
        c.font = date_font
        c.fill = light_bg
        c.alignment = left_align
        ws.row_dimensions[5].height = 18
        for col in range(2, max_col + 1):
            ws.cell(row=5, column=col).fill = light_bg

        # Linha 6: Estatísticas
        stats_list = list(stats_info.items())
        for i in range(max_col):
            cell = ws.cell(row=6, column=i + 1)
            cell.fill = dark_navy
            if i < len(stats_list):
                label, value = stats_list[i]
                cell.value = f'{label}: {value}'
                cell.font = gold_stat
                cell.alignment = center
        ws.row_dimensions[6].height = 24

        # Linha 7: Separador
        for i in range(1, max_col + 1):
            cell = ws.cell(row=7, column=i)
            cell.fill = white_bg
            cell.border = Border(bottom=Side(style='thin', color='D4AF37'))
        ws.row_dimensions[7].height = 4

        # Linha 8: Cabeçalho da tabela
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=8, column=col_idx)
            cell.font = header_font
            cell.fill = navy
            cell.alignment = wrap_center
            cell.border = gold_bottom
        ws.row_dimensions[8].height = 28

        # Linhas de dados
        for row_idx in range(9, 9 + len(df_exib)):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = left_align
                cell.border = thin_bottom
                if (row_idx - 9) % 2 == 1:
                    cell.fill = alt_row
                else:
                    cell.fill = white_bg

        # Auto-ajustar largura das colunas
        for col_idx in range(1, max_col + 1):
            col_l = get_column_letter(col_idx)
            header_len = len(str(ws.cell(row=8, column=col_idx).value or ''))
            max_len = header_len
            for row_idx in range(9, min(9 + len(df_exib), 109)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            adjusted = min(max_len * 1.15 + 2, 50)
            ws.column_dimensions[col_l].width = max(adjusted, 10)

        # Congelar painéis abaixo do cabeçalho
        ws.freeze_panes = 'A9'

        # --- Justificativa no final ---
        just_start = 9 + len(df_exib) + 2  # 2 linhas em branco
        just_font_bold = Font(name='Calibri', bold=True, color='001A4D', size=11)
        just_font = Font(name='Calibri', color='333333', size=10)
        just_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        just_align_body = Alignment(horizontal='justify', vertical='top', wrap_text=True)

        just_paragraphs = JUSTIFICATIVA_NF.split('\n\n')
        current_row = just_start

        # Título
        ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = just_paragraphs[0]
        cell.font = just_font_bold
        cell.alignment = just_align
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        # Parágrafos
        for par in just_paragraphs[1:]:
            ws.merge_cells(f'A{current_row}:{last_col}{current_row}')
            cell = ws.cell(row=current_row, column=1)
            cell.value = par
            cell.font = just_font
            cell.alignment = just_align_body
            ws.row_dimensions[current_row].height = max(60, len(par) // 2)
            current_row += 1

    return output.getvalue()


# ==================== ABA 1: PESQUISA NF (PORTAL DA TRANSPARÊNCIA) ====================
with tab_pesquisa:
    # --- Carregar lista de arquivos ---
    with st.spinner("Carregando lista de arquivos do Portal da Transparência..."):
        arquivos_portal = listar_arquivos_disponiveis(PASTA_ID)

    if not arquivos_portal:
        arquivos_portal = ARQUIVOS_FALLBACK.copy()
        st.warning("⚠️ Não foi possível listar os arquivos. Usando dados conhecidos como fallback.")

    col_info, col_refresh = st.columns([4, 1])
    with col_info:
        st.markdown(f"**{len(arquivos_portal)}** arquivo(s) disponível(is) no Portal da Transparência.")
    with col_refresh:
        if st.button("🔄 Atualizar lista", use_container_width=True):
            listar_arquivos_disponiveis.clear()
            st.rerun()

    # --- Formulário de busca ---
    st.markdown('<div class="filtros-container">', unsafe_allow_html=True)
    st.subheader("🔍 Filtros de Pesquisa")

    nomes_arquivos = list(arquivos_portal.values())
    ids_arquivos = list(arquivos_portal.keys())

    # Ordenar por nome decrescente para o arquivo mais recente ficar primeiro
    ordem = sorted(range(len(nomes_arquivos)), key=lambda i: nomes_arquivos[i], reverse=True)
    nomes_arquivos = [nomes_arquivos[i] for i in ordem]
    ids_arquivos = [ids_arquivos[i] for i in ordem]

    arquivo_idx = st.selectbox(
        "📂 Arquivo / Período",
        options=range(len(nomes_arquivos)),
        format_func=lambda i: nomes_arquivos[i],
        help="Selecione o arquivo CSV a ser consultado",
    )

    col1, col2 = st.columns(2)

    with col1:
        filtro_produto = st.text_input(
            "Descrição do Produto / Serviço",
            placeholder="Ex: TINTA, PARAFUSO, DIESEL...",
            help="Pesquisa parcial na descrição do produto/serviço",
        )

    with col2:
        filtro_nome_dest = st.text_input(
            "Nome Destinatário",
            placeholder="Ex: ARSENAL, HOSPITAL NAVAL...",
            help="Pesquisa parcial no nome do destinatário",
        )

    col3, col4, col5 = st.columns(3)

    with col3:
        filtro_uf_dest = st.text_input(
            "UF Destinatário",
            placeholder="Ex: RJ, SP, DF...",
            help="Sigla do estado do destinatário (filtro exato)",
            max_chars=2,
        )

    with col4:
        filtro_uf_emit = st.text_input(
            "UF Emitente",
            placeholder="Ex: RJ, SP, MG...",
            help="Sigla do estado do emitente/fornecedor (filtro exato)",
            max_chars=2,
        )

    with col5:
        max_resultados = st.number_input(
            "Máximo de resultados",
            min_value=100,
            max_value=10000,
            value=500,
            step=100,
            help="Limita a quantidade de registros retornados para otimizar a performance",
        )

    buscar = st.button("🔎 Pesquisar Notas Fiscais", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # --- Processar busca ---
    if buscar:
        if not filtro_produto and not filtro_nome_dest and not filtro_uf_dest and not filtro_uf_emit:
            st.warning("⚠️ Informe pelo menos um filtro para realizar a pesquisa.")
        else:
            file_id = ids_arquivos[arquivo_idx]
            file_name = nomes_arquivos[arquivo_idx]

            # Etapa 1: Download do arquivo
            progress_bar = st.progress(0, text="Verificando arquivo...")
            cache_path = os.path.join(CACHE_DIR, f"{file_id}.csv")

            # Invalidar cache corrompido (HTML salvo como CSV)
            if os.path.exists(cache_path) and os.path.getsize(cache_path) < 10000:
                os.remove(cache_path)

            if os.path.exists(cache_path):
                progress_bar.progress(1.0, text="✅ Arquivo já em cache local")
            else:
                try:
                    cache_path = baixar_arquivo_csv(file_id, progress_bar)
                except Exception as e:
                    progress_bar.empty()
                    st.error(f"❌ {str(e)}")
                    st.stop()

            # Etapa 2: Pesquisa nos dados
            progress_bar.progress(1.0, text="🔍 Pesquisando nos dados...")

            try:
                df_resultado, total_linhas = pesquisar_notas(
                    cache_path, filtro_produto, filtro_nome_dest, filtro_uf_dest, filtro_uf_emit, max_resultados
                )
            except Exception as e:
                progress_bar.empty()
                st.error(f"❌ Erro ao processar o arquivo: {str(e)}")
                st.stop()

            progress_bar.empty()

            if df_resultado.empty:
                st.info("📭 Nenhuma nota fiscal encontrada para os filtros informados.")
            else:
                # Salvar resultados no session_state para persistir entre interações
                st.session_state["nf_resultado"] = df_resultado
                st.session_state["nf_total_linhas"] = total_linhas
                st.session_state["nf_file_name"] = file_name
                st.session_state["nf_max_resultados"] = max_resultados
                st.session_state["nf_item_pesquisado"] = gerar_nome_pesquisa(filtro_produto, filtro_nome_dest, filtro_uf_dest, filtro_uf_emit)
                st.session_state["nf_filtros_texto"] = gerar_filtros_texto(filtro_produto, filtro_nome_dest, filtro_uf_dest, filtro_uf_emit)

    # --- Exibir resultados (persistidos no session_state) ---
    if "nf_resultado" in st.session_state and not st.session_state["nf_resultado"].empty:
        df_resultado = st.session_state["nf_resultado"]
        total_linhas = st.session_state["nf_total_linhas"]
        file_name = st.session_state["nf_file_name"]
        max_resultados = st.session_state["nf_max_resultados"]

        # Colunas e ordem definidas pelo usuário
        colunas_exibicao = {
            "DATA EMISSÃO": "Data",
            "DESCRIÇÃO DO PRODUTO/SERVIÇO": "Produto",
            "UNIDADE": "Unidade",
            "QUANTIDADE": "Quantidade",
            "VALOR UNITÁRIO": "Valor Unitário",
            "VALOR TOTAL": "Valor Total",
            "UF DESTINATÁRIO": "UF Destinatário",
            "NOME DESTINATÁRIO": "Nome Destinatário",
            "RAZÃO SOCIAL EMITENTE": "Razão Social Emitente",
            "CPF/CNPJ Emitente": "CNPJ Emitente",
            "UF EMITENTE": "UF Emitente",
            "MUNICÍPIO EMITENTE": "Município",
            "NCM/SH (TIPO DE PRODUTO)": "NCM/SH (Tipo de Produto)",
            "CHAVE DE ACESSO": "Chave de Acesso",
            "NATUREZA DA OPERAÇÃO": "Natureza da Operação",
        }
        colunas_disponiveis = [c for c in colunas_exibicao if c in df_resultado.columns]
        df_exib = df_resultado[colunas_disponiveis].rename(columns=colunas_exibicao).copy()

        # --- Filtro de Outliers ---
        _col_vunit = "Valor Unitário"
        _valores_num = pd.Series(dtype=float)
        if _col_vunit in df_exib.columns:
            _valores_num = pd.to_numeric(
                df_exib[_col_vunit].astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False),
                errors="coerce"
            )

        col_outlier, _ = st.columns([1, 2])
        with col_outlier:
            remover_outliers = st.checkbox(
                "🎯 Remover Outliers (IQR)",
                value=False,
                help="Remove valores unitários discrepantes usando o método IQR (Intervalo Interquartil). "
                     "Valores abaixo de Q1 − 1.5×IQR ou acima de Q3 + 1.5×IQR são removidos.",
            )

        if remover_outliers and not _valores_num.dropna().empty:
            q1 = _valores_num.quantile(0.25)
            q3 = _valores_num.quantile(0.75)
            iqr = q3 - q1
            lim_inf = max(q1 - 1.5 * iqr, 0)
            lim_sup = q3 + 1.5 * iqr
            mask_inlier = ((_valores_num >= lim_inf) & (_valores_num <= lim_sup)) | _valores_num.isna()
            n_removidos = int((~mask_inlier).sum())
            df_exib = df_exib[mask_inlier].reset_index(drop=True)
            _valores_num = _valores_num[mask_inlier].reset_index(drop=True)
            if n_removidos > 0:
                st.info(f"🎯 {n_removidos} registro(s) com valor unitário discrepante removido(s) "
                        f"(limites: {_fmt_brl(lim_inf)} — {_fmt_brl(lim_sup)}).")

        # Calcular estatísticas de preço unitário
        _vnum_valid = _valores_num.dropna()
        if not _vnum_valid.empty:
            preco_menor = _vnum_valid.min()
            preco_medio = _vnum_valid.mean()
            preco_mediana = _vnum_valid.median()
            preco_maior = _vnum_valid.max()
        else:
            preco_menor = preco_medio = preco_mediana = preco_maior = 0

        # --- Cards de estatísticas ---
        st.markdown("---")
        col_s1, col_s2, col_s3, col_s4 = st.columns(4)

        with col_s1:
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">REGISTROS ENCONTRADOS</div>
                <div style="color: #ffffff; font-size: 32px; font-weight: bold;">{len(df_resultado):,}</div>
                <div style="color: #aaaaaa; font-size: 12px;">{total_linhas:,} linhas analisadas</div>
            </div>
            """, unsafe_allow_html=True)

        with col_s2:
            fornecedores = (
                df_resultado["RAZÃO SOCIAL EMITENTE"].nunique()
                if "RAZÃO SOCIAL EMITENTE" in df_resultado.columns
                else 0
            )
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">FORNECEDORES</div>
                <div style="color: #ffffff; font-size: 32px; font-weight: bold;">{fornecedores}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Distintos</div>
            </div>
            """, unsafe_allow_html=True)

        with col_s3:
            orgaos = (
                df_resultado["ÓRGÃO DESTINATÁRIO"].nunique()
                if "ÓRGÃO DESTINATÁRIO" in df_resultado.columns
                else 0
            )
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">ÓRGÃOS DESTINO</div>
                <div style="color: #ffffff; font-size: 32px; font-weight: bold;">{orgaos}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Distintos</div>
            </div>
            """, unsafe_allow_html=True)

        with col_s4:
            try:
                total_valor = pd.to_numeric(
                    df_resultado["VALOR TOTAL"]
                    .str.replace(".", "", regex=False)
                    .str.replace(",", ".", regex=False),
                    errors="coerce",
                ).sum()
                valor_fmt = (
                    f"R$ {total_valor:,.2f}"
                    .replace(",", "X")
                    .replace(".", ",")
                    .replace("X", ".")
                )
            except Exception:
                valor_fmt = "—"
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">VALOR TOTAL</div>
                <div style="color: #ffffff; font-size: 20px; font-weight: bold;">{valor_fmt}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Soma dos registros</div>
            </div>
            """, unsafe_allow_html=True)

        # --- Cards de preço unitário ---
        col_p1, col_p2, col_p3, col_p4 = st.columns(4)

        with col_p1:
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">MENOR PREÇO UNIT.</div>
                <div style="color: #00ff88; font-size: 22px; font-weight: bold;">{_fmt_brl(preco_menor)}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Valor unitário mínimo</div>
            </div>
            """, unsafe_allow_html=True)

        with col_p2:
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">PREÇO MÉDIO</div>
                <div style="color: #ffffff; font-size: 22px; font-weight: bold;">{_fmt_brl(preco_medio)}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Média aritmética</div>
            </div>
            """, unsafe_allow_html=True)

        with col_p3:
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">MEDIANA</div>
                <div style="color: #ffffff; font-size: 22px; font-weight: bold;">{_fmt_brl(preco_mediana)}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Valor central</div>
            </div>
            """, unsafe_allow_html=True)

        with col_p4:
            st.markdown(f"""
            <div class="stats-card">
                <div style="color: #d4af37; font-size: 14px; font-weight: 600;">MAIOR PREÇO UNIT.</div>
                <div style="color: #ff6666; font-size: 22px; font-weight: bold;">{_fmt_brl(preco_maior)}</div>
                <div style="color: #aaaaaa; font-size: 12px;">Valor unitário máximo</div>
            </div>
            """, unsafe_allow_html=True)

        # --- Tabela de resultados ---
        st.markdown("---")
        st.subheader("📋 Resultados")
        st.markdown(f"**Arquivo:** {file_name}")
        st.dataframe(
            df_exib,
            use_container_width=True,
            hide_index=True,
            height=min(len(df_exib) * 38 + 40, 700),
        )

        # --- Preparar dados para exportação ---
        item_pesquisado = st.session_state.get("nf_item_pesquisado", "Consulta Geral")
        filtros_texto = st.session_state.get("nf_filtros_texto", "")
        nome_arquivo_base = sanitizar_nome_arquivo(item_pesquisado)

        # Montar estatísticas para os relatórios
        _fornecedores_exp = (
            df_exib["Razão Social Emitente"].nunique()
            if "Razão Social Emitente" in df_exib.columns else 0
        )
        try:
            _total_valor_exp = pd.to_numeric(
                df_exib["Valor Total"].astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False),
                errors="coerce",
            ).sum()
            _valor_fmt_exp = _fmt_brl(_total_valor_exp)
        except Exception:
            _valor_fmt_exp = "—"

        stats_export = {
            "REGISTROS": f"{len(df_exib):,}",
            "FORNECEDORES": str(_fornecedores_exp),
            "VALOR TOTAL": _valor_fmt_exp,
            "MENOR PREÇO": _fmt_brl(preco_menor),
            "PREÇO MÉDIO": _fmt_brl(preco_medio),
            "MEDIANA": _fmt_brl(preco_mediana),
            "MAIOR PREÇO": _fmt_brl(preco_maior),
        }

        # --- Botões de download ---
        st.markdown("---")
        col_d1, col_d2, col_d3 = st.columns(3)

        with col_d1:
            _csv_stats = (
                f"# Estatísticas de Preço Unitário: "
                f"Menor={_fmt_brl(preco_menor)} | Médio={_fmt_brl(preco_medio)} "
                f"| Mediana={_fmt_brl(preco_mediana)} | Maior={_fmt_brl(preco_maior)}\n"
            )
            _csv_just = "\n\n" + JUSTIFICATIVA_NF.replace('\n\n', '\n') + "\n"
            csv_data = _csv_stats + df_exib.to_csv(index=False, sep=";", encoding="utf-8-sig") + _csv_just
            st.download_button(
                "📥 Baixar CSV",
                data=csv_data,
                file_name=f"NF_{nome_arquivo_base}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with col_d2:
            try:
                excel_data = gerar_excel_formatado(
                    df_exib, item_pesquisado, filtros_texto, stats_export, file_name
                )
                st.download_button(
                    "📥 Baixar Planilha Excel",
                    data=excel_data,
                    file_name=f"NF_{nome_arquivo_base}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Erro ao gerar Excel: {e}")

        with col_d3:
            try:
                pdf_data = gerar_pdf_notas(
                    df_exib, item_pesquisado, filtros_texto, stats_export, file_name
                )
                st.download_button(
                    "📥 Baixar PDF",
                    data=pdf_data,
                    file_name=f"NF_{nome_arquivo_base}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Erro ao gerar PDF: {e}")

        # --- Botão Consultar Fornecedores ---
        st.markdown("<br>", unsafe_allow_html=True)
        col_forn_btn, _ = st.columns([1, 2])
        with col_forn_btn:
            if st.button("📞 Consultar Fornecedores", use_container_width=True, key="btn_fornecedores_nf"):
                cnpjs_unicos = df_resultado["CPF/CNPJ Emitente"].dropna().unique() if "CPF/CNPJ Emitente" in df_resultado.columns else []
                if len(cnpjs_unicos) == 0:
                    st.warning("⚠️ Nenhum CNPJ encontrado nos resultados filtrados.")
                else:
                    with st.spinner(f"Consultando dados de {len(cnpjs_unicos)} fornecedor(es) na API..."):
                        html_fornecedores = gerar_html_fornecedores_nf(df_resultado)
                        st.session_state["nf_html_fornecedores"] = html_fornecedores

        # Mostrar botão de download do relatório se já foi gerado
        if "nf_html_fornecedores" in st.session_state:
            col_dl_forn, _ = st.columns([1, 2])
            with col_dl_forn:
                data_hora_arq = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="📋 Baixar Relatório de Fornecedores (HTML)",
                    data=st.session_state["nf_html_fornecedores"].encode('utf-8'),
                    file_name=f"fornecedores_NF_{data_hora_arq}.html",
                    mime="text/html",
                    use_container_width=True,
                    key="btn_download_fornecedores_nf",
                )

        # --- Informações ---
        st.markdown("---")
        st.markdown(f"""
        <div style="background: #0a2540; border: 1px solid #333; border-radius: 8px; padding: 1rem; margin-top: 0.5rem;">
            <p style="color: #d4af37; font-weight: bold; margin-bottom: 0.5rem;">ℹ️ Informações:</p>
            <ul style="color: #cccccc; font-size: 13px; line-height: 1.8;">
                <li>Dados extraídos do <b>Portal da Transparência</b> — Notas Fiscais Eletrônicas.</li>
                <li>Na primeira consulta, os dados são carregados e armazenados em <b>cache local</b> (pode levar alguns minutos para arquivos grandes).</li>
                <li>Consultas subsequentes no mesmo arquivo são <b>muito mais rápidas</b>.</li>
                <li>Limite de <b>{max_resultados}</b> resultados para otimizar a performance.</li>
                <li>Os resultados podem ser exportados em <b>CSV</b> ou <b>Excel</b>.</li>
            </ul>
            <p style="color: #d4af37; text-align: center; margin-top: 1rem; font-size: 14px; font-weight: 600;">
                Notas Fiscais podem ser usadas na pesquisa de preço, acesse para baixar com a chave de acesso:
                <a href="https://www.nfe.fazenda.gov.br/portal/principal.aspx" target="_blank" style="color: #ffd700; text-decoration: underline;">https://www.nfe.fazenda.gov.br/portal/principal.aspx</a>
            </p>
        </div>
        """, unsafe_allow_html=True)
