import base64
import io
import os
import time
import requests
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ── Caminhos dos dados ─────────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Projeto Adesões")

# ── Constantes da API ──────────────────────────────────────────────────────
API_URL_FORNECEDOR = "https://dadosabertos.compras.gov.br/modulo-fornecedor/1_consultarFornecedor"

# Configuração da página
st.set_page_config(
    page_title="AtaCotada - Banco de Fornecedores",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# CSS customizado (mesmo padrão visual do projeto)
st.markdown("""
    <style>
        /* Anti-flash: forçar fundo escuro desde o início */
        html, body, [data-testid="stAppViewContainer"],
        .main, [data-testid="stApp"], .stApp {
            background-color: #001a4d !important;
            color: #ffffff !important;
        }

        /* Transição suave ao carregar */
        .stApp {
            animation: fadeIn 0.3s ease-in;
        }

        @keyframes fadeIn {
            from { opacity: 0.7; }
            to { opacity: 1; }
        }

        /* Header customizado */
        .header-container {
            background: linear-gradient(135deg, #001a4d 0%, #0033cc 100%);
            padding: 2rem;
            border-radius: 10px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3);
            text-align: center;
        }

        .logo-text {
            color: #ffffff;
            font-size: 14px;
            font-weight: 600;
            letter-spacing: 2px;
            margin-bottom: 0.5rem;
        }

        .sistema-nome {
            color: #d4af37;
            font-size: 48px;
            font-weight: bold;
            letter-spacing: 3px;
            text-shadow: 2px 2px 4px rgba(0, 0, 0, 0.5);
            margin: 0.5rem 0;
            font-family: 'Arial Black', sans-serif;
        }

        .subtitulo {
            color: #ffffff;
            font-size: 14px;
            margin-top: 0.5rem;
            letter-spacing: 1px;
        }

        /* ===== SIDEBAR MODERNA - PRETA COM BORDA DOURADA ===== */
        [data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0a0a0a 0%, #111111 50%, #0a0a0a 100%) !important;
            border-right: 3px solid #d4af37 !important;
            box-shadow: 4px 0 15px rgba(0, 0, 0, 0.5);
        }

        [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
            background: transparent !important;
        }

        /* Esconder navegação padrão do Streamlit */
        [data-testid="stSidebarNav"] {
            display: none !important;
        }

        /* Título da sidebar */
        [data-testid="stSidebar"] .stMarkdown h2 {
            color: #d4af37 !important;
            font-family: 'Arial Black', sans-serif;
            font-size: 22px;
            text-align: center;
            letter-spacing: 2px;
            text-shadow: 1px 1px 3px rgba(0, 0, 0, 0.8);
            border-bottom: 2px solid #d4af37;
            padding-bottom: 0.75rem;
            margin-bottom: 1.5rem;
        }

        /* Links de navegação na sidebar */
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] {
            background: linear-gradient(135deg, #1a1a1a 0%, #252525 100%) !important;
            color: #ffffff !important;
            border: 1px solid #333333 !important;
            border-radius: 10px !important;
            margin: 0.4rem 0 !important;
            padding: 0.85rem 1.2rem !important;
            font-weight: 600 !important;
            font-size: 15px !important;
            transition: all 0.3s ease !important;
            text-decoration: none !important;
            display: flex !important;
            align-items: center !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"] span {
            color: #ffffff !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover {
            background: linear-gradient(135deg, #252525 0%, #353535 100%) !important;
            border: 1px solid #d4af37 !important;
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.25) !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"]:hover span {
            color: #d4af37 !important;
        }

        /* Link ativo / página atual */
        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] {
            background: linear-gradient(135deg, #d4af37 0%, #c5a028 100%) !important;
            color: #0a0a0a !important;
            border: 1px solid #d4af37 !important;
            font-weight: bold !important;
            box-shadow: 0 4px 15px rgba(212, 175, 55, 0.4) !important;
        }

        [data-testid="stSidebar"] a[data-testid="stPageLink-NavLink"][aria-current="page"] span {
            color: #0a0a0a !important;
        }

        /* Texto e labels na sidebar */
        [data-testid="stSidebar"] label,
        [data-testid="stSidebar"] .stText,
        [data-testid="stSidebar"] p {
            color: #ffffff !important;
        }

        /* Separadores na sidebar */
        [data-testid="stSidebar"] hr {
            border-color: #333333 !important;
            margin: 1rem 0 !important;
        }

        /* Rodapé da sidebar */
        .sidebar-footer {
            color: #666666;
            font-size: 11px;
            text-align: center;
            padding: 1rem 0;
            border-top: 1px solid #333333;
            margin-top: 2rem;
        }

        /* ===== ESTILOS ESPECÍFICOS DA PÁGINA FORNECEDORES ===== */
        .metric-card {
            background: rgba(0, 26, 77, 0.6);
            border: 1px solid #d4af37;
            border-radius: 12px;
            padding: 1rem 1.25rem;
            box-shadow: 0 10px 30px rgba(0,0,0,0.35);
        }

        .result-card {
            background: rgba(0, 26, 77, 0.45);
            border-radius: 10px;
            padding: 0.9rem 1.1rem;
            margin-bottom: 0.55rem;
            border: 1px solid rgba(212, 175, 55, 0.3);
            box-shadow: 0 6px 24px rgba(0,0,0,0.28);
        }

        .result-card a {
            color: #d4af37 !important;
            text-decoration: none !important;
            font-weight: 600;
        }

        .result-card a:hover {
            color: #f0d060 !important;
            text-decoration: underline !important;
        }

        .status-text {
            color: #cbd5e1;
            font-size: 0.95rem;
            margin-bottom: 0.25rem;
        }

        /* Botão primário amarelo */
        .stButton > button[kind="primary"],
        .stButton > button[data-testid="stBaseButton-primary"] {
            background-color: #d4af37 !important;
            color: #ffffff !important;
            border: none !important;
            font-weight: bold !important;
            font-size: 1rem !important;
        }
        .stButton > button[kind="primary"]:hover,
        .stButton > button[data-testid="stBaseButton-primary"]:hover {
            background-color: #c5a028 !important;
            color: #ffffff !important;
        }

        /* Ícone de telefone vermelho no botão de contatos */
        [data-testid="stBaseButton-primary"] p {
            color: #ffffff !important;
        }

        /* Tag de status */
        .tag-ativo {
            display: inline-block;
            background: rgba(34, 197, 94, 0.2);
            color: #22c55e;
            border: 1px solid #22c55e;
            border-radius: 6px;
            padding: 0.15rem 0.5rem;
            font-size: 0.78rem;
            font-weight: 700;
        }
        .tag-inativo {
            display: inline-block;
            background: rgba(239, 68, 68, 0.2);
            color: #ef4444;
            border: 1px solid #ef4444;
            border-radius: 6px;
            padding: 0.15rem 0.5rem;
            font-size: 0.78rem;
            font-weight: 700;
        }
        .tag-habilitado {
            display: inline-block;
            background: rgba(59, 130, 246, 0.2);
            color: #60a5fa;
            border: 1px solid #60a5fa;
            border-radius: 6px;
            padding: 0.15rem 0.5rem;
            font-size: 0.78rem;
            font-weight: 700;
        }
    </style>
""", unsafe_allow_html=True)

# ===== SIDEBAR - Navegação customizada =====
_acanto_path = os.path.join(DATA_DIR, "acanto.png")
if os.path.exists(_acanto_path):
    with open(_acanto_path, "rb") as _f:
        _acanto_b64 = base64.b64encode(_f.read()).decode()
else:
    _acanto_b64 = None

with st.sidebar:
    if _acanto_b64:
        st.markdown(f'<div style="text-align:center;padding:1rem 0 0.5rem 0;"><img src="data:image/png;base64,{_acanto_b64}" style="max-width:70%;height:auto;"></div>', unsafe_allow_html=True)
    st.markdown("## MENU")
    st.markdown("---")
    st.page_link("streamlit_app.py", label="⚓ Cotação", icon="📊")
    st.page_link("pages/Adesões.py", label="🤝 Adesões", icon="📋")
    st.page_link("pages/Notas_Fiscais.py", label="📄 Notas Fiscais", icon="🧾")
    st.page_link("pages/Banco_de_Fornecedores.py", label="🏢 Fornecedores", icon="🔍")
    st.markdown("---")
    st.markdown('<div style="text-align:center;color:#d4af37;font-size:10px;font-weight:600;padding:0.3rem 0;white-space:nowrap;">Centro de Operações do Abastecimento</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-footer">Marinha do Brasil<br>AtaCotada v1.0</div>', unsafe_allow_html=True)

# Header
st.markdown("""
    <div class="header-container">
        <div class="logo-text">SISTEMA DE ACOMPANHAMENTO</div>
        <div class="sistema-nome">AtaCotada</div>
        <div class="subtitulo">Banco de Fornecedores</div>
    </div>
""", unsafe_allow_html=True)

st.caption("Consulta fornecedores registrados no Compras.gov.br por atividade econômica (CNAE).")
st.write("Digite o nome da atividade econômica (CNAE) para encontrar fornecedores habilitados.")


# ═══════════════════════════════════════════════════════════════════════════
# LISTA DE CNAEs — completa via API do IBGE (1.300+ subclasses)
# ═══════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=86400, show_spinner=False)   # cache por 24 h
def _fetch_cnaes_ibge() -> dict:
    """Busca TODAS as subclasses CNAE da API pública do IBGE/CONCLA."""
    url = "https://servicodados.ibge.gov.br/api/v2/cnae/subclasses"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    cnae_dict: dict[str, int] = {}
    for item in data:
        code_str = item["id"]           # ex.: "4761003"
        code_int = int(code_str)
        desc = item["descricao"].strip().title()

        # Formatar código CNAE legível: XXXX-X/XX
        if len(code_str) == 7:
            fmt = f"{code_str[:4]}-{code_str[4]}/{code_str[5:]}"
        else:
            fmt = code_str

        label = f"{desc} ({fmt})"
        cnae_dict[label] = code_int

    return cnae_dict


# Fallback mínimo caso a API do IBGE esteja fora do ar
_CNAE_FALLBACK = {
    "Comércio Varejista De Artigos De Escritório (4761-0/03)": 4761003,
    "Comércio Atacadista De Equipamentos De Informática (4651-6/01)": 4651601,
    "Comércio Atacadista De Produtos Alimentícios Em Geral (4639-7/01)": 4639701,
    "Comércio Atacadista De Medicamentos E Drogas De Uso Humano (4644-3/01)": 4644301,
    "Comércio Atacadista De Materiais De Construção Em Geral (4679-6/99)": 4679699,
    "Serviços De Engenharia (7112-0/00)": 7112000,
    "Serviços De Tecnologia Da Informação (6311-9/00)": 6311900,
    "Serviços De Limpeza Em Prédios E Domicílios (8121-4/00)": 8121400,
    "Manutenção E Reparação De Veículos Automotores (4520-0/01)": 4520001,
}

try:
    CNAE_LIST = _fetch_cnaes_ibge()
    if not CNAE_LIST:
        raise ValueError("Lista vazia")
except Exception:
    CNAE_LIST = _CNAE_FALLBACK


def format_cnpj(cnpj: str) -> str:
    """Formata CNPJ: XX.XXX.XXX/XXXX-XX"""
    cnpj = cnpj.strip()
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj


def search_fornecedores(codigo_cnae: int, apenas_ativos: bool = True) -> list:
    """Busca fornecedores na API do Compras.gov.br pelo código CNAE."""
    all_results = []
    page = 1
    max_pages = 50  # limite de segurança

    while page <= max_pages:
        params = {
            "codigoCnae": codigo_cnae,
            "ativo": str(apenas_ativos).lower(),
            "pagina": page,
            "tamanhoPagina": 100,
        }
        try:
            resp = requests.get(API_URL_FORNECEDOR, params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except Exception:
            break

        resultado = data.get("resultado", [])
        if not resultado:
            break

        all_results.extend(resultado)

        restantes = data.get("paginasRestantes", 0)
        if restantes <= 0:
            break
        page += 1

    return all_results


def fetch_contato_cnpj(cnpj_raw: str) -> dict:
    """Busca email e telefone de um CNPJ via múltiplas APIs."""
    cnpj = cnpj_raw.replace(".", "").replace("/", "").replace("-", "").strip()
    if len(cnpj) != 14:
        return {}

    email = ""
    telefone = ""

    # Tentativa 1: OpenCNPJ
    try:
        resp = requests.get(f"https://api.opencnpj.org/{cnpj}", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            # OpenCNPJ pode ter estrutura com estabelecimento ou campos na raiz
            estab = data.get("estabelecimento", data)
            email = (estab.get("email") or data.get("email") or "").strip()
            # Telefone
            ddd1 = str(estab.get("ddd1") or data.get("ddd1") or "").strip()
            tel1 = str(estab.get("telefone1") or data.get("telefone1") or "").strip()
            ddd2 = str(estab.get("ddd2") or data.get("ddd2") or "").strip()
            tel2 = str(estab.get("telefone2") or data.get("telefone2") or "").strip()
            telefones = []
            if ddd1 and tel1:
                telefones.append(f"({ddd1}) {tel1}")
            if ddd2 and tel2 and f"{ddd2}{tel2}" != f"{ddd1}{tel1}":
                telefones.append(f"({ddd2}) {tel2}")
            telefone = " / ".join(telefones)
            if email and telefone:
                return {"email": email, "telefone": telefone}
    except Exception:
        pass

    # Tentativa 2: BrasilAPI
    try:
        resp = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            email_br = (data.get("email") or "").strip()
            if email_br:
                email = email_br
            ddd1 = str(data.get("ddd_telefone_1") or "").strip()
            ddd2 = str(data.get("ddd_telefone_2") or "").strip()
            tel_parts = []
            if ddd1:
                tel_parts.append(ddd1)
            if ddd2 and ddd2 != ddd1:
                tel_parts.append(ddd2)
            if tel_parts and not telefone:
                telefone = " / ".join(tel_parts)
            if email and telefone:
                return {"email": email, "telefone": telefone}
    except Exception:
        pass

    # Tentativa 3: ReceitaWS
    try:
        resp = requests.get(f"https://receitaws.com.br/v1/cnpj/{cnpj}", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("status") != "ERROR":
                email_rws = (data.get("email") or "").strip()
                tel_rws = (data.get("telefone") or "").strip()
                if email_rws and not email:
                    email = email_rws
                if tel_rws and not telefone:
                    telefone = tel_rws
    except Exception:
        pass

    return {"email": email, "telefone": telefone}


def enrich_contatos(fornecedores: list, progress_bar, status_text) -> dict:
    """Busca contatos para uma lista de fornecedores. Retorna dict cnpj_limpo -> {email, telefone}."""
    contatos = {}
    total = len(fornecedores)
    for i, forn in enumerate(fornecedores):
        cnpj_raw = forn.get("cnpj") or ""
        cnpj_limpo = cnpj_raw.replace(".", "").replace("/", "").replace("-", "").strip()
        if not cnpj_limpo or cnpj_limpo in contatos:
            progress_bar.progress((i + 1) / total)
            status_text.text(f"Buscando contatos… {i + 1}/{total}")
            continue
        contato = fetch_contato_cnpj(cnpj_limpo)
        contatos[cnpj_limpo] = contato
        progress_bar.progress((i + 1) / total)
        status_text.text(f"Buscando contatos… {i + 1}/{total}")
        # Pequeno delay para evitar rate limit
        time.sleep(0.35)
    return contatos


# ═══════════════════════════════════════════════════════════════════════════
# INTERFACE
# ═══════════════════════════════════════════════════════════════════════════

# Selectbox com filtro por digitação
selected_cnae = st.selectbox(
    "Atividade Econômica (CNAE)",
    sorted(CNAE_LIST.keys()),
    index=None,
    placeholder="Digite para filtrar pelo nome do CNAE…",
)

apenas_ativos = True
if selected_cnae:
    apenas_ativos = st.checkbox("Buscar somente fornecedores ativos", value=True)

buscar = st.button(
    "Buscar fornecedores",
    type="primary",
    use_container_width=True,
    disabled=not selected_cnae,
)

if buscar and selected_cnae:
    codigo_cnae = CNAE_LIST[selected_cnae]
    with st.spinner("Consultando fornecedores, aguarde…"):
        resultados = search_fornecedores(codigo_cnae, apenas_ativos)
    st.session_state["fornecedores"] = resultados
    st.session_state["cnae_pesquisado"] = selected_cnae
    st.session_state["contatos_fornecedores"] = {}  # limpar contatos da busca anterior

# ── Exibição de resultados ─────────────────────────────────────────────────
resultados = st.session_state.get("fornecedores", [])
cnae_pesquisado = st.session_state.get("cnae_pesquisado", "")

if resultados:
    st.markdown(
        f"""
        <div class="metric-card" style="text-align:center;margin-bottom:1.2rem;">
            <div style="color:#d4af37;font-size:2rem;font-weight:bold;">{len(resultados)}</div>
            <div style="color:#cbd5e1;font-size:0.95rem;">fornecedor{"es" if len(resultados) != 1 else ""} encontrado{"s" if len(resultados) != 1 else ""}</div>
            <div style="color:#94a3b8;font-size:0.8rem;margin-top:0.3rem;">{cnae_pesquisado}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Filtro por texto dentro dos resultados
    filtro_texto = st.text_input(
        "🔍 Filtrar resultados por nome do fornecedor, município ou UF",
        placeholder="Ex: São Paulo, ME, equipamentos…",
    )

    display = resultados
    if filtro_texto:
        filtro_lower = filtro_texto.lower()
        display = [
            r for r in resultados
            if filtro_lower in (r.get("nomeRazaoSocialFornecedor") or "").lower()
            or filtro_lower in (r.get("nomeMunicipio") or "").lower()
            or filtro_lower in (r.get("ufSigla") or r.get("uf") or "").lower()
            or filtro_lower in (r.get("cnpj") or "").lower()
            or filtro_lower in (r.get("porteEmpresaNome") or "").lower()
            or filtro_lower in (r.get("naturezaJuridicaNome") or "").lower()
        ]

    if filtro_texto:
        st.markdown(
            f"""
            <div class="metric-card" style="text-align:center;margin-bottom:1rem;padding:0.6rem 1rem;">
                <div style="color:#d4af37;font-size:1.6rem;font-weight:bold;">{len(display)}</div>
                <div style="color:#cbd5e1;font-size:0.85rem;">fornecedor{"es" if len(display) != 1 else ""} após filtro</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # Botão para buscar contatos via CNPJ
    contatos = st.session_state.get("contatos_fornecedores", {})
    if not contatos:
        st.markdown(
            '<style>'
            '#btn_contatos span::before { content: "\260e\fe0f "; color: #e53e3e; font-size: 1.2rem; }'
            '</style>',
            unsafe_allow_html=True,
        )
        if st.button("\U0000260e\ufe0f Buscar e-mail e telefone dos fornecedores", use_container_width=True, type="primary", key="btn_contatos"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            contatos = enrich_contatos(display, progress_bar, status_text)
            st.session_state["contatos_fornecedores"] = contatos
            progress_bar.empty()
            status_text.success(f"✅ Contatos carregados para {len(contatos)} fornecedor(es).")
    else:
        st.caption(f"✅ Contatos carregados para {len(contatos)} fornecedor(es).")
        if st.button("🔄 Recarregar contatos", use_container_width=False):
            st.session_state["contatos_fornecedores"] = {}
            st.rerun()

    if not display:
        st.info("Nenhum fornecedor corresponde ao filtro digitado.")
    else:
        for forn in display:
            nome = forn.get("nomeRazaoSocialFornecedor", "Nome não informado")
            cnpj = forn.get("cnpj") or forn.get("cpf") or "N/I"
            if cnpj != "N/I" and len(cnpj.replace(".", "").replace("/", "").replace("-", "")) == 14:
                cnpj_fmt = format_cnpj(cnpj.replace(".", "").replace("/", "").replace("-", ""))
            else:
                cnpj_fmt = cnpj

            municipio = forn.get("nomeMunicipio", "N/I")
            uf = forn.get("ufSigla") or forn.get("uf") or "N/I"
            porte = forn.get("porteEmpresaNome", "N/I")
            natureza = forn.get("naturezaJuridicaNome", "N/I")
            cnae_nome = forn.get("nomeCnae", "")

            ativo = forn.get("ativo")
            if ativo is True:
                tag_ativo = '<span class="tag-ativo">Ativo</span>'
            elif ativo is False:
                tag_ativo = '<span class="tag-inativo">Inativo</span>'
            else:
                tag_ativo = ""

            habilitado = forn.get("habilitadoLicitar")
            if habilitado is True:
                tag_hab = '<span class="tag-habilitado">Habilitado a licitar</span>'
            elif habilitado is False:
                tag_hab = '<span class="tag-inativo">Não habilitado</span>'
            else:
                tag_hab = ""

            # Contato (buscado via BrasilAPI / publica.cnpj.ws)
            cnpj_limpo = (cnpj or "").replace(".", "").replace("/", "").replace("-", "").strip()
            contato_info = contatos.get(cnpj_limpo, {})
            email = contato_info.get("email", "")
            telefone = contato_info.get("telefone", "")

            contato_email = f'<span>📧 <b>E-mail:</b> {email}</span>' if email else ""
            contato_tel = f'<span>📞 <b>Telefone:</b> {telefone}</span>' if telefone else ""

            st.markdown(
                '<div class="result-card">'
                '<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:0.3rem;">'
                f'<div class="status-text" style="font-size:1rem;font-weight:600;color:#ffffff;">{nome}</div>'
                f'<div style="display:flex;gap:0.4rem;">{tag_ativo} {tag_hab}</div>'
                '</div>'
                '<div style="display:flex;flex-wrap:wrap;gap:0.5rem 1.4rem;margin-top:0.45rem;font-size:0.85rem;color:#cbd5e1;">'
                f'<span>📄 <b>CNPJ:</b> {cnpj_fmt}</span>'
                f'<span>📍 <b>Município:</b> {municipio} – {uf}</span>'
                f'<span>🏢 <b>Porte:</b> {porte}</span>'
                f'<span>⚖️ <b>Natureza Jurídica:</b> {natureza}</span>'
                f'{contato_email}'
                f'{contato_tel}'
                '</div>'
                '</div>',
                unsafe_allow_html=True,
            )

    # ── Exportação ──────────────────────────────────────────────────────
    if display:
        rows = [
            {
                "Razão Social": f.get("nomeRazaoSocialFornecedor", ""),
                "CNPJ": f.get("cnpj", ""),
                "Município": f.get("nomeMunicipio", ""),
                "UF": f.get("ufSigla") or f.get("uf") or "",
                "Porte": f.get("porteEmpresaNome", ""),
                "Natureza Jurídica": f.get("naturezaJuridicaNome", ""),
                "E-mail": contatos.get((f.get("cnpj") or "").replace(".", "").replace("/", "").replace("-", "").strip(), {}).get("email", ""),
                "Telefone": contatos.get((f.get("cnpj") or "").replace(".", "").replace("/", "").replace("-", "").strip(), {}).get("telefone", ""),
                "Ativo": "Sim" if f.get("ativo") is True else ("Não" if f.get("ativo") is False else ""),
                "Habilitado a Licitar": "Sim" if f.get("habilitadoLicitar") is True else ("Não" if f.get("habilitadoLicitar") is False else ""),
            }
            for f in display
        ]
        df = pd.DataFrame(rows)

        # ── Helper: gerar Excel formatado ──────────────────────────────────
        def build_excel(df: pd.DataFrame, cnae_nome: str) -> bytes:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fornecedores"

            # Cores
            azul_escuro = "001A4D"
            dourado = "D4AF37"
            branco = "FFFFFF"
            cinza_claro = "F2F2F2"
            borda = Side(style="thin", color="AAAAAA")

            n_cols = len(df.columns)

            # ── Linha 1: Titulo
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
            titulo_cell = ws.cell(row=1, column=1, value="FORNECEDORES ATIVOS NO COMPRASNET")
            titulo_cell.font = Font(name="Arial", size=14, bold=True, color=branco)
            titulo_cell.fill = PatternFill("solid", fgColor=azul_escuro)
            titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35

            # ── Linha 2: CNAE
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
            cnae_cell = ws.cell(row=2, column=1, value=f"CNAE: {cnae_nome}")
            cnae_cell.font = Font(name="Arial", size=11, bold=True, color=azul_escuro)
            cnae_cell.fill = PatternFill("solid", fgColor=dourado)
            cnae_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 25

            # ── Linha 3: Data de geração
            from datetime import datetime
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)
            data_cell = ws.cell(row=3, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(df)} fornecedor(es)")
            data_cell.font = Font(name="Arial", size=9, italic=True, color="555555")
            data_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[3].height = 20

            # ── Linha 4: vazia (espaçamento)
            ws.row_dimensions[4].height = 8

            # ── Linha 5: Cabeçalho da tabela
            header_row = 5
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                cell.font = Font(name="Arial", size=10, bold=True, color=branco)
                cell.fill = PatternFill("solid", fgColor=azul_escuro)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(top=borda, bottom=borda, left=borda, right=borda)
            ws.row_dimensions[header_row].height = 28

            # ── Dados
            for row_idx, row_data in enumerate(df.itertuples(index=False), header_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = Border(top=borda, bottom=borda, left=borda, right=borda)
                    if (row_idx - header_row) % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=cinza_claro)

            # ── Ajustar largura das colunas
            for col_idx in range(1, n_cols + 1):
                max_len = len(str(df.columns[col_idx - 1]))
                for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

            # ── Congelar painel no cabeçalho
            ws.freeze_panes = f"A{header_row + 1}"

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # ── Helper: gerar PDF ──────────────────────────────────────────────
        def build_pdf(df: pd.DataFrame, cnae_nome: str) -> bytes:
            from datetime import datetime

            class PDFReport(FPDF):
                def header(self):
                    self.set_fill_color(0, 26, 77)  # azul escuro
                    self.rect(0, 0, 300, 22, "F")
                    self.set_font("Helvetica", "B", 14)
                    self.set_text_color(255, 255, 255)
                    self.cell(0, 10, "FORNECEDORES ATIVOS NO COMPRASNET", align="C", new_x="LMARGIN", new_y="NEXT")
                    self.set_font("Helvetica", "B", 10)
                    self.set_text_color(212, 175, 55)  # dourado
                    self.cell(0, 8, f"CNAE: {cnae_nome}", align="C", new_x="LMARGIN", new_y="NEXT")
                    self.set_font("Helvetica", "I", 8)
                    self.set_text_color(200, 200, 200)
                    self.cell(0, 6, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total: {len(df)} fornecedor(es)", align="C", new_x="LMARGIN", new_y="NEXT")
                    self.ln(3)

                def footer(self):
                    self.set_y(-15)
                    self.set_font("Helvetica", "I", 7)
                    self.set_text_color(150, 150, 150)
                    self.cell(0, 10, f"AtaCotada  -  Pag. {self.page_no()}/{{nb}}", align="C")

            pdf = PDFReport(orientation="L", unit="mm", format="A4")
            pdf.alias_nb_pages()
            pdf.set_auto_page_break(auto=True, margin=18)
            pdf.add_page()

            # Colunas selecionadas para o PDF (cabe na página)
            pdf_cols = ["Razão Social", "CNPJ", "Município", "UF", "Porte", "E-mail", "Telefone", "Ativo"]
            pdf_cols = [c for c in pdf_cols if c in df.columns]
            col_widths = {
                "Razão Social": 70, "CNPJ": 38, "Município": 35, "UF": 12,
                "Porte": 30, "E-mail": 50, "Telefone": 30, "Ativo": 12,
            }

            # Cabeçalho da tabela
            pdf.set_fill_color(0, 26, 77)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 7)
            for col in pdf_cols:
                w = col_widths.get(col, 30)
                pdf.cell(w, 7, col, border=1, fill=True, align="C")
            pdf.ln()

            # Dados
            pdf.set_font("Helvetica", "", 6.5)
            fill = False
            for _, row in df.iterrows():
                if fill:
                    pdf.set_fill_color(240, 240, 240)
                else:
                    pdf.set_fill_color(255, 255, 255)
                pdf.set_text_color(30, 30, 30)
                for col in pdf_cols:
                    w = col_widths.get(col, 30)
                    txt = str(row.get(col, ""))[:60]  # truncar
                    pdf.cell(w, 6, txt, border=1, fill=True, align="L")
                pdf.ln()
                fill = not fill

            return bytes(pdf.output())

        # ── Botões de exportação ───────────────────────────────────────────
        col_xl, col_pdf = st.columns(2)
        with col_xl:
            excel_data = build_excel(df, cnae_pesquisado)
            st.download_button(
                label="📥 Exportar planilha Excel",
                data=excel_data,
                file_name=f"fornecedores_{cnae_pesquisado[:30].replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_pdf:
            pdf_data = build_pdf(df, cnae_pesquisado)
            st.download_button(
                label="📄 Exportar relatório PDF",
                data=pdf_data,
                file_name=f"fornecedores_{cnae_pesquisado[:30].replace(' ', '_')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

elif buscar and selected_cnae:
    st.info("Nenhum fornecedor encontrado para este CNAE.")
